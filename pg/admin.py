from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from accounts.models import UserAccount
from colleges.models import College
from import_export import resources, fields
from import_export.widgets import ForeignKeyWidget
from import_export.admin import ImportExportModelAdmin
from .models import (
    PGFaculty, PGDepartment, PGDegree, PGProgram, PGBatch, PGStudentProfile,
    PGCourseStructure, PGStudentCourseAssessment, PGSemesterRegistration, PGExamRegistration,
    PGCommonCourseStructure, PGExamResult,
    PGExam, PGExamCenterMapping, PGGroup, PGExamSchedule,
    PGExamRegistrationPayment,
)


class SafeForeignKeyWidget(ForeignKeyWidget):
    """
    Custom widget that uses filter().first() instead of get() to avoid
    MultipleObjectsReturned error when duplicate related objects exist.
    """
    def clean(self, value, row=None, *args, **kwargs):
        val = super(ForeignKeyWidget, self).clean(value, row=row, *args, **kwargs)
        if val:
            return self.model.objects.filter(**{self.field: val}).first()
        return None


@admin.register(PGFaculty)
class PGFacultyAdmin(admin.ModelAdmin):
    list_display = ('name', 'short_name', 'university', 'created_at')
    list_filter = ('university',)
    search_fields = ('name', 'short_name')
    ordering = ('name',)


@admin.register(PGDepartment)
class PGDepartmentAdmin(admin.ModelAdmin):
    list_display = ('uid', 'name', 'code', 'faculty', 'head_of_department', 'created_at')
    list_filter = ('faculty',)
    search_fields = ('uid', 'name', 'code', 'faculty__name')
    readonly_fields = ('uid', 'created_at', 'updated_at')
    ordering = ('faculty', 'name')


@admin.register(PGDegree)
class PGDegreeAdmin(admin.ModelAdmin):
    list_display = ('name', 'short_name', 'total_semesters', 'total_years', 'created_at')
    search_fields = ('name', 'short_name')
    ordering = ('name',)


@admin.register(PGProgram)
class PGProgramAdmin(admin.ModelAdmin):
    list_display = ('name', 'short_name', 'degree', 'department', 'created_at')
    list_filter = ('degree', 'department__faculty')
    search_fields = ('name', 'short_name', 'degree__name', 'department__name')
    ordering = ('name',)


@admin.register(PGBatch)
class PGBatchAdmin(admin.ModelAdmin):
    list_display = ('name', 'program', 'created_at')
    list_filter = ('program',)
    search_fields = ('name', 'program__name')
    ordering = ('name',)


class PGStudentProfileResource(resources.ModelResource):
    """
    Resource for importing/exporting PGStudentProfile records.

    FK columns in the import file:
      - user       → UserAccount.username
      - college    → College.college_code
      - department → PGDepartment.name
      - program    → PGProgram.name
      - degree     → PGDegree.name
    """

    user = fields.Field(
        column_name='user',
        attribute='user',
        widget=ForeignKeyWidget(UserAccount, field='username')
    )
    college = fields.Field(
        column_name='college',
        attribute='college',
        widget=SafeForeignKeyWidget(College, field='college_code')
    )
    department = fields.Field(
        column_name='department',
        attribute='department',
        widget=SafeForeignKeyWidget(PGDepartment, field='name')
    )
    program = fields.Field(
        column_name='program',
        attribute='program',
        widget=SafeForeignKeyWidget(PGProgram, field='name')
    )
    degree = fields.Field(
        column_name='degree',
        attribute='degree',
        widget=SafeForeignKeyWidget(PGDegree, field='name')
    )

    class Meta:
        model = PGStudentProfile
        exclude = ('uid', 'profile_image', 'signature')
        import_id_fields = ('registration_no',)
        export_order = (
            'id', 'registration_no', 'roll_no', 'first_name', 'last_name',
            'hindi_name', 'user', 'college', 'department', 'program', 'degree',
            'gender', 'date_of_birth', 'mobile_no', 'aadhar_no', 'apaar_id',
            'address', 'father_name', 'mother_name', 'religion', 'nationality',
            'medium_of_student', 'caste', 'current_semester', 'session', 'batch',
            'status', 'is_active', 'admission_date', 'enrollment_date',
            'migration_submitted', 'last_university',
            'cc_course', 'sec_course', 'ec_course',
            'created_at', 'updated_at',
        )


@admin.register(PGStudentProfile)
class PGStudentProfileAdmin(ImportExportModelAdmin):
    resource_class = PGStudentProfileResource
    list_display = ('id', 'registration_no', 'first_name', 'last_name', 'hindi_name', 'roll_no', 'college', 
                   'department', 'program', 'current_semester', 'status', 'is_active', 'batch')
    list_filter = ('status', 'gender', 'religion', 'nationality', 'medium_of_student', 'college', 'department', 'program', 'degree', 
                  'current_semester', 'batch')
    search_fields = ('registration_no', 'user__username', 'roll_no', 'first_name', 'last_name', 
                    'mobile_no', 'aadhar_no')
    readonly_fields = ('uid', 'created_at', 'updated_at')
    ordering = ('-created_at',)
    
    # Performance optimizations
    list_select_related = ('user', 'college', 'department', 'program', 'degree')
    autocomplete_fields = ['college', 'department', 'program', 'degree']
    raw_id_fields = ('user',)
    list_per_page = 50  # Limit records per page for better performance
    
    def get_queryset(self, request):
        """Optimize queryset with select_related for foreign keys"""
        qs = super().get_queryset(request)
        return qs.select_related(
            'user',
            'college', 
            'department',
            'program',
            'degree'
        )
    
    fieldsets = (
        ('Personal Information', {
            'fields': ('uid', 'user', 'first_name', 'last_name', 'hindi_name',
                      'date_of_birth', 'gender', 'caste', 'religion', 'nationality', 'medium_of_student', 'mobile_no', 'aadhar_no', 'address')
        }),
        ('Academic Information', {
            'fields': ('registration_no', 'roll_no', 'college', 'department', 
                      'program', 'degree', 'current_semester', 'session', 'batch', 'status', 'is_active')
        }),
        ('Admission Details', {
            'fields': ('admission_date', 'enrollment_date', 'migration_submitted', 'last_university')
        }),
        ('Course Selections (PG)', {
            'fields': ('cc_course', 'sec_course', 'ec_course')
        }),
        ('Family Information', {
            'fields': ('father_name', 'mother_name')
        }),
        ('Documents', {
            'fields': ('profile_image', 'signature'),
            'classes': ('collapse',)
        }),
        ('Additional Data', {
            'fields': ('json_data',),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )


@admin.register(PGCourseStructure)
class PGCourseStructureAdmin(admin.ModelAdmin):
    list_display = ('course_name', 'department', 'course_type', 'code', 'paper_code', 'semester', 'max_marks', 'max_credit', 'effective_credit', 'label')
    list_filter = ('department__faculty', 'department', 'course_type', 'semester','code',)
    search_fields = ('course_name', 'course_short_name', 'code', 'paper_code', 'department__name', 'label')
    ordering = ('department', 'semester', 'code')
    readonly_fields = ('uid', 'created_at', 'updated_at')
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('uid', 'course_name', 'course_short_name', 'department', 'course_type', 'code', 'paper_code', 'semester')
        }),
        ('Credits & Marks', {
            'fields': ('max_credit', 'effective_credit', 'max_marks', 'min_marks')
        }),
        ('Assessment Details', {
            'fields': ('label', 'description')
        }),
        ('Additional Data', {
            'fields': ('json_data',),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )






class PGStudentCourseAssessmentResource(resources.ModelResource):
    student = fields.Field(
        column_name='student',
        attribute='student',
        widget=SafeForeignKeyWidget(PGStudentProfile, field='registration_no')
    )
    batch = fields.Field(
        column_name='batch',
        attribute='batch',
        widget=SafeForeignKeyWidget(PGBatch, field='name')
    )
    department = fields.Field(
        column_name='department',
        attribute='department',
        widget=SafeForeignKeyWidget(PGDepartment, field='name')
    )
    student_department = fields.Field(
        column_name='student_department',
        attribute='student__department__name',
        readonly=True
    )

    class Meta:
        model = PGStudentCourseAssessment
        exclude = ('uid', 'json_data')
        export_order = (
            'id', 'student', 'student_department', 'course_name', 'course_code',
            'paper_code', 'semester', 'label', 'department', 'session',
            'batch', 'college_code', 'exam_type', 'ind_marks_obtained',
            'ind_max_marks', 'ind_pass_marks', 'ind_is_pass', 'ind_is_absent',
            'sem_result', 'is_ese_fill'
        )

    def get_queryset(self):
        """Optimize queryset for export with select_related"""
        return super().get_queryset().select_related('student', 'department', 'batch')

    def skip_row(self, instance, original, row, import_validation_errors=None):
        """Skip completely blank rows (Excel files often have trailing empty rows)."""
        student_val = row.get('student')
        if not student_val or str(student_val).strip() in ('', 'None', 'nan'):
            return True  # Skip — student is blank
        return super().skip_row(instance, original, row, import_validation_errors)

    def before_import_row(self, row, **kwargs):
        """
        Normalize the 'student' column before import.

        Handles two cases:
          1. student column is a numeric id (int or float like 1048.0  from Excel)
             → look up the PGStudentProfile by pk and replace with registration_no
          2. student column is a registration_no string
             → auto-create PGStudentProfile if UserAccount exists but profile doesn't
        """
        student_val = row.get('student')
        print(f"[DEBUG before_import_row] student raw value: {repr(student_val)} | type: {type(student_val).__name__}")
        if not student_val:
            return

        # Normalize: Excel often stores numbers as floats (e.g. 1048.0)
        student_str = str(student_val).strip()
        if student_str.endswith('.0'):
            student_str = student_str[:-2]   # '1048.0' → '1048'

        # Case 1: numeric id → look up registration_no and substitute
        if student_str.isdigit():
            profile = PGStudentProfile.objects.filter(pk=int(student_str)).first()
            if profile:
                row['student'] = profile.registration_no
            # else: leave as-is; import will fail with a clear FK error
            return

        # Case 2: registration_no string
        registration_no = student_str
        row['student'] = registration_no  # ensure cleaned value is used

        if not PGStudentProfile.objects.filter(registration_no=registration_no).exists():
            from accounts.models import UserAccount
            from colleges.models import College

            user = UserAccount.objects.filter(username=registration_no).first()
            if user:
                batch_name   = row.get('batch')
                dept_name    = row.get('department')
                college_code = row.get('college_code')

                dept_obj    = PGDepartment.objects.filter(name=dept_name).first() if dept_name else None
                college_obj = College.objects.filter(college_code=college_code).first() if college_code else None

                PGStudentProfile.objects.create(
                    user=user,
                    registration_no=registration_no,
                    first_name=user.get_full_name(),
                    last_name='',
                    batch=batch_name,
                    department=dept_obj,
                    college=college_obj,
                    status='Active'
                )
                print(f"[Import] Auto-created PGStudentProfile for: {registration_no}")

@admin.register(PGStudentCourseAssessment)
class PGStudentCourseAssessmentAdmin(ImportExportModelAdmin):
    resource_class = PGStudentCourseAssessmentResource

    # ── List view: only essential columns ─────────────────────────────────
    list_display = (
        'get_regno', 'get_student_name', 'department', 'paper_code', 'semester',
        'label', 'exam_type', 'session',
        'ind_marks_obtained', 'ind_max_marks', 'ind_is_absent', 'ind_is_pass',
        'sem_result', 'is_ese_fill',
    )

    # Prevent N+1: join student in the same query
    list_select_related = ('student', 'department', 'batch')

    # Raw-id widgets → avoids loading full dropdown lists
    raw_id_fields = ('student', 'department', 'batch')

    # No COUNT(*) on 400K rows
    show_full_result_count = False

    # Only FK / indexed / low-cardinality fields as filters
    list_filter = (
        'semester', 'batch__name','session', 'exam_type', 'label',
        'ind_is_absent', 'sem_result',
        'is_cia_fill', 'is_ese_fill',
        'department',
        'college_code',
        'course_code'
    )

    # Search on indexed fields only
    search_fields = (
        'student__registration_no',
        'student__roll_no',
        'paper_code',
        'college_code',
    )

    list_per_page = 50

    # No default ordering → avoids full-table sort on 400K rows
    # Use search / filters to narrow first, then sort
    ordering = []

    actions = ['export_assessments_to_excel']

    def export_assessments_to_excel(self, request, queryset):
        """
        Fast Excel export for PG Student Course Assessments.
        Uses openpyxl directly for better performance on large datasets.
        Respects current filters and selection.
        """
        from openpyxl import Workbook
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Assessments"
        
        # Define headers
        headers = [
            'Registration No', 'Roll No', 'Name', 'Department', 'College', 
            'Batch', 'Semester', 'Session', 'Paper Code', 'Label', 
            'Exam Type', 'Max Marks', 'Pass Marks', 'Marks Obtained', 
            'Is Absent', 'Is Pass', 'Result', 'ESE Fill'
        ]
        
        # Style header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, head in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=head)
            cell.fill = header_fill
            cell.font = header_font
        
        # Optimize queryset
        queryset = queryset.select_related('student', 'student__college', 'department', 'batch').only(
            'student__registration_no', 'student__roll_no', 'student__first_name', 'student__last_name',
            'student__college__college_code', 'department__name', 'batch__name',
            'semester', 'session', 'paper_code', 'label', 'exam_type',
            'ind_max_marks', 'ind_pass_marks', 'ind_marks_obtained',
            'ind_is_absent', 'ind_is_pass', 'sem_result', 'is_ese_fill'
        )
        
        # Write data
        for row_idx, obj in enumerate(queryset.iterator(), 2):
            student_name = f"{obj.student.first_name} {obj.student.last_name or ''}".strip() if obj.student else "-"
            college = obj.student.college.college_code if obj.student and obj.student.college else obj.college_code or "-"
            
            row = [
                obj.student.registration_no if obj.student else "-",
                obj.student.roll_no if obj.student else "-",
                student_name,
                obj.department.name if obj.department else (obj.student.department.name if obj.student and obj.student.department else "-"),
                college,
                obj.batch.name if obj.batch else "-",
                obj.semester,
                obj.session,
                obj.paper_code,
                obj.label,
                obj.exam_type,
                obj.ind_max_marks,
                float(obj.ind_pass_marks) if obj.ind_pass_marks else 0,
                float(obj.ind_marks_obtained) if obj.ind_marks_obtained is not None else 0,
                "YES" if obj.ind_is_absent else "NO",
                "PASS" if obj.ind_is_pass else "FAIL" if obj.ind_is_pass is False else "PENDING",
                obj.sem_result,
                "YES" if obj.is_ese_fill else "NO"
            ]
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="pg_assessments_export.xlsx"'
        wb.save(response)
        return response

    export_assessments_to_excel.short_description = "🚀 Fast Excel Export (Optimized)"

    readonly_fields = (
        'uid', 'created_at', 'updated_at',
        'comb_max_marks', 'comb_final_marks_obtained', 'comb_grade_point',
        'course_max_marks', 'course_final_marks_obtained', 'course_grade_point',
        'sem_max_credit', 'sgpa', 'sem_result',
    )

    fieldsets = (
        ('Student & Course Info', {
            'fields': ('uid', 'student', 'course_name', 'course_short_name',
                       'course_type', 'course_code', 'paper_code', 'semester', 'label')
        }),
        ('Exam Details', {
            'fields': ('exam_type', 'session', 'batch', 'department',
                       'degree', 'college_code', 'attendance', 'is_cia_fill', 'is_ese_fill')
        }),
        ('Individual Assessment', {
            'fields': ('ind_max_marks', 'ind_pass_marks', 'ind_marks_obtained',
                       'ind_grace_obtained', 'ind_final_marks_obtained', 'ind_is_pass', 'ind_is_absent')
        }),
        ('Combined Assessment', {
            'fields': ('comb_max_marks', 'comb_max_credits', 'comb_pass_marks', 'comb_marks_obtained',
                       'comb_grace_obtained', 'comb_final_marks_obtained', 'comb_credit_obtained',
                       'comb_numeric_grade', 'comb_letter_grade', 'comb_grade_point'),
            'classes': ('collapse',)
        }),
        ('Course Assessment', {
            'fields': ('course_max_marks', 'course_max_credits', 'course_pass_marks', 'course_marks_obtained',
                       'course_grace_obtained', 'course_final_marks_obtained',
                       'course_credit_obtained', 'course_grade_point'),
            'classes': ('collapse',)
        }),
        ('Semester Assessment', {
            'fields': ('sem_max_credit', 'sem_credit_obtained', 'sgpa',
                       'sem_result', 'next_sem_status', 'sem_grace_obtained')
        }),
        ('Temp / JSON', {
            'fields': ('temp_total_gp', 'json_data'),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

    # ── Queryset: select_related to prevent N+1 on list page ──────────────
    def get_queryset(self, request):
        return super().get_queryset(request).select_related(
            'student', 'department', 'batch'
        )

    # ── Compact display helpers ────────────────────────────────────────────
    def get_regno(self, obj):
        return obj.student.registration_no if obj.student else '-'
    get_regno.short_description = 'Reg No'
    get_regno.admin_order_field = 'student__registration_no'

    def get_student_name(self, obj):
        if obj.student:
            return f"{obj.student.first_name or ''} {obj.student.last_name or ''}".strip()
        return '-'
    get_student_name.short_description = 'Name'
    get_student_name.admin_order_field = 'student__first_name'


class PGSemesterRegistrationResource(resources.ModelResource):
    student = fields.Field(
        column_name='student',
        attribute='student',
        widget=ForeignKeyWidget(PGStudentProfile, field='registration_no')
    )

    class Meta:
        model = PGSemesterRegistration
        exclude = ('uid',)
        import_id_fields = ('student', 'sem', 'session')


@admin.register(PGSemesterRegistration)
class PGSemesterRegistrationAdmin(ImportExportModelAdmin):
    resource_class = PGSemesterRegistrationResource
    list_display = ('student', 'sem', 'session', 'status', 'is_open', 'exam_eligible', 'start_date', 'end_date')
    list_filter = ('sem', 'is_open', 'status', 'exam_eligible', 'session')
    search_fields = (
        'student__registration_no',
        'student__roll_no',
        'student__first_name',
        'student__last_name',
        'student__aadhar_no',
        'student__mobile_no',
        'session',
        'remarks',
    )
    raw_id_fields = ('student',)
    ordering = ('-created_at',)
    readonly_fields = ('uid', 'created_at', 'updated_at')
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('uid', 'student', 'sem', 'session')
        }),
        ('Registration Period', {
            'fields': ('start_date', 'end_date', 'is_open', 'status')
        }),
        ('Eligibility & Remarks', {
            'fields': ('exam_eligible', 'remarks')
        }),
        ('Additional Data', {
            'fields': ('json_data',),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )


class PGExamRegistrationResource(resources.ModelResource):
    student = fields.Field(
        column_name='student',
        attribute='student',
        widget=ForeignKeyWidget(PGStudentProfile, field='registration_no')
    )

    class Meta:
        model = PGExamRegistration
        exclude = ('uid', 'id')
        import_id_fields = ('student', 'sem', 'session', 'exam_type')

    def before_import_row(self, row, **kwargs):
        """
        Hook called before importing each row.
        Used here to create PGStudentProfile if it doesn't exist but UserAccount does.
        """
        registration_no = row.get('student')
        if not registration_no:
            return

        # Check if Profile exists
        if not PGStudentProfile.objects.filter(registration_no=registration_no).exists():
            from accounts.models import UserAccount
            from colleges.models import College
            
            # Check if UserAccount exists
            user = UserAccount.objects.filter(username=registration_no).first()
            if user:
                # print(f"Creating missing profile for User: {registration_no}")
                
                # Resolving Foreign Keys from row data if available (though these are on registration, not profile usually)
                # For Profile creation, we try to use defaults or data from account
                
                # Create the Profile
                PGStudentProfile.objects.create(
                    user=user,
                    registration_no=registration_no,
                    first_name=user.get_full_name(), 
                    last_name="", 
                    status='Active'
                )

@admin.register(PGExamRegistration)
class PGExamRegistrationAdmin(ImportExportModelAdmin):
    resource_class = PGExamRegistrationResource
    list_display = ('student', 'exam', 'sem', 'session', 'status', 'is_open', 'exam_type', 'fees', 'start_date', 'end_date',)
    list_filter = ('sem', 'is_open', 'status', 'session', 'exam_type', 'exam', 'student__department', 'created_at')
    search_fields = (
        'student__registration_no',
        'student__roll_no',
        'student__first_name',
        'student__last_name',
        'student__aadhar_no',
        'student__mobile_no',
        'session',
    )
    raw_id_fields = ('student', 'exam')
    ordering = ('-created_at',)
    readonly_fields = ('uid', 'created_at', 'updated_at')
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('uid', 'student', 'exam', 'sem', 'session', 'exam_type')
        }),
        ('Registration Period', {
            'fields': ('start_date', 'end_date', 'is_open', 'status')
        }),
        ('Fees', {
            'fields': ('fees',)
        }),
        ('Documents', {
            'fields': ('admission_receipt',),
            'classes': ('collapse',)
        }),
        ('Additional Data', {
            'fields': ('json_data',),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )


@admin.register(PGCommonCourseStructure)
class CommonCourseStructureAdmin(admin.ModelAdmin):
    list_display = ('semester', 'course_code', 'course_type', 'course_name', 'get_departments_count', 'credit', 'marks', 'cia_marks', 'ese_marks')
    list_filter = ('semester', 'credit', 'course_type', 'departments', 'course_code')
    search_fields = ('course_name', 'course_type', 'course_code', 'old_code', 'new_code')
    ordering = ('semester', 'course_code')
    readonly_fields = ('uid', 'created_at', 'updated_at', 'get_departments_list')
    filter_horizontal = ('departments',)  # Nice interface for ManyToMany
    list_per_page = 50
    
    def get_departments_count(self, obj):
        """Display count of departments offering this course"""
        return obj.departments.count()
    get_departments_count.short_description = 'Dept Count'
    
    def get_departments_list(self, obj):
        """Display list of all departments"""
        depts = obj.departments.all()
        if depts.exists():
            return ', '.join([d.name for d in depts])
        return 'No departments'
    get_departments_list.short_description = 'Departments Offering This Course'
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('uid', 'semester', 'course_code', 'course_type', 'course_name')
        }),
        ('Departments', {
            'fields': ('departments', 'get_departments_list'),
            'description': 'Select all departments that offer this common course'
        }),
        ('Credits & Marks', {
            'fields': ('credit', 'marks', 'cia_marks', 'ese_marks')
        }),
        ('Course Codes', {
            'fields': ('old_code', 'new_code')
        }),
        ('Additional Data', {
            'fields': ('json_data',),
            'classes': ('collapse',)
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )

class PGExamResultResource(resources.ModelResource):
    student = fields.Field(
        column_name='student',
        attribute='student',
        widget=ForeignKeyWidget(PGStudentProfile, field='registration_no')
    )
    department = fields.Field(
        column_name='department',
        attribute='student__department__name',
        readonly=True
    )

    class Meta:
        model = PGExamResult
        exclude = ('uid',)
        import_id_fields = ('student', 'semester', 'session')
        export_order = (
            'id', 'student', 'department', 'semester', 'session',
            'cia_pass', 'ese_pass', 'semester_result', 'sgpa',
            'semester_max_credit', 'semester_credit_earned',
            'next_semester', 'next_sem_status', 'is_legacy',
            'published_at', 'created_at', 'updated_at'
        )


@admin.register(PGExamResult)
class PGExamResultAdmin(ImportExportModelAdmin):
    resource_class = PGExamResultResource
    list_display = (
        'get_student_regno',
        'get_student_name',
        'get_student_dept',
        'semester',
        'session',
        'cia_pass_display',
        'ese_pass_display',
        'semester_result',
        'sgpa',
        'semester_credit_earned',
        'next_sem_status',
        'created_at'
    )
    
    # Optimization: Reduce database queries
    list_select_related = ('student', 'student__department', 'student__program', 'student__college')
    
    list_filter = (
        'semester',
        'session',
         'student__batch',
        'cia_pass',
        'ese_pass',
        'semester_result',
        'next_sem_status',
        'is_legacy',
        'student__department',
    )
    
    # Optimized search fields
    search_fields = (
        'student__registration_no',
        'student__first_name',
        'student__last_name',
        'student__roll_no',
        'student__department__name',
        'uid'
    )
    
    readonly_fields = (
        'uid',
        # 'get_student_full_info',
        'get_cia_courses',
        'get_ese_courses',
        'created_at',
        'updated_at'
    )
    
    ordering = ('-created_at',)
    raw_id_fields = ('student',)
    list_per_page = 50
    
    # Add custom actions
    actions = ['export_pass_promoted_to_excel']
    
    def export_pass_promoted_to_excel(self, request, queryset):
        """
        Export PGExamResult records to Excel, filtering only PASS and PROMOTED students
        from Semester 2, Session 2024-25
        """
        # Ignore the selected queryset and fetch ALL matching records from database
        # This ensures we export all students, not just those on the current page
        filtered_queryset = PGExamResult.objects.filter(
            semester_result__in=['PASS', 'PROMOTED'],
            semester='2ND',
            session='2024-25'
        )
        
        if not filtered_queryset.exists():
            self.message_user(request, "No PASS or PROMOTED records found for Semester 2, Session 2024-25.", level='warning')
            return
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "PG Exam Results"
        
        # Define headers
        headers = [
            'Registration No',
            'Student Name',
            'College Code',
            'Department',
            'Program',
            'Batch',
            'Semester',
            'Session',
            'CIA Status',
            'ESE Status',
            'Semester Result',
            'SGPA',
            'Credits Earned',
            'Max Credits',
            'Next Semester',
            'Next Sem Status',
            'Published At',
            'Created At'
        ]
        
        # Style the header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Set column widths
        column_widths = [15, 25, 15, 20, 25, 15, 10, 10, 12, 12, 15, 8, 12, 12, 12, 15, 18, 18]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
        
        # Write data rows
        for row_num, result in enumerate(filtered_queryset.select_related('student', 'student__department', 'student__program', 'student__college'), 2):
            # Helper function to get CIA/ESE status display
            def get_status_display(status):
                if status is None:
                    return 'Pending'
                return 'PASS' if status else 'FAIL'
            
            row_data = [
                result.student.registration_no if result.student else 'N/A',
                result.student.first_name if result.student else 'N/A',
                result.student.college.college_code if result.student and result.student.college else 'N/A',
                result.student.department.name if result.student and result.student.department else 'N/A',
                result.student.program.name if result.student and result.student.program else 'N/A',
                result.student.batch if result.student else 'N/A',
                result.semester,
                result.session,
                get_status_display(result.cia_pass),
                get_status_display(result.ese_pass),
                result.semester_result,
                float(result.sgpa) if result.sgpa else 0.0,
                result.semester_credit_earned,
                result.semester_max_credit,
                result.next_semester if result.next_semester else 'N/A',
                result.next_sem_status if result.next_sem_status else 'N/A',
                result.published_at.strftime('%Y-%m-%d %H:%M:%S') if result.published_at else 'Not Published',
                result.created_at.strftime('%Y-%m-%d %H:%M:%S') if result.created_at else 'N/A'
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(vertical="center")
        
        # Freeze the header row
        ws.freeze_panes = ws['A2']
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=pg_exam_results_pass_promoted.xlsx'
        
        # Save workbook to response
        wb.save(response)
        
        # Show success message
        self.message_user(request, f"Successfully exported {filtered_queryset.count()} PASS/PROMOTED records (Sem 2, 2024-25) to Excel.", level='success')
        
        return response
    
    export_pass_promoted_to_excel.short_description = "Export Sem 2 (2024-25) PASS/PROMOTED to Excel"
    
    # Custom methods for list display
    def get_student_regno(self, obj):
        return obj.student.registration_no if obj.student else 'N/A'
    get_student_regno.short_description = 'Reg No'
    get_student_regno.admin_order_field = 'student__registration_no'
    
    def get_student_name(self, obj):
        if obj.student:
            return f"{obj.student.first_name} {obj.student.last_name}"
        return 'N/A'
    get_student_name.short_description = 'Student Name'
    get_student_name.admin_order_field = 'student__first_name'
    
    def get_student_dept(self, obj):
        if obj.student and obj.student.department:
            return obj.student.department.name
        return 'N/A'
    get_student_dept.short_description = 'Department'
    
    def cia_pass_display(self, obj):
        if obj.cia_pass is None:
            return '⏳ Pending'
        return '✅ PASS' if obj.cia_pass else '❌ FAIL'
    cia_pass_display.short_description = 'CIA Status'
    
    def ese_pass_display(self, obj):
        if obj.ese_pass is None:
            return '⏳ Pending'
        return '✅ PASS' if obj.ese_pass else '❌ FAIL'
    ese_pass_display.short_description = 'ESE Status'
    
    def get_student_full_info(self, obj):
        if not obj.student:
            return 'N/A'
        
        student = obj.student
        from django.utils.html import format_html
        return format_html(
            '<strong>Registration No:</strong> {}<br>'
            '<strong>Name:</strong> {} {}<br>'
            '<strong>Roll No:</strong> {}<br>'
            '<strong>Department:</strong> {}<br>'
            '<strong>Batch:</strong> {}',
            student.registration_no,
            student.first_name, student.last_name,
            student.roll_no if student.roll_no else 'N/A',
            student.department.name if student.department else 'N/A',
            student.batch if student.batch else 'N/A'
        )
    get_student_full_info.short_description = 'Student Information'
    
    def get_cia_courses(self, obj):
        """Show all CIA courses for this student in this semester"""
        from .models import PGStudentCourseAssessment
        from django.utils.html import format_html
        
        student_id = obj.student.id
        semester = obj.semester
        
        cia_courses = PGStudentCourseAssessment.objects.filter(
            student_id=student_id,
            semester=semester,
            label__icontains='CIA'
        )
        
        if not cia_courses.exists():
            return 'No CIA assessments found'
        
        rows = []
        for assessment in cia_courses:
            # Calculate pass status
            is_pass = False
            if assessment.ind_marks_obtained is not None and assessment.ind_pass_marks is not None:
                if not assessment.ind_is_absent:
                    is_pass = assessment.ind_marks_obtained >= assessment.ind_pass_marks
            
            status = '✅ PASS' if is_pass else '❌ FAIL'
            status_color = '#d4edda' if is_pass else '#f8d7da'
            
            rows.append(f'<tr style="background-color: {status_color};"><td style="padding: 5px;">{assessment.course_name[:50]}</td>'
                       f'<td style="text-align: center;">{assessment.ind_marks_obtained}/{assessment.ind_max_marks}</td>'
                       f'<td style="text-align: center;">{assessment.ind_pass_marks}</td>'
                       f'<td style="text-align: center;">{status}</td></tr>')
        
        html = ('<table style="width:100%; border-collapse: collapse;">'
                '<tr style="background-color: #f0f0f0;"><th style="padding: 5px; text-align: left;">Course</th><th>Marks</th><th>Pass Marks</th><th>Status</th></tr>'
                + ''.join(rows) + '</table>')
        
        return format_html(html)
    get_cia_courses.short_description = 'CIA Course Details'
    
    def get_ese_courses(self, obj):
        """Show all ESE courses for this student in this semester"""
        from .models import PGStudentCourseAssessment
        from django.utils.html import format_html
        
        student_id = obj.student.id
        semester = obj.semester
        
        ese_courses = PGStudentCourseAssessment.objects.filter(
            student_id=student_id,
            semester=semester,
            label__icontains='ESE'
        )
        
        if not ese_courses.exists():
            return 'No ESE assessments found (or not yet entered)'
        
        rows = []
        for assessment in ese_courses:
            # Calculate pass status
            is_pass = False
            if assessment.ind_marks_obtained is not None and assessment.ind_pass_marks is not None:
                if not assessment.ind_is_absent:
                    is_pass = assessment.ind_marks_obtained >= assessment.ind_pass_marks
            
            status = '✅ PASS' if is_pass else '❌ FAIL'
            status_color = '#d4edda' if is_pass else '#f8d7da'
            
            rows.append(f'<tr style="background-color: {status_color};"><td style="padding: 5px;">{assessment.course_name[:50]}</td>'
                       f'<td style="text-align: center;">{assessment.ind_marks_obtained}/{assessment.ind_max_marks}</td>'
                       f'<td style="text-align: center;">{assessment.ind_pass_marks}</td>'
                       f'<td style="text-align: center;">{status}</td></tr>')
        
        html = ('<table style="width:100%; border-collapse: collapse;">'
                '<tr style="background-color: #f0f0f0;"><th style="padding: 5px; text-align: left;">Course</th><th>Marks</th><th>Pass Marks</th><th>Status</th></tr>'
                + ''.join(rows) + '</table>')
        
        return format_html(html)
    get_ese_courses.short_description = 'ESE Course Details'
    
    fieldsets = (
        ('Student Information', {
            'fields': ('student',)
        }),
        ('Exam Details', {
            'fields': ('uid', 'semester', 'session')
        }),
        ('CIA Assessment', {
            'fields': ('cia_pass', 'get_cia_courses'),
            'description': 'CIA examination results'
        }),
        ('ESE Assessment', {
            'fields': ('ese_pass', 'get_ese_courses'),
            'description': 'ESE examination results'
        }),
        ('Semester Result', {
            'fields': (
                'semester_result',
                'sgpa',
                'semester_max_credit',
                'semester_credit_earned'
            )
        }),
        ('Promotion', {
            'fields': ('next_semester', 'next_sem_status')
        }),
        ('Meta', {
            'fields': ('is_legacy', 'published_at', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )

@admin.register(PGExam)
class PGExamAdmin(admin.ModelAdmin):
    list_display = ('name', 'year', 'session', 'batch', 'exam_month_year', 'publication_date', 'created_at')
    list_filter = ('session', 'year', 'batch')
    search_fields = ('name', 'session', 'batch')
    readonly_fields = ('uid', 'created_at', 'updated_at')
    ordering = ('-created_at',)

@admin.register(PGExamCenterMapping)
class PGExamCenterMappingAdmin(admin.ModelAdmin):
    list_display = ('center', 'get_exams_count', 'get_attached_colleges_count', 'created_at')
    list_filter = ('center',)
    search_fields = ('center__name', 'center__code')
    filter_horizontal = ('exams', 'attached_colleges')
    
    def get_exams_count(self, obj):
        return obj.exams.count()
    get_exams_count.short_description = 'Exams Count'
    
    def get_attached_colleges_count(self, obj):
        return obj.attached_colleges.count()
    get_attached_colleges_count.short_description = 'Attached Colleges Count'

@admin.register(PGGroup)
class PGGroupAdmin(admin.ModelAdmin):
    list_display = ('name', 'get_departments', 'created_at')
    list_filter = ('department',)
    search_fields = ('name', 'department__name')
    ordering = ('name',)
    filter_horizontal = ('department',)

    def get_departments(self, obj):
        return ", ".join([d.name for d in obj.department.all() if d.name])
    get_departments.short_description = 'Departments'

@admin.register(PGExamSchedule)
class PGExamScheduleAdmin(admin.ModelAdmin):
    list_display = ('exam', 'group', 'common_course_structure', 'exam_date', 'exam_time', 'sitting')
    list_filter = ('exam', 'group', 'exam_date', 'sitting')
    search_fields = ('exam__name', 'common_course_structure__course_code', 'common_course_structure__course_name')
    ordering = ('exam_date', 'exam_time')


@admin.register(PGExamRegistrationPayment)
class PGExamRegistrationPaymentAdmin(admin.ModelAdmin):
    list_display = (
        'order_id', 
        'get_student_info', 
        'amount', 
        'payment_status', 
        'tracking_id', 
        'payment_mode', 
        'created_at'
    )
    list_filter = ('payment_status', 'payment_mode', 'created_at')
    search_fields = (
        'order_id', 
        'tracking_id', 
        'bank_ref_no',
        'registration__student__registration_no',
        'registration__student__first_name',
        'registration__student__mobile_no'
    )
    readonly_fields = (
        'uid', 
        'registration', 
        'order_id', 
        'tracking_id', 
        'bank_ref_no', 
        'raw_response', 
        'created_at', 
        'updated_at'
    )
    ordering = ('-created_at',)
    
    def get_student_info(self, obj):
        student = obj.registration.student
        return f"{student.first_name} ({student.registration_no})"
    get_student_info.short_description = 'Student'

