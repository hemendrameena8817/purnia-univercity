import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.conf import settings
from pg.models import (
    PGExamRegistration,
    PGExamSchedule,
    PGCommonCourseStructure,
    PGExamCenterMapping,
    PGStudentCourseAssessment,
    PGDepartment,
    PGExamResult,
)
import re as _re
import logging
logger = logging.getLogger(__name__)

# ── Common Semester/Roman Mappings ───────────────────────────────────────
_roman_str_to_int = {
    'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5,
    'VI': 6, 'VII': 7, 'VIII': 8, 'IX': 9, 'X': 10,
}
roman_to_arabic_map = {k: str(v) for k, v in _roman_str_to_int.items()}
roman_to_arabic_map.update({'XI': '11', 'XII': '12', 'XIII': '13', 'XIV': '14', 'XV': '15'})
roman_map_inv = {v: k for k, v in roman_to_arabic_map.items()}

def normalize_code(code):
    if not code:
        return ""
    code = code.upper().strip().replace(" ", "-")
    m = _re.match(r'^([A-Z]+)(\d+)$', code)
    if m:
        code = f"{m.group(1)}-{m.group(2)}"
    
    if "-" in code:
        parts = code.rsplit("-", 1)
        prefix, suffix = parts[0], parts[1]
        if suffix in roman_to_arabic_map:
            return f"{prefix}-{roman_to_arabic_map[suffix]}"
        return code
    if code in roman_to_arabic_map:
        return roman_to_arabic_map[code]
    return code

def get_suffix(code):
    nc = normalize_code(code)
    return nc.split("-")[-1] if "-" in nc else nc

def _roman_to_int(roman):
    """Convert Roman numeral string to integer."""
    roman_map = {'I': 1, 'V': 5, 'X': 10, 'L': 50, 'C': 100}
    res = 0
    for i in range(len(roman)):
        if i > 0 and roman_map.get(roman[i], 0) > roman_map.get(roman[i-1], 0):
            res += roman_map.get(roman[i], 0) - 2 * roman_map.get(roman[i-1], 0)
        else:
            res += roman_map.get(roman[i], 0)
    return res

def generate_pg_roll_sheet_excel(exam, college, department=None, registration_no=None):
    """
    Generates and returns Exam Roll Sheet Excel for PG.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    ey = str(exam.year) if exam.year else ""
    sem_variants_int = set()
    if ey.isdigit():
        sem_variants_int.add(int(ey))

    if exam.name:
        roman_m = _re.search(r'\b(?:SEM|SEMESTER)[-\s]*(I{1,3}|IV|VI{0,3}|IX|X)\b', exam.name, _re.IGNORECASE)
        if roman_m:
            rn = roman_m.group(1).upper()
            if rn in _roman_str_to_int:
                sem_variants_int.add(_roman_str_to_int[rn])
        digit_m = _re.search(r'(?<!\d)([1-8])(?!\d)', exam.name)
        if digit_m:
            sem_variants_int.add(int(digit_m.group(1)))

    sem_variants_int = list(sem_variants_int)
    sem_variants = set()
    for s_int in sem_variants_int:
        s_str = str(s_int)
        roman_s = roman_map_inv.get(s_str, "")
        variants = [s_str, roman_s]
        if s_str == '1': variants.extend(['1ST', 'FIRST'])
        elif s_str == '2': variants.extend(['2ND', 'SECOND'])
        elif s_str == '3': variants.extend(['3RD', 'THIRD'])
        elif s_str == '4': variants.extend(['4TH', 'FOURTH'])
        
        for v in variants:
            if v:
                v_up = v.upper()
                sem_variants.add(v_up)
                sem_variants.add(f"SEM-{v_up}")
                sem_variants.add(f"SEM {v_up}")
                sem_variants.add(f"SEMESTER-{v_up}")
                sem_variants.add(f"SEMESTER {v_up}")
    
    sem_variants = list(sem_variants)

    _is_year_range = bool(_re.match(r'^\d{4}-\d{2,4}$', exam.session or ''))

    regs_qs = PGExamRegistration.objects.filter(
        exam=exam,
        student__college=college,
        status='REGISTERED',
    )

    if department:
        regs_qs = regs_qs.filter(student__department=department)

    if registration_no:
        regs_qs = regs_qs.filter(student__registration_no=registration_no)

    regs_qs = regs_qs.select_related(
        'student', 'student__department', 'student__program', 'student__degree'
    ).order_by('student__roll_no', 'student__registration_no')

    if not regs_qs.exists():
        return None

    # ── Exam center ──────────────────────────────────────────────────────────
    center_name = "-"
    cm = PGExamCenterMapping.objects.filter(exams=exam, attached_colleges=college).first()
    if cm and cm.center:
        center_name = cm.center.name

    schedules_all = PGExamSchedule.objects.filter(exam=exam).select_related('common_course_structure', 'group')

    # Group by department
    if department:
        depts_to_process = [department]
    else:
        dept_ids = regs_qs.exclude(student__department__isnull=True).values_list('student__department_id', flat=True).distinct()
        depts_to_process = list(PGDepartment.objects.filter(id__in=dept_ids).order_by('name'))
        if regs_qs.filter(student__department__isnull=True).exists():
            depts_to_process.append(None)

    for dept_obj in depts_to_process:
        dept_name = dept_obj.name if dept_obj else "General"
        sheet_title = _re.sub(r'[\\*?:/\[\]]', '', dept_name)[:30]
        ws = wb.create_sheet(title=sheet_title)

        dept_regs = regs_qs.filter(student__department=dept_obj) if dept_obj else regs_qs.filter(student__department__isnull=True)
        stu_ids = [r.student_id for r in dept_regs]

        # ── Subjects Column Headers ──────────────────────────────────────────
        subjects = []
        seen_norms = set()
        code_to_id = {}
        norm_to_id = {}

        if dept_obj:
            # Derive sessions and semesters from actual registrations in this dept
            dept_sessions = list(set(r.session for r in dept_regs if r.session))
            if not dept_sessions:
                dept_sessions = [exam.session] if exam.session else []

            dept_sems = list(set(r.sem for r in dept_regs if r.sem is not None))
            dept_sem_variants = set()
            for s_int in dept_sems:
                s_str = str(s_int)
                r_s = roman_map_inv.get(s_str, "")
                variants = [s_str, r_s]
                if s_str == '1': variants.extend(['1ST', 'FIRST'])
                elif s_str == '2': variants.extend(['2ND', 'SECOND'])
                elif s_str == '3': variants.extend(['3RD', 'THIRD'])
                elif s_str == '4': variants.extend(['4TH', 'FOURTH'])
                for v in variants:
                    if v:
                        v_u = v.upper()
                        dept_sem_variants.add(v_u)
                        dept_sem_variants.add(f"SEM-{v_u}")
                        dept_sem_variants.add(f"SEM {v_u}")
                        dept_sem_variants.add(f"SEMESTER-{v_u}")
                        dept_sem_variants.add(f"SEMESTER {v_u}")

            ese_rows = PGStudentCourseAssessment.objects.filter(
                student_id__in=stu_ids, 
                label__iregex=r'^ESE', 
                semester__in=list(dept_sem_variants) if dept_sem_variants else sem_variants,
                session__in=dept_sessions
            ).values('course_code', 'course_name').distinct()
            for row in ese_rows:
                nc = normalize_code(row['course_code'])
                if nc and nc not in seen_norms:
                    seen_norms.add(nc)
                    subjects.append({'code': row['course_code'], 'name': row['course_name'], 'norm': nc})
                    code_to_id[row['course_code'].upper()] = nc
                    norm_to_id[nc] = nc
            
            for s in schedules_all.filter(group__department=dept_obj):
                if s.common_course_structure:
                    sc = s.common_course_structure
                    nc = normalize_code(sc.course_code)
                    if nc and nc not in seen_norms:
                        seen_norms.add(nc)
                        subjects.append({'code': sc.course_code, 'name': sc.course_name, 'norm': nc})
                        code_to_id[sc.course_code.upper()] = nc
                        norm_to_id[nc] = nc
        else:
            for s in schedules_all:
                if s.common_course_structure:
                    sc = s.common_course_structure
                    nc = normalize_code(sc.course_code)
                    if nc and nc not in seen_norms:
                        seen_norms.add(nc)
                        subjects.append({'code': sc.course_code, 'name': sc.course_name, 'norm': nc})
                        code_to_id[sc.course_code.upper()] = nc

        def _sort_sub(s):
            n = s['norm']
            pref = n.split('-')[0]
            is_aecc = 1 if pref in ('AECC', 'AEC') else 0
            m = _re.search(r'(\d+)', n)
            # Use string for second element to avoid comparison errors between int and str
            val = f"{int(m.group(1)):03d}" if m else n
            return (is_aecc, val)
        subjects.sort(key=_sort_sub)

        # ── Styles ───────────────────────────────────────────────────────────
        header_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # ── Header Rows ──────────────────────────────────────────────────────
        ws.merge_cells('A1:E1')
        ws['A1'] = f"PURNEA UNIVERSITY, PURNIA"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_align

        ws.merge_cells('A2:E2')
        ws['A2'] = f"EXAMINATION ROLL SHEET"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = center_align

        ws['A4'] = f"Exam: {exam.name}"
        ws['A5'] = f"College: {college.name}"
        ws['A6'] = f"Department: {dept_name}"
        ws['D4'] = f"Session: {exam.session}"
        ws['D5'] = f"Center: {center_name}"

        # ── Table Header ─────────────────────────────────────────────────────
        start_row = 8
        headers = ["Sl No.", "Roll No.", "Reg. No.", "Candidate Name"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=i, value=h)
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_align

        for i, sub in enumerate(subjects, 5):
            val = f"{sub['code']}\n({sub['name']})" if sub['name'] else sub['code']
            cell = ws.cell(row=start_row, column=i, value=val)
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border
            cell.alignment = center_align

        # ── Data Rows ────────────────────────────────────────────────────────
        for idx, reg in enumerate(dept_regs, 1):
            curr_row = start_row + idx
            stu = reg.student
            ws.cell(row=curr_row, column=1, value=idx).border = thin_border
            ws.cell(row=curr_row, column=2, value=stu.roll_no).border = thin_border
            ws.cell(row=curr_row, column=3, value=stu.registration_no).border = thin_border
            ws.cell(row=curr_row, column=4, value=stu.get_full_name()).border = thin_border

            # Marks/Subject Registration check
            # For each student, use their specific registration's session and sem variants
            stu_sem_variants = set()
            if reg.sem:
                s_int = reg.sem
                s_str = str(s_int)
                r_s = roman_map_inv.get(s_str, "")
                variants = [s_str, r_s]
                if s_str == '1': variants.extend(['1ST', 'FIRST'])
                elif s_str == '2': variants.extend(['2ND', 'SECOND'])
                elif s_str == '3': variants.extend(['3RD', 'THIRD'])
                elif s_str == '4': variants.extend(['4TH', 'FOURTH'])
                for v in variants:
                    if v:
                        v_u = v.upper()
                        stu_sem_variants.add(v_u)
                        stu_sem_variants.add(f"SEM-{v_u}")
                        stu_sem_variants.add(f"SEM {v_u}")
                        stu_sem_variants.add(f"SEMESTER-{v_u}")
                        stu_sem_variants.add(f"SEMESTER {v_u}")

            stu_assessments = PGStudentCourseAssessment.objects.filter(
                student=stu, 
                semester__in=list(stu_sem_variants) if stu_sem_variants else sem_variants, 
                label__icontains='ESE', 
                session=reg.session if reg.session else exam.session
            )
            stu_codes = set()
            for a in stu_assessments:
                uc = (a.course_code or "").upper()
                sid = code_to_id.get(uc) or normalize_code(uc)
                stu_codes.add(sid)

            for j, sub in enumerate(subjects, 5):
                val = "√" if sub['norm'] in stu_codes else ""
                if reg.exam_type == 'REGULAR' and not val: # Fallback for regular if no assessment yet
                    val = "√"
                cell = ws.cell(row=curr_row, column=j, value=val)
                cell.border = thin_border
                cell.alignment = center_align

        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        for col in range(5, 5 + len(subjects)):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def generate_pg_tsi_excel(exam, college, department=None, registration_no=None):
    """
    Generates and returns TSI (Tabulation Sheet I) Excel for PG.
    """
    wb = openpyxl.Workbook()
    ws_base = wb.active if wb.active else wb.create_sheet()
    wb.remove(ws_base)

    ey = str(exam.year) if exam.year else ""
    sem_variants_int = set()
    if ey.isdigit(): sem_variants_int.add(int(ey))
    if exam.name:
        digit_m = _re.search(r'(?<!\d)([1-8])(?!\d)', exam.name)
        if digit_m: sem_variants_int.add(int(digit_m.group(1)))
    
    sem_variants_int = list(sem_variants_int)
    sem_variants = set()
    for s_int in sem_variants_int:
        s_str = str(s_int)
        roman_s = roman_map_inv.get(s_str, "")
        variants = [s_str, roman_s]
        if s_str == '1': variants.extend(['1ST', 'FIRST'])
        elif s_str == '2': variants.extend(['2ND', 'SECOND'])
        elif s_str == '3': variants.extend(['3RD', 'THIRD'])
        elif s_str == '4': variants.extend(['4TH', 'FOURTH'])
        
        for v in variants:
            if v:
                v_up = v.upper()
                sem_variants.add(v_up)
                sem_variants.add(f"SEM-{v_up}")
                sem_variants.add(f"SEM {v_up}")
                sem_variants.add(f"SEMESTER-{v_up}")
                sem_variants.add(f"SEMESTER {v_up}")
    
    sem_variants = list(sem_variants)

    regs_qs = PGExamRegistration.objects.filter(
        exam=exam,
        student__college=college, 
        status='REGISTERED'
    )

    regs_qs = regs_qs.select_related('student', 'student__department').order_by('student__roll_no')

    if not regs_qs.exists():
        return None

    # Group by department for sheets
    dept_ids = regs_qs.values_list('student__department_id', flat=True).distinct()
    departments = PGDepartment.objects.filter(id__in=dept_ids).order_by('name')

    for dept in departments:
        sheet_title = _re.sub(r'[\\*?:/\[\]]', '', dept.name)[:30]
        ws = wb.create_sheet(title=sheet_title)
        
        dept_regs = regs_qs.filter(student__department=dept)
        stu_ids = [r.student_id for r in dept_regs]

        # Derive sessions and semesters from registrations in this dept
        dept_sessions = list(set(r.session for r in dept_regs if r.session))
        if not dept_sessions:
            dept_sessions = [exam.session] if exam.session else []

        dept_sems = list(set(r.sem for r in dept_regs if r.sem is not None))
        dept_sem_variants = set()
        for s_int in dept_sems:
            s_str = str(s_int)
            r_s = roman_map_inv.get(s_str, "")
            variants = [s_str, r_s]
            if s_str == '1': variants.extend(['1ST', 'FIRST'])
            elif s_str == '2': variants.extend(['2ND', 'SECOND'])
            elif s_str == '3': variants.extend(['3RD', 'THIRD'])
            elif s_str == '4': variants.extend(['4TH', 'FOURTH'])
            for v in variants:
                if v:
                    v_u = v.upper()
                    dept_sem_variants.add(v_u)
                    dept_sem_variants.add(f"SEM-{v_u}")
                    dept_sem_variants.add(f"SEM {v_u}")
                    dept_sem_variants.add(f"SEMESTER-{v_u}")
                    dept_sem_variants.add(f"SEMESTER {v_u}")

        # Get relevant subjects for this department/semester
        subjects_qs = PGStudentCourseAssessment.objects.filter(
            student_id__in=stu_ids, 
            semester__in=list(dept_sem_variants) if dept_sem_variants else sem_variants, 
            label__iregex=r'^(ESE|CIA)',
            session__in=dept_sessions
        ).values('course_code', 'course_name').distinct()
        
        subjects = []
        seen_codes = set()
        for s in subjects_qs:
            code = (s['course_code'] or "").upper()
            if code and code not in seen_codes:
                seen_codes.add(code)
                subjects.append({'code': code, 'name': s['course_name']})
        
        def _sort_tsi(s):
            c = s['code']
            is_aecc = 1 if "AECC" in c or "AEC" in c else 0
            m = _re.search(r'(\d+)', c)
            val = f"{int(m.group(1)):03d}" if m else c
            return (is_aecc, val)
        subjects.sort(key=_sort_tsi)

        # ── TSI Headers ──────────────────────────────────────────────────────
        col_span = len(subjects)*3 + 7
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_span)
        ws['A1'] = f"PURNEA UNIVERSITY, PURNIA - TABULATION SHEET (TSI)"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A3'] = f"Exam: {exam.name}"
        ws['A4'] = f"College: {college.name}"
        ws['A5'] = f"Department: {dept.name}"

        # Table Header
        h_row = 7
        static_cols = ["Sl.", "Roll No.", "Registration No.", "Student Name"]
        for i, col in enumerate(static_cols, 1):
            ws.cell(row=h_row, column=i, value=col).font = Font(bold=True)
            ws.cell(row=h_row, column=i).border = Border(bottom=Side(style='thin'))
        c_idx = 5
        for sub in subjects:
            ws.merge_cells(start_row=h_row, start_column=c_idx, end_row=h_row, end_column=c_idx+2)
            val = f"{sub['code']}\n({sub['name']})" if sub['name'] else sub['code']
            cell = ws.cell(row=h_row, column=c_idx, value=val)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(bottom=Side(style='thin'))
            
            ws.cell(row=h_row+1, column=c_idx, value="CIA").font = Font(italic=True)
            ws.cell(row=h_row+1, column=c_idx+1, value="ESE").font = Font(italic=True)
            ws.cell(row=h_row+1, column=c_idx+2, value="TOT").font = Font(bold=True)
            c_idx += 3

        res_cols = ["Total", "SGPA", "Result"]
        for i, col in enumerate(res_cols, 0):
            ws.cell(row=h_row, column=c_idx + i, value=col).font = Font(bold=True)

        # ── Data Processing ───────────────────────────────────────────────────
        # For each student, use their specific registration contexts if possible, 
        # but for bulk load here we use the union of dept metadata
        assessments = PGStudentCourseAssessment.objects.filter(
            student_id__in=stu_ids, 
            semester__in=list(dept_sem_variants) if dept_sem_variants else sem_variants, 
            session__in=dept_sessions
        )
        asmt_map = {} # (student_id, course_code) -> {'CIA': marks, 'ESE': marks}
        for a in assessments:
            sid = a.student_id
            code = (a.course_code or "").upper()
            label = a.label.upper()
            if (sid, code) not in asmt_map: asmt_map[(sid, code)] = {'CIA': None, 'ESE': None}
            if 'CIA' in label: asmt_map[(sid, code)]['CIA'] = a.ind_marks_obtained
            if 'ESE' in label: asmt_map[(sid, code)]['ESE'] = a.ind_marks_obtained

        results = PGExamResult.objects.filter(student_id__in=stu_ids, semester=ey_for_str)
        res_map = {r.student_id: r for r in results}

        for idx, reg in enumerate(dept_regs, 1):
            curr_row = h_row + 2 + idx
            stu = reg.student
            ws.cell(row=curr_row, column=1, value=idx)
            ws.cell(row=curr_row, column=2, value=stu.roll_no)
            ws.cell(row=curr_row, column=3, value=stu.registration_no)
            ws.cell(row=curr_row, column=4, value=stu.get_full_name())

            c_idx = 5
            for sub in subjects:
                data = asmt_map.get((stu.id, sub['code']), {'CIA': '-', 'ESE': '-'})
                cia = data.get('CIA', '-')
                ese = data.get('ESE', '-')
                tot = (0 if cia == '-' else cia) + (0 if ese == '-' else ese)
                ws.cell(row=curr_row, column=c_idx, value=cia)
                ws.cell(row=curr_row, column=c_idx+1, value=ese)
                ws.cell(row=curr_row, column=c_idx+2, value=tot if (cia != '-' or ese != '-') else '-')
                c_idx += 3
            
            # Calculate total marks from asmt_map for this student
            stu_total_marks = 0
            has_marks = False
            for sub in subjects:
                data = asmt_map.get((stu.id, sub['code']), {})
                cia = data.get('CIA')
                ese = data.get('ESE')
                if cia is not None and cia != '-': stu_total_marks += float(cia); has_marks = True
                if ese is not None and ese != '-': stu_total_marks += float(ese); has_marks = True

            # Result data
            r_obj = res_map.get(stu.id)
            if r_obj:
                ws.cell(row=curr_row, column=c_idx, value=stu_total_marks if has_marks else "-")
                ws.cell(row=curr_row, column=c_idx+1, value=r_obj.sgpa)
                ws.cell(row=curr_row, column=c_idx+2, value=r_obj.semester_result)
            else:
                ws.cell(row=curr_row, column=c_idx, value=stu_total_marks if has_marks else "-")
                ws.cell(row=curr_row, column=c_idx+1, value="-")
                ws.cell(row=curr_row, column=c_idx+2, value="-")

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
