"""
Import just registered_applicant_master from XLSX.

Usage:
    poetry run python manage.py shell -c "exec(open('scripts/import_registered_applicant.py').read()); run()"
"""
from openpyxl import load_workbook
import os


def clean_value(value):
    if value is None:
        return None
    val_str = str(value).strip()
    if val_str in ('', 'None', 'NULL', '[NULL]', '\\N'):
        return None
    return val_str


def run():
    from staging.models import RegisteredApplicantMaster
    
    xlsx_path = 'old_data/registered_applicant_master_xml.xlsx'
    print(f"Loading: {xlsx_path}")
    
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).lower() if h else f'col_{i}' for i, h in enumerate(next(rows))]
    print(f"Columns: {headers}")
    
    # Delete existing
    existing = RegisteredApplicantMaster.objects.count()
    print(f"Deleting {existing} existing records...")
    RegisteredApplicantMaster.objects.all().delete()
    
    # Get model field names
    skip_fields = {'id', 'uid', 'is_migrated', 'migration_notes', 'imported_at'}
    model_fields = [f.name for f in RegisteredApplicantMaster._meta.fields if f.name not in skip_fields]
    
    imported = 0
    for row in rows:
        d = dict(zip(headers, row))
        
        kwargs = {}
        for field in model_fields:
            # Try to get value from XLSX using field name
            value = d.get(field)
            # Handle 'id' column -> 'csv_id' field
            if field == 'csv_id':
                value = d.get('id')
            kwargs[field] = clean_value(value)
        
        RegisteredApplicantMaster.objects.create(**kwargs)
        imported += 1
        if imported % 10000 == 0:
            print(f"...{imported} records")
    
    wb.close()
    print(f"✅ Done! Imported {imported} records")
    print(f"Total in table: {RegisteredApplicantMaster.objects.count()}")


if __name__ == '__main__':
    run()
