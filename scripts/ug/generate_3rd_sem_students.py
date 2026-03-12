import os
import django
import sys
from pathlib import Path
from datetime import datetime

BASE_DIR = Path(__file__).resolve().parent.parent.parent
sys.path.append(str(BASE_DIR))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'pup_umis_backend.settings.development')
django.setup()

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from ug.models import SemesterRegistration, CourseStructure


def get_aec_sec_papers(student, semester='3'):
    aec_paper = ""
    sec_paper = ""

    # if not student.batch:
    #     return aec_paper, sec_paper

    semester_course_structures = CourseStructure.objects.filter(
        semester=str(semester),
        course_code__in=['AEC-3', 'SEC-3']
    ).order_by('course_code', 'id')

    for course in semester_course_structures:
        if course.course_code == 'AEC-3':
            aec_paper = course.course_name or ""
        if course.course_code == 'SEC-3':
            sec_paper = course.course_name or ""

    return aec_paper, sec_paper


def generate_student_report():
    print("Starting student report generation...")

    semester_registrations = SemesterRegistration.objects.filter(
        sem=3,
        session='2025-26',
        status='REGISTERED'
    ).select_related(
        'student',
        'student__college',
        'student__major_course',
        'student__minor_course',
        'student__mdc_course',
        'student__batch'
    ).order_by('student__college__college_code', 'student__registration_no')

    print(f"Found {semester_registrations.count()} registered students in 3rd semester 2025-26")

    wb = Workbook()
    ws = wb.active
    ws.title = '3rd Sem Students 2025-26'

    headers = [
        'College Code',
        'College Name',
        'Student Name',
        'Roll No',
        'Registration No',
        'Major Course Name',
        'Minor Course Name',
        'MDC Course Name',
        'AEC Paper Name',
        'SEC Paper Name'
    ]

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    column_widths = {
        'A': 15,
        'B': 40,
        'C': 30,
        'D': 15,
        'E': 20,
        'F': 30,
        'G': 30,
        'H': 30,
        'I': 40,
        'J': 40,
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    row_num = 2
    for reg in semester_registrations:
        student = reg.student
        student_name = f"{student.first_name or ''} {student.last_name or ''}".strip() or 'N/A'
        college_code = student.college.college_code if student.college else 'N/A'
        college_name = student.college.name if student.college else 'N/A'
        major_course_name = student.major_course.name if student.major_course else 'N/A'
        minor_course_name = student.minor_course.name if student.minor_course else 'N/A'
        mdc_course_name = student.mdc_course.name if student.mdc_course else 'N/A'
        aec_paper, sec_paper = get_aec_sec_papers(student, semester='3')

        row_data = [
            college_code,
            college_name,
            student_name,
            student.roll_no or 'N/A',
            student.registration_no or 'N/A',
            major_course_name,
            minor_course_name,
            mdc_course_name,
            aec_paper or 'N/A',
            sec_paper or 'N/A'
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        row_num += 1

    ws.freeze_panes = 'A2'

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = BASE_DIR / f'3rd_sem_students_2025-26_{timestamp}.xlsx'
    wb.save(filename)

    print(f"\n✅ Report generated successfully: {filename.name}")
    print(f"Total students: {row_num - 2}")

    return filename


if __name__ == '__main__':
    try:
        filename = generate_student_report()
        print(f"\nFile saved at: {filename}")
    except Exception as e:
        print(f"\n❌ Error generating report: {str(e)}")
        import traceback
        traceback.print_exc()
