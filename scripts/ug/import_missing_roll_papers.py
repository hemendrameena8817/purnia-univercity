import argparse
import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from pathlib import Path

import django
from openpyxl import load_workbook
try:
    import xlrd
except ImportError:
    xlrd = None

BASE_DIR = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(BASE_DIR))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'pup_umis_backend.settings')
django.setup()

from django.db import transaction
from django.utils.timezone import localdate, make_aware
from ug.management.commands.map_mjc_mic_mdc_live import MJC_COURSE_MAP, MIC_COURSE_MAP, MDC_COURSE_MAP
from ug.models import CourseStructure, ExamRegistration, StudentCourseAssessment, UGDepartment, UGStudentProfile

DEFAULT_XLSX_PATH = BASE_DIR / 'scripts' / 'ug' / 'K_B_JHA_COLLEGE_KATIHAR_Upload_Template.xlsx'
TARGET_SESSION = '2025-26'
TARGET_SEMESTER = '1ST'
TARGET_SEM_INT = 1
TARGET_EXAM_TYPE = 'BACK'
TARGET_EXAM_REGISTRATION_FEES = 600
TARGET_EXAM_REGISTRATION_START = make_aware(datetime.strptime('09-03-2026 00:00:00', '%d-%m-%Y %H:%M:%S'))
TARGET_EXAM_REGISTRATION_END = make_aware(datetime.strptime('15-03-2026 00:00:00', '%d-%m-%Y %H:%M:%S'))
FALLBACK_PAPER_CODES = {
    'MJC': '1001',
    'MIC': '1002',
    'SEC': '1003',
    'VAC': '1004',
    'MDC': '1005',
    'AEC': '1006',
}
LABEL_MARKS = {
    'CIA-Theory': (30, Decimal('13.5')),
    'ESE-Theory': (70, Decimal('31.5')),
    'CIA-Practical': (30, Decimal('13.5')),
    'ESE-Practical': (70, Decimal('31.5')),
}
PAPER_NAME_ALIASES = {
    'understanding poltical theory': 'understanding political theory',
    'introduction to sociology -1': 'introduction to sociology - i',
    'introduction to sociology-1': 'introduction to sociology - i',
    'introduction to sociology-i': 'introduction to sociology - i',
    'introduction to sociology- i': 'introduction to sociology - i',
    'introduction to sociology -i': 'introduction to sociology - i',
    'introduction to sociology-i ': 'introduction to sociology - i',
    'introduction to sociology- i ': 'introduction to sociology - i',
    'introduction to sociology- i': 'introduction to sociology - i',
    'introduction to sociology-i': 'introduction to sociology - i',
    'introduction to sociology - i': 'introduction to sociology - i',
    'introduction to socilogy -1': 'introduction to sociology - i',
    'introduction to socilogy-1': 'introduction to sociology - i',
    'decductive logic': 'deductive logic',
    'study of urdu fiction': 'study of urdu fiction',
    'hindi': 'mil-hindi',
    'hindi communication': 'mil-hindi',
    'english communication': 'mil english communication',
}
DISPLAY_NAME_ALIASES = {
    'mil-hindi': 'MIL-Hindi',
    'mil english communication': 'MIL English Communication',
    'understanding political theory': 'Understanding Political Theory',
    'introduction to sociology - i': 'Introduction to Sociology - I',
    'study of urdu fiction': 'Study of Urdu Fiction',
    'deductive logic': 'Deductive Logic',
}
HEADER_ALIASES = {
    'reg no': 'Registration No',
    'reg. no': 'Registration No',
    'registration no': 'Registration No',
    'registration no.': 'Registration No',
    'registration number': 'Registration No',
    'regsitration no': 'Registration No',
    'regsitration number': 'Registration No',
    'course code': 'Course Code',
    'couse code': 'Course Code',
    'subject code': 'Course Code',
    'paper name': 'Paper Name',
    'ppaer name': 'Paper Name',
    'paper code': 'Paper Code',
    'roll no': 'Roll No',
    'roll no.': 'Roll No',
    'exam type': 'Exam Type',
    'semester': 'Semester',
    'theory': 'Theory',
    'practical': 'Practical',
}


def clean_text(value):
    text = str(value or '').replace('_x000D_', ' ').replace('\n', ' ').replace('\r', ' ')
    return ' '.join(text.split()).strip()


def normalize_header(value):
    text = clean_text(value)
    alias_key = text.lower().replace('_', ' ')
    return HEADER_ALIASES.get(alias_key, text)


def normalize_paper_name(value):
    text = clean_text(value).lower()
    text = text.replace('–', '-').replace('—', '-')
    text = re.sub(r'\s*-\s*1\b', ' - i', text)
    text = re.sub(r'\s*-\s*i\b', ' - i', text)
    text = PAPER_NAME_ALIASES.get(text, text)
    return text


def display_paper_name(value):
    key = normalize_paper_name(value)
    return DISPLAY_NAME_ALIASES.get(key, clean_text(value))


def normalize_course_code(value):
    text = clean_text(value).upper().replace(' ', '')
    text = text.replace('–', '-').replace('—', '-')
    text = re.sub(r'-(I|1)$', '-1', text)
    return text


def get_paper_code_suffix(value):
    text = re.sub(r'[^A-Z0-9]', '', clean_text(value).upper())
    return text[-4:] if len(text) >= 4 else text


def get_assessment_dedupe_key(paper_code, label, exam_type, semester, session):
    return (
        get_paper_code_suffix(paper_code),
        clean_text(label),
        clean_text(exam_type),
        clean_text(semester),
        clean_text(session),
    )


def get_student_assessment_dedupe_key(student_id, paper_code, label, exam_type, semester, session):
    return (student_id,) + get_assessment_dedupe_key(
        paper_code=paper_code,
        label=label,
        exam_type=exam_type,
        semester=semester,
        session=session,
    )


def dedupe_pending_assessments(assessments):
    unique_assessments = []
    seen_keys = set()

    for assessment in assessments:
        key = get_student_assessment_dedupe_key(
            student_id=assessment.student_id,
            paper_code=assessment.paper_code,
            label=assessment.label,
            exam_type=assessment.exam_type,
            semester=assessment.semester,
            session=assessment.session,
        )
        if key in seen_keys:
            continue
        seen_keys.add(key)
        unique_assessments.append(assessment)

    return unique_assessments


def normalize_semester(value):
    text = clean_text(value).upper().replace(' ', '')
    if text in {'1', '1ST', 'I'}:
        return '1ST'
    return clean_text(value).upper()


def normalize_exam_type(value):
    text = clean_text(value).upper().replace(' ', '')
    if text == 'REGULAR':
        return 'Regular'
    if text == 'BACK':
        return 'Back'
    return clean_text(value).title() or 'Regular'


def is_yes(value):
    return clean_text(value).lower() in {'yes', 'y', 'true'}


def build_name_map(raw_map):
    return {normalize_paper_name(name): code for name, code in raw_map.items()}


def load_excel_rows(file_path):
    suffix = file_path.suffix.lower()
    if suffix == '.xls':
        if xlrd is None:
            raise ImportError('xlrd is required to read .xls files. Install it and run again.')
        workbook = xlrd.open_workbook(str(file_path))
        sheet = workbook.sheet_by_index(0)
        return [tuple(sheet.row_values(row_index)) for row_index in range(sheet.nrows)]

    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active
    return list(sheet.iter_rows(values_only=True))


def load_students_from_file(file_path):
    rows = load_excel_rows(file_path)
    if not rows:
        return {}

    header_row_index = None
    headers = []
    for index, row in enumerate(rows):
        candidate_headers = [normalize_header(header) for header in row]
        if 'Registration No' in candidate_headers:
            header_row_index = index
            headers = candidate_headers
            break
    if header_row_index is None:
        return {}

    grouped = defaultdict(lambda: {
        'papers': {},
        'sheet_roll_assigned': False,
    })

    for row in rows[header_row_index + 1:]:
        data = dict(zip(headers, row))
        reg_no = clean_text(data.get('Registration No')).upper()
        if not reg_no:
            continue
        roll_no = clean_text(data.get('Roll No'))
        if roll_no:
            grouped[reg_no]['sheet_roll_assigned'] = True

        semester = normalize_semester(data.get('Semester'))
        exam_type = normalize_exam_type(data.get('Exam Type'))
        course_code = normalize_course_code(data.get('Course Code'))
        paper_name_key = normalize_paper_name(data.get('Paper Name'))
        paper_name_display = display_paper_name(data.get('Paper Name'))

        if not course_code or not paper_name_key:
            continue

        key = (semester, exam_type, course_code, paper_name_key)
        entry = grouped[reg_no]['papers'].get(key)
        if not entry:
            entry = {
                'semester': semester,
                'exam_type': exam_type,
                'course_code': course_code,
                'paper_name_key': paper_name_key,
                'paper_name_display': paper_name_display,
                'has_theory': False,
                'has_practical': False,
            }
            grouped[reg_no]['papers'][key] = entry

        entry['has_theory'] = entry['has_theory'] or is_yes(data.get('Theory'))
        entry['has_practical'] = entry['has_practical'] or is_yes(data.get('Practical'))
        if len(paper_name_display) > len(entry['paper_name_display']):
            entry['paper_name_display'] = paper_name_display

    return {
        reg_no: {
            'papers': list(payload['papers'].values()),
            'sheet_roll_assigned': payload['sheet_roll_assigned'],
        }
        for reg_no, payload in grouped.items()
    }


def build_course_structure_maps():
    course_lookup = defaultdict(list)
    default_paper_code = {}
    course_rows = CourseStructure.objects.filter(semester='1').select_related('department').order_by('id')
    for course in course_rows:
        course_code = normalize_course_code(course.course_code)
        paper_name_key = normalize_paper_name(course.course_name)
        course_lookup[(course_code, paper_name_key)].append(course)
        if course_code and course.paper_code and course_code not in default_paper_code:
            default_paper_code[course_code] = clean_text(course.paper_code)
    return course_lookup, default_paper_code


def pick_course_structure(course_lookup, course_code, paper_name_key, mapped_department):
    candidates = course_lookup.get((course_code, paper_name_key), [])
    if not candidates:
        return None
    if mapped_department:
        for candidate in candidates:
            if candidate.department_id == mapped_department.id:
                return candidate
    return candidates[0]


def get_department_for_course(course_code, paper_name_key, dept_map, mjc_map, mic_map, mdc_map):
    prefix = course_code.split('-', 1)[0]
    if prefix == 'MJC':
        dept_code = mjc_map.get(paper_name_key)
    elif prefix == 'MIC':
        dept_code = mic_map.get(paper_name_key)
    elif prefix == 'MDC':
        dept_code = mdc_map.get(paper_name_key)
    else:
        dept_code = None
    if not dept_code:
        return None
    return dept_map.get(dept_code.upper())


def create_assessment(student, semester, session, exam_type, course_code, course_name, course_short_name, paper_code, department, labels, source_file):
    course_type = course_code.split('-', 1)[0] if '-' in course_code else course_code
    degree_code = student.degree.short_name or student.degree.name if student.degree else None
    college_code = student.college.college_code if student.college else None
    assessments = []
    for label in labels:
        max_marks, pass_marks = LABEL_MARKS[label]
        assessments.append(StudentCourseAssessment(
            student=student,
            semester=semester,
            session=session,
            batch=student.batch,
            paper_code=paper_code,
            course_name=course_name,
            course_short_name=course_short_name,
            course_code=course_code,
            course_type=course_type,
            department=department,
            degree=degree_code,
            college_code=college_code,
            exam_type=exam_type,
            label=label,
            ind_max_marks=max_marks,
            ind_pass_marks=pass_marks,
            ind_is_absent=False,
            json_data={
                'source_file': str(source_file),
                'source_type': 'missing_roll_template_import',
            },
        ))
    return assessments


def delete_today_imported_assessments(reg_nos):
    if not reg_nos:
        print(
            f"Deleted today's imported assessments: 0 | "
            f"session={TARGET_SESSION} semester={TARGET_SEMESTER} exam_type={TARGET_EXAM_TYPE}"
        )
        return
    today = localdate()
    queryset = StudentCourseAssessment.objects.filter(
        student__registration_no__in=reg_nos,
        created_at__date=today,
    )
    deleted_count = queryset.count()
    if deleted_count:
        queryset.delete()
    print(
        f"Deleted today's imported assessments: {deleted_count} | "
        f"session={TARGET_SESSION} semester={TARGET_SEMESTER} exam_type={TARGET_EXAM_TYPE}"
    )


def remove_duplicate_assessments(reg_nos, dry_run=False):
    if not reg_nos:
        print(
            f"Removed duplicate assessments: 0 | "
            f"session={TARGET_SESSION} semester={TARGET_SEMESTER} exam_type={TARGET_EXAM_TYPE}"
        )
        return 0

    queryset = StudentCourseAssessment.objects.filter(
        student__registration_no__in=reg_nos,
        semester=TARGET_SEMESTER,
        session=TARGET_SESSION,
        exam_type=TARGET_EXAM_TYPE,
    ).values_list(
        'id',
        'student_id',
        'paper_code',
        'label',
    ).order_by('student_id', 'label', 'paper_code', 'id')

    ids_to_delete = []
    seen_keys = set()

    for assessment_id, student_id, paper_code, label in queryset.iterator(chunk_size=2000):
        key = get_student_assessment_dedupe_key(
            student_id=student_id,
            paper_code=paper_code,
            label=label,
            exam_type=TARGET_EXAM_TYPE,
            semester=TARGET_SEMESTER,
            session=TARGET_SESSION,
        )
        if key in seen_keys:
            ids_to_delete.append(assessment_id)
            continue
        seen_keys.add(key)

    if ids_to_delete and not dry_run:
        StudentCourseAssessment.objects.filter(id__in=ids_to_delete).delete()

    print(
        f"Removed duplicate assessments: {len(ids_to_delete)} | "
        f"session={TARGET_SESSION} semester={TARGET_SEMESTER} exam_type={TARGET_EXAM_TYPE}"
    )
    return len(ids_to_delete)


def cleanup_only(file_path=None, registration_no=None, dry_run=False):
    if registration_no:
        reg_nos = [clean_text(registration_no).upper()]
    elif file_path:
        reg_nos = list(load_students_from_file(file_path).keys())
    else:
        reg_nos = []

    removed_duplicates = remove_duplicate_assessments(reg_nos, dry_run=dry_run)
    if dry_run:
        print('DRY RUN: no changes saved')
    return removed_duplicates


def get_roll_number_state():
    existing_roll_nos = set()
    numeric_roll_nos = []

    for raw_roll_no in UGStudentProfile.objects.exclude(roll_no__isnull=True).exclude(roll_no='').values_list('roll_no', flat=True):
        roll_no = clean_text(raw_roll_no)
        if not roll_no:
            continue
        existing_roll_nos.add(roll_no)
        if roll_no.isdigit():
            numeric_roll_nos.append(roll_no)

    numeric_width = max((len(roll_no) for roll_no in numeric_roll_nos), default=1)
    next_roll_number = max((int(roll_no) for roll_no in numeric_roll_nos), default=0) + 1
    return existing_roll_nos, next_roll_number, numeric_width


def assign_next_roll_no(existing_roll_nos, next_roll_number, numeric_width):
    while True:
        candidate_number = str(next_roll_number)
        candidate_roll_no = candidate_number.zfill(numeric_width) if numeric_width > len(candidate_number) else candidate_number
        next_roll_number += 1
        if candidate_roll_no in existing_roll_nos:
            continue
        existing_roll_nos.add(candidate_roll_no)
        return candidate_roll_no, next_roll_number


def ensure_exam_registration(profile, source_file):
    registration = ExamRegistration.objects.filter(
        student=profile,
        sem=TARGET_SEM_INT,
        session=TARGET_SESSION,
        exam_type=TARGET_EXAM_TYPE,
    ).order_by('-created_at').first()

    update_fields = []
    if registration:
        if registration.status == 'REGISTERED':
            return registration, False, update_fields
        if registration.status != 'OPEN':
            registration.status = 'OPEN'
            update_fields.append('status')
        if registration.is_open is not True:
            registration.is_open = True
            update_fields.append('is_open')
        if registration.fees != TARGET_EXAM_REGISTRATION_FEES:
            registration.fees = TARGET_EXAM_REGISTRATION_FEES
            update_fields.append('fees')
        if registration.start_date != TARGET_EXAM_REGISTRATION_START:
            registration.start_date = TARGET_EXAM_REGISTRATION_START
            update_fields.append('start_date')
        if registration.end_date != TARGET_EXAM_REGISTRATION_END:
            registration.end_date = TARGET_EXAM_REGISTRATION_END
            update_fields.append('end_date')
        json_data = registration.json_data or {}
        if json_data.get('source_type') != 'missing_roll_template_import' or json_data.get('source_file') != str(source_file):
            json_data.update({
                'source_file': str(source_file),
                'source_type': 'missing_roll_template_import',
            })
            registration.json_data = json_data
            update_fields.append('json_data')
        return registration, False, update_fields

    registration = ExamRegistration(
        student=profile,
        sem=TARGET_SEM_INT,
        session=TARGET_SESSION,
        exam_type=TARGET_EXAM_TYPE,
        status='OPEN',
        is_open=True,
        fees=TARGET_EXAM_REGISTRATION_FEES,
        start_date=TARGET_EXAM_REGISTRATION_START,
        end_date=TARGET_EXAM_REGISTRATION_END,
        json_data={
            'source_file': str(source_file),
            'source_type': 'missing_roll_template_import',
        },
    )
    return registration, True, ['status', 'is_open', 'fees', 'start_date', 'end_date', 'json_data']


def process_file(file_path, dry_run=False):
    grouped_students = load_students_from_file(file_path)
    print(f'Found {len(grouped_students)} students with registration number in {file_path}')
    if not grouped_students:
        return

    reg_nos = list(grouped_students.keys())
    if not dry_run:
        delete_today_imported_assessments(reg_nos)
        removed_duplicates = remove_duplicate_assessments(reg_nos)
    else:
        removed_duplicates = 0

    profiles = {
        profile.registration_no: profile
        for profile in UGStudentProfile.objects.select_related('college', 'degree', 'batch', 'major_course', 'minor_course', 'mdc_course').filter(
            registration_no__in=reg_nos
        )
    }
    dept_map = {
        department.code.upper(): department
        for department in UGDepartment.objects.filter(is_publish=True)
        if department.code
    }
    mjc_map = build_name_map(MJC_COURSE_MAP)
    mic_map = build_name_map(MIC_COURSE_MAP)
    mdc_map = build_name_map(MDC_COURSE_MAP)
    course_lookup, default_paper_code = build_course_structure_maps()

    profile_updates = []
    assessments_to_create = []
    exam_registrations_to_create = []
    exam_registrations_to_update = []
    existing_roll_nos, next_roll_number, numeric_width = get_roll_number_state()
    stats = {
        'students_in_file': len(grouped_students),
        'profiles_found': 0,
        'profiles_updated': 0,
        'roll_numbers_assigned': 0,
        'assessments_created': 0,
        'assessment_skipped_existing': 0,
        'duplicate_assessments_removed': removed_duplicates,
        'exam_registrations_created': 0,
        'exam_registrations_updated': 0,
        'missing_profiles': 0,
        'missing_paper_code': 0,
        'conflicts': 0,
    }

    for reg_no, student_payload in grouped_students.items():
        profile = profiles.get(reg_no)
        if not profile:
            print(f'PROFILE NOT FOUND: {reg_no}')
            stats['missing_profiles'] += 1
            continue

        stats['profiles_found'] += 1

        papers = student_payload['papers']
        mapped_departments = {'MJC': set(), 'MIC': set(), 'MDC': set()}
        paper_payloads = []

        for paper in papers:
            mapped_department = get_department_for_course(
                paper['course_code'],
                paper['paper_name_key'],
                dept_map,
                mjc_map,
                mic_map,
                mdc_map,
            )
            prefix = paper['course_code'].split('-', 1)[0]
            if prefix in mapped_departments and mapped_department:
                mapped_departments[prefix].add(mapped_department)

            course_structure = pick_course_structure(
                course_lookup,
                paper['course_code'],
                paper['paper_name_key'],
                mapped_department,
            )
            paper_code = clean_text(course_structure.paper_code) if course_structure and course_structure.paper_code else default_paper_code.get(paper['course_code'])
            if not paper_code:
                paper_code = FALLBACK_PAPER_CODES.get(prefix)
            if not paper_code:
                print(f'PAPER CODE NOT FOUND: {reg_no} | {paper["course_code"]} | {paper["paper_name_display"]}')
                stats['missing_paper_code'] += 1
                continue

            labels = []
            if paper['has_theory']:
                labels.extend(['CIA-Theory', 'ESE-Theory'])
            if paper['has_practical']:
                labels.extend(['CIA-Practical', 'ESE-Practical'])
            if not labels:
                labels.extend(['CIA-Theory', 'ESE-Theory'])

            paper_payloads.append({
                'semester': TARGET_SEMESTER,
                'exam_type': TARGET_EXAM_TYPE,
                'course_code': paper['course_code'],
                'paper_code': paper_code,
                'course_name': clean_text(course_structure.course_name) if course_structure and course_structure.course_name else paper['paper_name_display'],
                'course_short_name': clean_text(course_structure.course_short_name) if course_structure and course_structure.course_short_name else None,
                'department': mapped_department or (course_structure.department if course_structure else None),
                'labels': labels,
            })

        conflict = False
        for prefix, departments in mapped_departments.items():
            if len(departments) > 1:
                dept_codes = ', '.join(sorted({department.code for department in departments if department.code}))
                print(f'CONFLICT: {reg_no} | {prefix} -> {dept_codes}')
                stats['conflicts'] += 1
                conflict = True
        if conflict:
            continue

        update_fields = []
        major_department = next(iter(mapped_departments['MJC'])) if mapped_departments['MJC'] else None
        minor_department = next(iter(mapped_departments['MIC'])) if mapped_departments['MIC'] else None
        mdc_department = next(iter(mapped_departments['MDC'])) if mapped_departments['MDC'] else None

        if major_department and profile.major_course_id != major_department.id:
            profile.major_course = major_department
            update_fields.append('major_course')
        if minor_department and profile.minor_course_id != minor_department.id:
            profile.minor_course = minor_department
            update_fields.append('minor_course')
        if mdc_department and profile.mdc_course_id != mdc_department.id:
            profile.mdc_course = mdc_department
            update_fields.append('mdc_course')
        if not student_payload['sheet_roll_assigned'] and not clean_text(profile.roll_no):
            assigned_roll_no, next_roll_number = assign_next_roll_no(existing_roll_nos, next_roll_number, numeric_width)
            profile.roll_no = assigned_roll_no
            update_fields.append('roll_no')
            stats['roll_numbers_assigned'] += 1
            print(f'ROLL NO ASSIGN: {reg_no} | {assigned_roll_no}')
        if update_fields:
            profile_updates.append((profile, update_fields))
            print(f'PROFILE UPDATE: {reg_no} | {", ".join(update_fields)}')

        exam_registration, registration_created, registration_update_fields = ensure_exam_registration(profile, file_path)
        if registration_created:
            exam_registrations_to_create.append(exam_registration)
            stats['exam_registrations_created'] += 1
            print(
                f'EXAM REGISTRATION CREATE: {reg_no} | '
                f'SEM={TARGET_SEM_INT} | SESSION={TARGET_SESSION} | TYPE={TARGET_EXAM_TYPE} | FEES={TARGET_EXAM_REGISTRATION_FEES}'
            )
        elif registration_update_fields:
            exam_registrations_to_update.append((exam_registration, registration_update_fields))
            stats['exam_registrations_updated'] += 1
            print(
                f'EXAM REGISTRATION UPDATE: {reg_no} | '
                f'{", ".join(registration_update_fields)}'
            )

        existing_keys = set(
            get_student_assessment_dedupe_key(
                student_id=profile.id,
                paper_code=paper_code,
                label=label,
                exam_type=exam_type,
                semester=semester,
                session=session,
            )
            for paper_code, label, exam_type, semester, session in StudentCourseAssessment.objects.filter(
                student=profile,
                semester=TARGET_SEMESTER,
                session=TARGET_SESSION,
                exam_type=TARGET_EXAM_TYPE,
            ).values_list('paper_code', 'label', 'exam_type', 'semester', 'session')
        )
        session_value = TARGET_SESSION

        for payload in paper_payloads:
            for assessment in create_assessment(
                student=profile,
                semester=payload['semester'],
                session=session_value,
                exam_type=payload['exam_type'],
                course_code=payload['course_code'],
                course_name=payload['course_name'],
                course_short_name=payload['course_short_name'],
                paper_code=payload['paper_code'],
                department=payload['department'],
                labels=payload['labels'],
                source_file=file_path,
            ):
                key = get_student_assessment_dedupe_key(
                    student_id=profile.id,
                    paper_code=assessment.paper_code,
                    label=assessment.label,
                    exam_type=assessment.exam_type,
                    semester=assessment.semester,
                    session=assessment.session,
                )
                if key in existing_keys:
                    stats['assessment_skipped_existing'] += 1
                    continue
                assessments_to_create.append(assessment)
                existing_keys.add(key)
                stats['assessments_created'] += 1
                print(f'ASSESSMENT CREATE: {reg_no} | {assessment.course_code} | {assessment.course_name} | {assessment.label}')

    if not dry_run:
        with transaction.atomic():
            for profile, update_fields in profile_updates:
                profile.save(update_fields=update_fields)
            for exam_registration, update_fields in exam_registrations_to_update:
                exam_registration.save(update_fields=update_fields)
            if exam_registrations_to_create:
                ExamRegistration.objects.bulk_create(exam_registrations_to_create, batch_size=500)
            if assessments_to_create:
                assessments_to_create = dedupe_pending_assessments(assessments_to_create)
                StudentCourseAssessment.objects.bulk_create(assessments_to_create, batch_size=500)

    stats['profiles_updated'] = len(profile_updates)

    print('')
    print(f"Students in file: {stats['students_in_file']}")
    print(f"Profiles found: {stats['profiles_found']}")
    print(f"Profiles updated: {stats['profiles_updated']}")
    print(f"Roll numbers assigned: {stats['roll_numbers_assigned']}")
    print(f"Assessments created: {stats['assessments_created']}")
    print(f"Assessments skipped existing: {stats['assessment_skipped_existing']}")
    print(f"Duplicate assessments removed: {stats['duplicate_assessments_removed']}")
    print(f"Exam registrations created: {stats['exam_registrations_created']}")
    print(f"Exam registrations updated: {stats['exam_registrations_updated']}")
    print(f"Profiles missing: {stats['missing_profiles']}")
    print(f"Paper code missing: {stats['missing_paper_code']}")
    print(f"Conflicts: {stats['conflicts']}")
    if dry_run:
        print('DRY RUN: no changes saved')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--file-path', default=str(DEFAULT_XLSX_PATH))
    parser.add_argument('--registration-no')
    parser.add_argument('--cleanup-only', action='store_true')
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()
    if args.cleanup_only:
        cleanup_only(
            file_path=Path(args.file_path) if args.file_path else None,
            registration_no=args.registration_no,
            dry_run=args.dry_run,
        )
        return
    process_file(Path(args.file_path), dry_run=args.dry_run)


if __name__ == '__main__':
    main()
