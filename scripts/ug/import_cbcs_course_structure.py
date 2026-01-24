"""
Script to import CBCS Course Structure from Excel file.
Creates Faculty, Department, and CourseStructure entries for UG app.

Usage:
    poetry run python manage.py shell -c "exec(open('scripts/import_cbcs_course_structure.py').read()); import_cbcs()"
"""
from openpyxl import load_workbook


# Common typo fixes for department names
TYPO_FIXES = {
    'Hstory': 'History',
    'Phychology': 'Psychology',
    'Develovent': 'Development',
    'Oceanogaphy': 'Oceanography',
    'Calcules': 'Calculus',
    'Fundamentls': 'Fundamentals',
    'Fandamentals': 'Fundamentals',
}


def clean_name(name):
    """Clean and fix typos in names."""
    if not name:
        return None
    name = str(name).strip()
    for typo, fix in TYPO_FIXES.items():
        name = name.replace(typo, fix)
    return name


def import_cbcs(clear_existing=True):
    from university.models import University, Faculty, Department
    from ug.models import CourseStructure
    
    XLSX_PATH = '/Users/anuprash/Desktop/projects/pup-umis-backend/ug_courses_cbcs_system/CBCS REGISTRATION COURSE  2ND Semester.xlsx'
    SEMESTER = 2
    BATCH = '2025-29'
    
    print("="*70)
    print("IMPORTING CBCS COURSE STRUCTURE - SEMESTER II")
    print("="*70)
    
    # Get university
    university = University.objects.first()
    if not university:
        print("❌ No university found!")
        return
    print(f"University: {university.name}")
    
    # Clear existing data if requested
    if clear_existing:
        print("\n🗑️  Clearing existing data...")
        CourseStructure.objects.all().delete()
        Department.objects.all().delete()
        # Keep faculties that were manually created or from previous imports
        print("   Cleared CourseStructure and Department tables")
    
    # Load Excel
    wb = load_workbook(XLSX_PATH, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    print(f"Total rows: {len(rows)}")
    
    # Track created records
    faculties_created = 0
    departments_created = 0
    courses_created = 0
    
    current_faculty = None
    
    for i, row in enumerate(rows):
        # Skip empty rows
        if not any(row):
            continue
        
        # Check for Faculty header (e.g., "Faculty of Social Science - Semester-II")
        if row[0] and 'Faculty of' in str(row[0]):
            faculty_name = str(row[0]).strip().split(' - ')[0].strip()
            
            # Create or get Faculty
            faculty, created = Faculty.objects.get_or_create(
                name=faculty_name,
                defaults={
                    'university': university,
                    'short_name': ''.join([w[0] for w in faculty_name.split() if w[0].isupper()])
                }
            )
            current_faculty = faculty
            if created:
                faculties_created += 1
                print(f"\n✅ Created Faculty: {faculty_name}")
            else:
                print(f"\n📌 Using Faculty: {faculty_name}")
            continue
        
        # Skip header row
        if row[0] and 'Department & Subject Name' in str(row[0]):
            continue
        
        # Skip note rows and instruction rows
        if row[0] and 'Note' in str(row[0]):
            continue
        if row[3] and 'MJC-2 में' in str(row[3]):
            continue
        
        # Skip title rows
        if row[0] and 'List of subjects' in str(row[0]):
            continue
            
        if not current_faculty:
            continue
        
        # Process data rows
        # Columns: 
        # 0: Department & Subject Name (for MJC)
        # 1: MJC-2 (Major Core Course) name
        # 2: MIC-2 Subject Name (department for minor)
        # 3: MIC-2 (Minor Core Course) name
        # 4: SEC-2 (Skill Enhancement Course)
        # 5: VAC-2 (Value Added Course)
        # 6: MDC-2 Subject Name (department for MDC)
        # 7: MDC-2 (Multi-Disciplinary Course) name
        # 8: AEC-2 (Ability Enhancement Course)
        
        # Process MJC (Major Core Course)
        if row[0] and row[1]:
            dept_name = clean_name(row[0])
            course_name = clean_name(row[1])
            
            if dept_name and course_name:
                # Create department - just clean name, no code
                dept, created = Department.objects.get_or_create(
                    name=dept_name,
                    faculty=current_faculty
                )
                if created:
                    departments_created += 1
                
                # Create course structure
                cs, created = CourseStructure.objects.get_or_create(
                    name=course_name,
                    department=dept,
                    course_type='MJC',
                    code='MJC-2',
                    semester=SEMESTER,
                    defaults={
                        'label': 'Major Core Course',
                        'description': 'Major Core Course for Semester 2',
                        'json_data': {'batch': BATCH, 'faculty': current_faculty.name}
                    }
                )
                if created:
                    courses_created += 1
        
        # Process MIC (Minor Core Course)
        if row[2] and row[3]:
            dept_name = clean_name(row[2])
            course_name = clean_name(row[3])
            
            if dept_name and course_name and 'MJC-2 में' not in str(row[3]):
                dept, created = Department.objects.get_or_create(
                    name=dept_name,
                    faculty=current_faculty
                )
                if created:
                    departments_created += 1
                
                cs, created = CourseStructure.objects.get_or_create(
                    name=course_name,
                    department=dept,
                    course_type='MIC',
                    code='MIC-2',
                    semester=SEMESTER,
                    defaults={
                        'label': 'Minor Core Course',
                        'description': 'Minor Core Course for Semester 2',
                        'json_data': {'batch': BATCH, 'faculty': current_faculty.name}
                    }
                )
                if created:
                    courses_created += 1
        
        # Process SEC (Skill Enhancement Course)
        if row[4]:
            course_name = clean_name(row[4])
            if course_name:
                # SEC is common across faculty, use first department
                dept = Department.objects.filter(faculty=current_faculty).first()
                if dept:
                    cs, created = CourseStructure.objects.get_or_create(
                        name=course_name,
                        department=dept,
                        course_type='SEC',
                        code='SEC-2',
                        semester=SEMESTER,
                        defaults={
                            'label': 'Skill Enhancement Course',
                            'description': 'Skill Enhancement Course for Semester 2',
                            'json_data': {'batch': BATCH, 'faculty': current_faculty.name}
                        }
                    )
                    if created:
                        courses_created += 1
        
        # Process VAC (Value Added Course)
        if row[5]:
            course_name = clean_name(row[5])
            if course_name:
                dept = Department.objects.filter(faculty=current_faculty).first()
                if dept:
                    cs, created = CourseStructure.objects.get_or_create(
                        name=course_name,
                        department=dept,
                        course_type='VAC',
                        code='VAC-2',
                        semester=SEMESTER,
                        defaults={
                            'label': 'Value Added Course',
                            'description': 'Value Added Course for Semester 2',
                            'json_data': {'batch': BATCH, 'faculty': current_faculty.name}
                        }
                    )
                    if created:
                        courses_created += 1
        
        # Process MDC (Multi-Disciplinary Course)
        if row[6] and row[7]:
            dept_name = clean_name(row[6])
            course_name = clean_name(row[7])
            
            if dept_name and course_name and 'Note' not in str(row[6]):
                # For MDC, create department under current faculty
                dept, created = Department.objects.get_or_create(
                    name=dept_name,
                    faculty=current_faculty
                )
                if created:
                    departments_created += 1
                
                cs, created = CourseStructure.objects.get_or_create(
                    name=course_name,
                    department=dept,
                    course_type='MDC',
                    code='MDC-2',
                    semester=SEMESTER,
                    defaults={
                        'label': 'Multi-Disciplinary Course',
                        'description': 'Multi-Disciplinary Course for Semester 2',
                        'json_data': {'batch': BATCH, 'faculty': current_faculty.name}
                    }
                )
                if created:
                    courses_created += 1
        
        # Process AEC (Ability Enhancement Course)
        if row[8]:
            course_name = clean_name(row[8])
            if course_name:
                dept = Department.objects.filter(faculty=current_faculty).first()
                if dept:
                    cs, created = CourseStructure.objects.get_or_create(
                        name=course_name,
                        department=dept,
                        course_type='AEC',
                        code='AEC-2',
                        semester=SEMESTER,
                        defaults={
                            'label': 'Ability Enhancement Course - MIL',
                            'description': 'Ability Enhancement Course for Semester 2',
                            'json_data': {'batch': BATCH, 'faculty': current_faculty.name}
                        }
                    )
                    if created:
                        courses_created += 1
    
    wb.close()
    
    print("\n" + "="*70)
    print("IMPORT SUMMARY")
    print("="*70)
    print(f"   Faculties created: {faculties_created}")
    print(f"   Departments created: {departments_created}")
    print(f"   Course Structures created: {courses_created}")
    print(f"\n   Total Faculties: {Faculty.objects.count()}")
    print(f"   Total Departments: {Department.objects.count()}")
    print(f"   Total Course Structures: {CourseStructure.objects.count()}")


if __name__ == '__main__':
    import_cbcs()
