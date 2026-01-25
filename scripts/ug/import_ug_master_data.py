"""
Script to import UG Faculty, Department, Degree, Program, and CourseStructure data.
Reads all semester files from courses_data/ug/all_sem_courses/ folder.
IMPORTANT: Reads exact data from files - does NOT generate any data automatically.

Usage:
    poetry run python manage.py shell -c "exec(open('scripts/ug/import_ug_master_data.py').read()); import_ug_master_data()"
"""
import os
import re
from openpyxl import load_workbook


# Common typo fixes for names
TYPO_FIXES = {
    'Hstory': 'History',
    'Histoy': 'History',
    'Phychology': 'Psychology',
    'Develovent': 'Development',
    'Oceanogaphy': 'Oceanography',
    'Calcules': 'Calculus',
    'Fundamentls': 'Fundamentals',
    'Fandamentals': 'Fundamentals',
    'Genral': 'General',
}


def clean_name(name):
    """Clean and fix typos in names."""
    if not name:
        return None
    name = str(name).strip()
    for typo, fix in TYPO_FIXES.items():
        name = name.replace(typo, fix)
    return name if name else None


def extract_faculty_name(row_text):
    """Extract faculty name from row like 'Faculty of Social Science - Semester-I'"""
    if not row_text or 'Faculty of' not in str(row_text):
        return None
    faculty_name = str(row_text).strip().split(' - ')[0].strip()
    return faculty_name


def extract_semester_from_text(text):
    """Extract semester number from text."""
    if not text:
        return None
    text = str(text).upper()
    # Match patterns like "Semester-III", "Semester-I", "Semester-VIII"
    roman_numerals = {'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5, 'VI': 6, 'VII': 7, 'VIII': 8}
    match = re.search(r'SEMESTER[-\s]*([IVX]+)', text)
    if match:
        roman = match.group(1)
        return roman_numerals.get(roman)
    return None


def parse_header_column(header_text):
    """
    Parse header column to extract course type and code.
    Examples:
        'MJC-I (Major Core Course)' -> ('MJC', 'MJC-I', 'Major Core Course')
        'MJC-3 (Major Core Course)' -> ('MJC', 'MJC-3', 'Major Core Course')
        'SEC-3 (Skill Enhancement Course)' -> ('SEC', 'SEC-3', 'Skill Enhancement Course')
    """
    if not header_text:
        return None, None, None
    
    text = str(header_text).strip()
    
    # Extract course code like MJC-I, MJC-3, SEC-3, etc.
    code_match = re.match(r'^([A-Z]+)-?(\d+|[IVX]+)', text)
    if not code_match:
        return None, None, None
    
    course_type = code_match.group(1)  # MJC, MIC, SEC, VAC, MDC, AEC, INT, RP
    full_code = code_match.group(0)    # MJC-I, MJC-3, etc.
    
    # Extract description in parentheses
    desc_match = re.search(r'\(([^)]+)\)', text)
    description = desc_match.group(1).strip() if desc_match else None
    
    return course_type, full_code, description


def import_ug_master_data(clear_existing=False, courses_dir=None):
    """Import UG Faculty, Department, Degree, Program, and CourseStructure."""
    from university.models import University
    from ug.models import UGFaculty, UGDepartment, UGDegree, UGProgram, UGBatch, CourseStructure
    
    # Prompt for file path if not provided
    default_path = '/Users/anuprash/Desktop/projects/pup-umis-backend/courses_data/ug/all_sem_courses'
    if courses_dir is None:
        courses_dir = input(f"Enter UG courses folder path [{default_path}]: ").strip()
        if not courses_dir:
            courses_dir = default_path
    
    print("="*70)
    print("IMPORTING UG MASTER DATA")
    print("(Faculty, Department, Degree, Program, Batch, CourseStructure)")
    print("="*70)
    print(f"Courses folder: {courses_dir}")
    
    # Get university
    university = University.objects.first()
    if not university:
        print("❌ No university found! Please create one first.")
        return
    print(f"University: {university.name}")
    
    # Clear existing data if requested
    if clear_existing:
        print("\n🗑️  Clearing existing UG master data...")
        CourseStructure.objects.all().delete()
        UGBatch.objects.all().delete()
        UGProgram.objects.all().delete()
        UGDepartment.objects.all().delete()
        UGDegree.objects.all().delete()
        UGFaculty.objects.all().delete()
        print("   Cleared all UG master data tables")
    
    # Track created records
    faculties_created = set()
    departments_created = set()
    batches_created = set()
    courses_created = 0
    
    # Get all xlsx files
    xlsx_files = sorted([f for f in os.listdir(courses_dir) if f.endswith('.xlsx')])
    print(f"\nFound {len(xlsx_files)} xlsx files to process")
    
    for xlsx_file in xlsx_files:
        xlsx_path = os.path.join(courses_dir, xlsx_file)
        print(f"\n📄 Processing: {xlsx_file}")
        
        wb = load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        
        current_faculty = None
        current_semester = None
        current_batch = None  # Track batch for this file
        header_columns = []  # List of (col_index, course_type, course_code, description)
        file_courses = 0
        
        for i, row in enumerate(rows):
            # Skip empty rows
            if not any(row):
                continue
            
            first_cell = str(row[0]).strip() if row[0] else ''
            
            # Check for title row with batch info (e.g., "Four Year Under Graduate course 2025-29")
            if 'List of subjects' in first_cell or 'Under Graduate course' in first_cell:
                # Extract batch from title like "2025-29"
                batch_match = re.search(r'(\d{4}-\d{2,4})', first_cell)
                if batch_match:
                    batch_name = batch_match.group(1)
                    if batch_name not in batches_created:
                        current_batch, created = UGBatch.objects.get_or_create(
                            name=batch_name
                        )
                        if created:
                            batches_created.add(batch_name)
                            print(f"   📦 Created Batch: {batch_name}")
                    else:
                        current_batch = UGBatch.objects.filter(name=batch_name).first()
                continue
            
            # Check for Faculty header (also contains semester info)
            if 'Faculty of' in first_cell:
                faculty_name = extract_faculty_name(first_cell)
                semester = extract_semester_from_text(first_cell)
                
                if faculty_name:
                    # Create or get Faculty
                    faculty, created = UGFaculty.objects.get_or_create(
                        name=faculty_name,
                        defaults={
                            'university': university,
                        }
                    )
                    if created:
                        faculties_created.add(faculty_name)
                        print(f"   ✅ Created Faculty: {faculty_name}")
                    current_faculty = faculty
                
                if semester:
                    current_semester = semester
                    print(f"   📅 Semester: {current_semester}")
                continue
            
            # Check for header row (contains course type columns)
            if 'Department' in first_cell or 'Subject Name' in first_cell:
                header_columns = []
                for col_idx, cell in enumerate(row):
                    if col_idx == 0:
                        continue  # Skip first column (Department name)
                    if cell:
                        course_type, course_code, description = parse_header_column(cell)
                        if course_type and course_code:
                            header_columns.append({
                                'col_idx': col_idx,
                                'course_type': course_type,
                                'course_code': course_code,
                                'description': description
                            })
                print(f"   📋 Found {len(header_columns)} course columns: {[h['course_code'] for h in header_columns]}")
                continue
            
            # Skip note rows
            if 'Note' in first_cell:
                continue
            
            if not current_faculty or not current_semester or not header_columns:
                continue
            
            # Process data row
            dept_name = clean_name(row[0])
            if not dept_name or len(dept_name) < 2 or len(dept_name) > 100:
                continue
            
            # Create department
            dept_key = (dept_name, current_faculty.name)
            if dept_key not in departments_created:
                dept, created = UGDepartment.objects.get_or_create(
                    name=dept_name,
                    faculty=current_faculty
                )
                if created:
                    departments_created.add(dept_key)
            else:
                dept = UGDepartment.objects.filter(name=dept_name, faculty=current_faculty).first()
            
            if not dept:
                continue
            
            # Process each course column
            for header in header_columns:
                col_idx = header['col_idx']
                if col_idx >= len(row) or not row[col_idx]:
                    continue
                
                course_name = clean_name(row[col_idx])
                if not course_name or len(course_name) < 3:
                    continue
                
                # Skip if it looks like a department name in MDC column
                if header['course_type'] == 'MDC' and len(course_name) < 30:
                    # Check if next column has the actual course name
                    next_col = col_idx + 1
                    if next_col < len(row) and row[next_col]:
                        # This is MDC subject name, use it as department for MDC course
                        mdc_dept_name = course_name
                        mdc_course_name = clean_name(row[next_col])
                        
                        if mdc_course_name:
                            # Create MDC department if needed
                            mdc_dept_key = (mdc_dept_name, current_faculty.name)
                            if mdc_dept_key not in departments_created:
                                mdc_dept, created = UGDepartment.objects.get_or_create(
                                    name=mdc_dept_name,
                                    faculty=current_faculty
                                )
                                if created:
                                    departments_created.add(mdc_dept_key)
                            else:
                                mdc_dept = UGDepartment.objects.filter(name=mdc_dept_name, faculty=current_faculty).first()
                            
                            if mdc_dept:
                                # Find the MDC course header (next column)
                                for h in header_columns:
                                    if h['col_idx'] == next_col and h['course_type'] == 'MDC':
                                        cs, created = CourseStructure.objects.get_or_create(
                                            name=mdc_course_name,
                                            department=mdc_dept,
                                            code=h['course_code'],  # From file header
                                            semester=current_semester,
                                            defaults={
                                                'course_type': h['course_type'],
                                                # label will be set by another script (CIA Theory, ESE Practical, etc.)
                                                'batch': current_batch,
                                            }
                                        )
                                        if created:
                                            courses_created += 1
                                            file_courses += 1
                                        break
                        continue
                
                # Create course structure
                cs, created = CourseStructure.objects.get_or_create(
                    name=course_name,
                    department=dept,
                    code=header['course_code'],  # From file header
                    semester=current_semester,
                    defaults={
                        'course_type': header['course_type'],
                        # label will be set by another script (CIA Theory, ESE Practical, etc.)
                        'batch': current_batch,
                    }
                )
                if created:
                    courses_created += 1
                    file_courses += 1
        
        print(f"   📊 Courses created from this file: {file_courses}")
        wb.close()
    
    # Create UG Degrees (not in xlsx file, using standard values)
    print("\n📚 Creating UG Degrees...")
    ug_degrees_data = [
        {'name': 'Bachelor of Arts (Honours)', 'short_name': 'B.A. (Hons)', 'total_semesters': 8, 'total_years': 4},
        {'name': 'Bachelor of Science (Honours)', 'short_name': 'B.Sc. (Hons)', 'total_semesters': 8, 'total_years': 4},
        {'name': 'Bachelor of Commerce (Honours)', 'short_name': 'B.Com. (Hons)', 'total_semesters': 8, 'total_years': 4},
    ]
    
    degrees_created = 0
    for deg_data in ug_degrees_data:
        degree, created = UGDegree.objects.get_or_create(
            name=deg_data['name'],
            defaults={
                'short_name': deg_data['short_name'],
                'total_semesters': deg_data['total_semesters'],
                'total_years': deg_data['total_years']
            }
        )
        if created:
            degrees_created += 1
            print(f"   ✅ Created Degree: {deg_data['name']} ({deg_data['short_name']})")
    
    print("\n" + "="*70)
    print("IMPORT SUMMARY")
    print("="*70)
    print(f"   Faculties created: {len(faculties_created)}")
    print(f"   Departments created: {len(departments_created)}")
    print(f"   Batches created: {len(batches_created)}")
    print(f"   Degrees created: {degrees_created}")
    print(f"   CourseStructures created: {courses_created}")
    print(f"\n   Total UGFaculty: {UGFaculty.objects.count()}")
    print(f"   Total UGDepartment: {UGDepartment.objects.count()}")
    print(f"   Total UGBatch: {UGBatch.objects.count()}")
    print(f"   Total UGDegree: {UGDegree.objects.count()}")
    print(f"   Total UGProgram: {UGProgram.objects.count()}")
    print(f"   Total CourseStructure: {CourseStructure.objects.count()}")


if __name__ == '__main__':
    import_ug_master_data()
