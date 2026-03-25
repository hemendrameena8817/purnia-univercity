#!/usr/bin/env python
import argparse
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import django
from django.db import transaction
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[2]
sys.path.append(str(BASE_DIR))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'pup_umis_backend.settings')
django.setup()

from ug.models import CourseStructure, StudentCourseAssessment

DEFAULT_XLSX_PATH = BASE_DIR / 'scripts' / 'ug' / 'UPDATED COURSE CODES.xlsx'
TARGET_SEMESTER = '1ST'
TARGET_SESSION = '2025-26'
THEORY_LABELS = {'CIA-Theory', 'ESE-Theory'}
PRACTICAL_LABELS = {'CIA-Practical', 'ESE-Practical'}
ALL_LABELS = THEORY_LABELS | PRACTICAL_LABELS
CATEGORY_TO_DB_COURSE_CODE = {
    'MAJOR': 'MJC-1',
    'MINOR': 'MIC-1',
    'MULTIDISCIPLINARY': 'MDC-1',
    'MDC': 'MDC-1',
    'ABILITY ENHANCEMENT': 'AEC-1',
    'ABILITY ENHANCEMENT COURSE': 'AEC-1',
    'AEC': 'AEC-1',
    'SKILL ENHANCEMENT': 'SEC-1',
    'SKILL ENHANCEMENT COURSE': 'SEC-1',
    'SEC': 'SEC-1',
    'VALUE ADDED': 'VAC-1',
    'VALUE ADDED COURSE': 'VAC-1',
    'VAC': 'VAC-1',
}
NAME_ALIASES = {
    'understanding poltical theory': 'understanding political theory',
    'introduction to sociology -1': 'introduction to sociology - i',
    'introduction to sociology-1': 'introduction to sociology - i',
    'introduction to sociology-i': 'introduction to sociology - i',
    'introduction to sociology- i': 'introduction to sociology - i',
    'introduction to sociology -i': 'introduction to sociology - i',
    'introduction to sociology- i ': 'introduction to sociology - i',
    'introduction to socilogy -1': 'introduction to sociology - i',
    'introduction to socilogy-1': 'introduction to sociology - i',
    'decductive logic': 'deductive logic',
    'indian classical literaure': 'indian classical literature',
    'fandamentals of hrm': 'fundamentals of hrm',
}
HEADER_ALIASES = {
    'sl. no.': 'SL. NO.',
    'course codes': 'COURSE CODES',
    'name of courses': 'NAME OF COURSES',
    'credit': 'CREDIT',
    'type': 'TYPE',
    'course category': 'COURSE CATEGORY',
    'category': 'CATEGORY',
    'discipline': 'DISCIPLINE',
    'stream': 'STREAM',
}
REQUIRED_HEADERS = ('COURSE CODES', 'NAME OF COURSES', 'TYPE', 'CATEGORY')


def clean_text(value):
    text = str(value or '')
    text = text.replace('_x000D_', ' ').replace('\n', ' ').replace('\r', ' ')
    return ' '.join(text.split()).strip()


def normalize_header(value):
    text = clean_text(value).lower().replace('_', ' ')
    return HEADER_ALIASES.get(text, clean_text(value).upper())


def normalize_course_name(value):
    text = clean_text(value).lower()
    text = text.replace('`', "'").replace('’', "'").replace('–', '-')
    text = re.sub(r'\s*-\s*', ' - ', text)
    text = ' '.join(text.split())
    return NAME_ALIASES.get(text, text)


def normalize_category(value):
    text = clean_text(value).upper()
    return ' '.join(text.split())


def normalize_context_value(value):
    return clean_text(value).lower()


def get_components_from_type(type_value):
    text = clean_text(type_value).upper()
    if not text:
        return set()
    if 'THEORY' in text and 'PRACTICAL' in text:
        return {'THEORY', 'PRACTICAL'}
    if text == 'BOTH':
        return {'THEORY', 'PRACTICAL'}
    if 'THEORY' in text:
        return {'THEORY'}
    if 'PRACTICAL' in text:
        return {'PRACTICAL'}
    return set()


def get_component_from_label(label):
    label_value = clean_text(label)
    if label_value in THEORY_LABELS:
        return 'THEORY'
    if label_value in PRACTICAL_LABELS:
        return 'PRACTICAL'
    return None


def load_sheet_mapping(file_path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    raw_header = next(rows)
    headers = [normalize_header(value) for value in raw_header]
    index_map = {header: idx for idx, header in enumerate(headers) if header}

    missing_headers = [header for header in REQUIRED_HEADERS if header not in index_map]
    if missing_headers:
        raise ValueError(f'Missing required headers in sheet: {", ".join(missing_headers)}')

    mapping = {}
    conflicts = []
    sheet_course_names = set()
    stats = {
        'rows_read': 0,
        'rows_skipped_missing_course_name': 0,
        'rows_skipped_missing_new_code': 0,
        'rows_skipped_missing_category': 0,
        'rows_skipped_missing_type': 0,
        'rows_skipped_unknown_category': 0,
        'mapping_keys': 0,
        'conflicts': 0,
    }

    for excel_row_no, row in enumerate(rows, start=2):
        stats['rows_read'] += 1
        new_course_code = clean_text(row[index_map['COURSE CODES']])
        course_name = clean_text(row[index_map['NAME OF COURSES']])
        type_value = clean_text(row[index_map['TYPE']])
        category_value = clean_text(row[index_map['CATEGORY']])
        stream_value = normalize_context_value(row[index_map['STREAM']]) if 'STREAM' in index_map else ''
        discipline_value = normalize_context_value(row[index_map['DISCIPLINE']]) if 'DISCIPLINE' in index_map else ''

        if not course_name:
            stats['rows_skipped_missing_course_name'] += 1
            continue
        if not new_course_code:
            stats['rows_skipped_missing_new_code'] += 1
            continue
        if not category_value:
            stats['rows_skipped_missing_category'] += 1
            continue
        if not type_value:
            stats['rows_skipped_missing_type'] += 1
            continue

        db_course_code = CATEGORY_TO_DB_COURSE_CODE.get(normalize_category(category_value))
        if not db_course_code:
            stats['rows_skipped_unknown_category'] += 1
            continue

        components = get_components_from_type(type_value)
        if not components:
            stats['rows_skipped_missing_type'] += 1
            continue

        course_name_key = normalize_course_name(course_name)
        sheet_course_names.add(course_name_key)
        for component in components:
            key = (course_name_key, db_course_code, component)
            candidate = {
                'new_course_code': new_course_code,
                'stream': stream_value,
                'discipline': discipline_value,
                'excel_row_no': excel_row_no,
            }
            existing_candidates = mapping.setdefault(key, [])

            duplicate_candidate = next(
                (
                    item for item in existing_candidates
                    if item['new_course_code'] == new_course_code
                    and item['stream'] == stream_value
                    and item['discipline'] == discipline_value
                ),
                None,
            )
            if duplicate_candidate:
                continue

            conflicting_candidate = next(
                (
                    item for item in existing_candidates
                    if item['new_course_code'] != new_course_code
                    and item['stream'] == stream_value
                    and item['discipline'] == discipline_value
                ),
                None,
            )
            if conflicting_candidate:
                conflicts.append((excel_row_no, key, conflicting_candidate['new_course_code'], new_course_code))
                continue

            existing_candidates.append(candidate)

    stats['mapping_keys'] = len(mapping)
    stats['conflicts'] = len(conflicts)
    return mapping, conflicts, stats, sheet_course_names


def build_name_index(mapping):
    name_index = defaultdict(dict)
    for (course_name_key, db_course_code, component), candidates in mapping.items():
        name_index[course_name_key][(db_course_code, component)] = candidates
    return name_index


def get_row_context_text(row):
    parts = []
    department = getattr(row, 'department', None)
    if department and getattr(department, 'name', None):
        parts.append(department.name)
    degree = getattr(row, 'degree', None)
    if degree:
        parts.append(degree)
    return normalize_context_value(' '.join(parts))


def candidate_matches_context(candidate, stream_terms=None, discipline_terms=None):
    stream_terms = stream_terms or []
    discipline_terms = discipline_terms or []
    candidate_stream = candidate.get('stream', '') or ''
    candidate_discipline = candidate.get('discipline', '') or ''

    stream_ok = all(term in candidate_stream for term in stream_terms)
    discipline_ok = all(
        any(option in candidate_discipline for option in options)
        if isinstance(options, tuple)
        else options in candidate_discipline
        for options in discipline_terms
    )
    return stream_ok and discipline_ok


def resolve_special_case(row, course_name_key, detail_key, candidates):
    row_context = get_row_context_text(row)

    if course_name_key == 'fundamentals of the earth system':
        return None

    if course_name_key == 'algebra' and detail_key == ('MJC-1', 'THEORY'):
        if 'mathematics' in row_context:
            preferred = next(
                (candidate for candidate in candidates if candidate['new_course_code'] == 'U11013'),
                None,
            )
            if preferred:
                return preferred['new_course_code']

    if 'commerce hrm' in row_context or ('commerce' in row_context and 'hrm' in row_context):
        matches = [
            candidate for candidate in candidates
            if candidate_matches_context(
                candidate,
                stream_terms=['commerce'],
                discipline_terms=[('human resource', 'hrm')],
            )
        ]
        if len({candidate['new_course_code'] for candidate in matches}) == 1:
            return matches[0]['new_course_code']

    if 'commerce marketing' in row_context or ('commerce' in row_context and 'marketing' in row_context):
        matches = [
            candidate for candidate in candidates
            if candidate_matches_context(
                candidate,
                stream_terms=['commerce'],
                discipline_terms=['marketing'],
            )
        ]
        if len({candidate['new_course_code'] for candidate in matches}) == 1:
            return matches[0]['new_course_code']

    return None


def score_candidate_for_context(candidate, row_context):
    candidate_context = ' '.join(
        part for part in [candidate.get('stream'), candidate.get('discipline')] if part
    ).strip()
    if not candidate_context or not row_context:
        return 0

    score = 0
    if candidate_context in row_context:
        score += 100
    if row_context in candidate_context:
        score += 25

    for token in sorted(set(candidate_context.split()), key=len, reverse=True):
        if token and token in row_context:
            score += len(token)

    return score


def resolve_candidate_by_context(candidates, row):
    unique_codes = {candidate['new_course_code'] for candidate in candidates}
    if len(unique_codes) == 1:
        return next(iter(unique_codes))

    course_name_key = normalize_course_name(row.course_name)
    detail_key = (clean_text(row.course_code), get_component_from_label(row.label))
    special_case_code = resolve_special_case(row, course_name_key, detail_key, candidates)
    if special_case_code:
        return special_case_code

    row_context = get_row_context_text(row)
    scored_candidates = []
    for candidate in candidates:
        score = score_candidate_for_context(candidate, row_context)
        scored_candidates.append((score, candidate))

    scored_candidates.sort(
        key=lambda item: (
            item[0],
            len(item[1].get('stream', '')),
            len(item[1].get('discipline', '')),
        ),
        reverse=True,
    )

    best_score, best_candidate = scored_candidates[0]
    if best_score <= 0:
        return None

    best_codes = {
        candidate['new_course_code']
        for score, candidate in scored_candidates
        if score == best_score
    }
    if len(best_codes) != 1:
        return None

    return best_candidate['new_course_code']


def resolve_new_course_code(row, name_index):
    course_name_key = normalize_course_name(row.course_name)
    name_matches = name_index.get(course_name_key)
    if not name_matches:
        return None, 'course_name_not_matched'

    component = get_component_from_label(row.label)
    if not component:
        return None, 'course_name_matched_but_details_failed'
    detail_key = (clean_text(row.course_code), component)
    candidates = name_matches.get(detail_key)
    if not candidates:
        return None, 'course_name_matched_but_details_failed'

    new_course_code = resolve_candidate_by_context(candidates, row)
    if not new_course_code:
        return None, 'course_name_matched_but_details_failed'

    return new_course_code, None


def update_course_structures(name_index, dry_run=False):
    queryset = CourseStructure.objects.filter(
        semester=TARGET_SEMESTER,
        label__in=ALL_LABELS,
    ).select_related('department').only('id', 'course_name', 'course_code', 'label', 'new_course_code', 'department__name')

    to_update = []
    unmatched_course_names = []
    detail_failed_rows = []
    stats = {
        'processed': 0,
        'matched': 0,
        'updated': 0,
        'already_set': 0,
        'course_name_not_matched': 0,
        'course_name_matched_but_details_failed': 0,
    }

    for row in queryset.iterator(chunk_size=500):
        stats['processed'] += 1
        new_code, failure_reason = resolve_new_course_code(row, name_index)
        if not new_code:
            stats[failure_reason] += 1
            if failure_reason == 'course_name_not_matched':
                unmatched_course_names.append(
                    f"CourseStructure#{row.id} | course_name={clean_text(row.course_name) or '-'} | course_code={clean_text(row.course_code) or '-'} | department={clean_text(getattr(row.department, 'name', '')) or '-'} | label={clean_text(row.label) or '-'} | component={get_component_from_label(row.label) or '-'}"
                )
            elif failure_reason == 'course_name_matched_but_details_failed':
                detail_failed_rows.append(
                    f"CourseStructure#{row.id} | course_name={clean_text(row.course_name) or '-'} | course_code={clean_text(row.course_code) or '-'} | department={clean_text(getattr(row.department, 'name', '')) or '-'} | label={clean_text(row.label) or '-'} | component={get_component_from_label(row.label) or '-'}"
                )
            continue

        stats['matched'] += 1
        if row.new_course_code == new_code:
            stats['already_set'] += 1
            continue

        row.new_course_code = new_code
        to_update.append(row)
        stats['updated'] += 1

    if to_update and not dry_run:
        CourseStructure.objects.bulk_update(to_update, ['new_course_code'], batch_size=500)

    return stats, unmatched_course_names, detail_failed_rows


def update_student_assessments(name_index, dry_run=False):
    queryset = StudentCourseAssessment.objects.filter(
        semester=TARGET_SEMESTER,
        session=TARGET_SESSION,
        label__in=ALL_LABELS,
    ).select_related('department').only('id', 'course_name', 'course_code', 'label', 'new_course_code', 'degree', 'department__name')

    to_update = []
    unmatched_course_names = []
    stats = {
        'processed': 0,
        'matched': 0,
        'updated': 0,
        'already_set': 0,
        'course_name_not_matched': 0,
        'course_name_matched_but_details_failed': 0,
    }

    for row in queryset.iterator(chunk_size=500):
        stats['processed'] += 1
        new_code, failure_reason = resolve_new_course_code(row, name_index)
        if not new_code:
            stats[failure_reason] += 1
            if failure_reason == 'course_name_not_matched':
                unmatched_course_names.append(
                    f"StudentCourseAssessment#{row.id} | course_name={clean_text(row.course_name) or '-'} | course_code={clean_text(row.course_code) or '-'} | department={clean_text(getattr(row.department, 'name', '')) or '-'} | degree={clean_text(row.degree) or '-'} | label={clean_text(row.label) or '-'} | component={get_component_from_label(row.label) or '-'}"
                )
            elif failure_reason == 'course_name_matched_but_details_failed':
                detail_failed_rows.append(
                    f"StudentCourseAssessment#{row.id} | course_name={clean_text(row.course_name) or '-'} | course_code={clean_text(row.course_code) or '-'} | department={clean_text(getattr(row.department, 'name', '')) or '-'} | degree={clean_text(row.degree) or '-'} | label={clean_text(row.label) or '-'} | component={get_component_from_label(row.label) or '-'}"
                )
            continue

        stats['matched'] += 1
        if row.new_course_code == new_code:
            stats['already_set'] += 1
            continue

        row.new_course_code = new_code
        to_update.append(row)
        stats['updated'] += 1

    if to_update and not dry_run:
        StudentCourseAssessment.objects.bulk_update(to_update, ['new_course_code'], batch_size=500)

    return stats, unmatched_course_names, detail_failed_rows


def print_stats(title, stats):
    print('\n' + '=' * 100)
    print(title)
    print('=' * 100)
    for key, value in stats.items():
        print(f'{key.replace("_", " ").title():<35} {value:,}')


def print_unmatched_course_names(title, rows):
    if not rows:
        return

    print('\n' + '=' * 100)
    print(title)
    print('=' * 100)
    for row in rows:
        print(row)


def get_db_course_name_sets():
    course_structure_names = {
        normalize_course_name(name)
        for name in CourseStructure.objects.filter(
            semester=TARGET_SEMESTER,
            label__in=ALL_LABELS,
        ).values_list('course_name', flat=True)
        if clean_text(name)
    }

    assessment_names = {
        normalize_course_name(name)
        for name in StudentCourseAssessment.objects.filter(
            semester=TARGET_SEMESTER,
            session=TARGET_SESSION,
            label__in=ALL_LABELS,
        ).values_list('course_name', flat=True)
        if clean_text(name)
    }

    return course_structure_names, assessment_names


def print_name_set_difference(title, names):
    if not names:
        return

    print('\n' + '=' * 100)
    print(title)
    print('=' * 100)
    for name in sorted(names):
        print(name)


def run(file_path, dry_run=False):
    mapping, conflicts, sheet_stats, sheet_course_names = load_sheet_mapping(file_path)
    name_index = build_name_index(mapping)
    course_structure_db_names, assessment_db_names = get_db_course_name_sets()
    print_stats('SHEET SUMMARY', sheet_stats)

    if conflicts:
        print('\nConflicts found in sheet mapping:')
        for excel_row_no, key, old_value, new_value in conflicts[:20]:
            print(f'row={excel_row_no} | key={key} | existing={old_value} | new={new_value}')
        raise SystemExit('Resolve sheet conflicts before updating.')

    if dry_run:
        course_structure_stats, course_structure_unmatched, course_structure_detail_failed = update_course_structures(name_index, dry_run=True)
        assessment_stats, assessment_unmatched, assessment_detail_failed = update_student_assessments(name_index, dry_run=True)
        print_stats('COURSE STRUCTURE DRY RUN', course_structure_stats)
        print_stats('STUDENT ASSESSMENT DRY RUN', assessment_stats)
        print_name_set_difference('COURSE NAMES IN SHEET BUT NOT FOUND IN COURSE STRUCTURE DB', sheet_course_names - course_structure_db_names)
        print_name_set_difference('COURSE NAMES IN SHEET BUT NOT FOUND IN STUDENT ASSESSMENT DB', sheet_course_names - assessment_db_names)
        print_name_set_difference('COURSE NAMES IN COURSE STRUCTURE DB BUT NOT FOUND IN SHEET', course_structure_db_names - sheet_course_names)
        print_name_set_difference('COURSE NAMES IN STUDENT ASSESSMENT DB BUT NOT FOUND IN SHEET', assessment_db_names - sheet_course_names)
        print_unmatched_course_names('COURSE STRUCTURE COURSE NAME NOT FOUND IN THE SHEET', course_structure_unmatched)
        print_unmatched_course_names('COURSE STRUCTURE COURSE NAME MATCHED BUT DETAILS FAILED', course_structure_detail_failed)
        print_unmatched_course_names('STUDENT ASSESSMENT COURSE NAME NOT FOUND IN THE SHEET', assessment_unmatched)
        print_unmatched_course_names('STUDENT ASSESSMENT COURSE NAME MATCHED BUT DETAILS FAILED', assessment_detail_failed)
        print('\nDRY RUN: no changes saved')
        return

    with transaction.atomic():
        course_structure_stats, course_structure_unmatched, course_structure_detail_failed = update_course_structures(name_index, dry_run=False)
        assessment_stats, assessment_unmatched, assessment_detail_failed = update_student_assessments(name_index, dry_run=False)

    print_stats('COURSE STRUCTURE UPDATE', course_structure_stats)
    print_stats('STUDENT ASSESSMENT UPDATE', assessment_stats)
    print_name_set_difference('COURSE NAMES IN SHEET BUT NOT FOUND IN COURSE STRUCTURE DB', sheet_course_names - course_structure_db_names)
    print_name_set_difference('COURSE NAMES IN SHEET BUT NOT FOUND IN STUDENT ASSESSMENT DB', sheet_course_names - assessment_db_names)
    print_name_set_difference('COURSE NAMES IN COURSE STRUCTURE DB BUT NOT FOUND IN SHEET', course_structure_db_names - sheet_course_names)
    print_name_set_difference('COURSE NAMES IN STUDENT ASSESSMENT DB BUT NOT FOUND IN SHEET', assessment_db_names - sheet_course_names)
    print_unmatched_course_names('COURSE STRUCTURE COURSE NAME NOT FOUND IN THE SHEET', course_structure_unmatched)
    print_unmatched_course_names('COURSE STRUCTURE COURSE NAME MATCHED BUT DETAILS FAILED', course_structure_detail_failed)
    print_unmatched_course_names('STUDENT ASSESSMENT COURSE NAME NOT FOUND IN THE SHEET', assessment_unmatched)
    print_unmatched_course_names('STUDENT ASSESSMENT COURSE NAME MATCHED BUT DETAILS FAILED', assessment_detail_failed)


def main():
    parser = argparse.ArgumentParser(description='Update new_course_code from UPDATED COURSE CODES.xlsx for 1ST sem course structures and 1ST sem 2025-26 assessments.')
    parser.add_argument('--file-path', type=str, default=str(DEFAULT_XLSX_PATH), help='Path to UPDATED COURSE CODES.xlsx')
    parser.add_argument('--dry-run', action='store_true', help='Preview matches without saving')
    args = parser.parse_args()
    run(Path(args.file_path), dry_run=args.dry_run)


if __name__ == '__main__':
    main()
