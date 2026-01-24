"""
Import script for applicant_qualification_detail.xlsx
Run from Django shell:
    exec(open('scripts/import_stagingapplicantqualificationdetail.py').read())
"""
import os
from openpyxl import load_workbook
from staging.models import StagingApplicantQualificationDetail

XLSX_PATH = 'old_data/applicant_qualification_detail.xlsx'

def import_data():
    if not os.path.exists(XLSX_PATH):
        print(f"Error: XLSX file not found at {XLSX_PATH}")
        return
    
    # Delete all existing data first
    existing_count = StagingApplicantQualificationDetail.objects.count()
    print(f"🗑️  Deleting all {existing_count} existing records...")
    StagingApplicantQualificationDetail.objects.all().delete()
    print(f"✅ All existing records deleted!")
    
    wb = load_workbook(XLSX_PATH, read_only=True)
    ws = wb.active
    row_iterator = ws.iter_rows(values_only=True)
    
    # Skip header
    try:
        header = next(row_iterator)
    except StopIteration:
        print("Empty file!")
        return
        
    print(f"Starting import...")
    
    imported = 0
    skipped = 0
    errors = []
    
    # Track csv_ids to avoid duplicates within the same file
    seen_csv_ids = set()
    
    for i, row in enumerate(row_iterator):
        try:
            csv_id = str(row[9]) if row[9] is not None else None
            
            # Skip if csv_id is None or already seen (duplicate in file)
            if csv_id is None:
                skipped += 1
                continue
            
            if csv_id in seen_csv_ids:
                skipped += 1
                continue
            
            seen_csv_ids.add(csv_id)
            
            StagingApplicantQualificationDetail.objects.create(
                    applied_class=str(row[0]) if row[0] is not None else None,
                    applied_program=str(row[1]) if row[1] is not None else None,
                    created_by=str(row[2]) if row[2] is not None else None,
                    created_on=str(row[3]) if row[3] is not None else None,
                    division_distinction=str(row[4]) if row[4] is not None else None,
                    exam_code=str(row[5]) if row[5] is not None else None,
                    full_mark=str(row[6]) if row[6] is not None else None,
                    grade=str(row[7]) if row[7] is not None else None,
                    grade_mark_flag=str(row[8]) if row[8] is not None else None,
                    csv_id=csv_id,
                    institute_code=str(row[10]) if row[10] is not None else None,
                    institute_name=str(row[11]) if row[11] is not None else None,
                    last_updated=str(row[12]) if row[12] is not None else None,
                    mark_secured=str(row[13]) if row[13] is not None else None,
                    math_grade=str(row[14]) if row[14] is not None else None,
                    math_mark=str(row[15]) if row[15] is not None else None,
                    percentage_mark=str(row[16]) if row[16] is not None else None,
                    qual_desc_1=str(row[17]) if row[17] is not None else None,
                    qual_desc_2=str(row[18]) if row[18] is not None else None,
                    reg_user_id=str(row[19]) if row[19] is not None else None,
                    roll_no=str(row[20]) if row[20] is not None else None,
                    status=str(row[21]) if row[21] is not None else None,
                    subjects_offered=str(row[22]) if row[22] is not None else None,
                    university_board=str(row[23]) if row[23] is not None else None,
                    updated_by=str(row[24]) if row[24] is not None else None,
                    updated_on=str(row[25]) if row[25] is not None else None,
                    year_of_passing=str(row[26]) if row[26] is not None else None,
            )
            imported += 1
            if imported % 1000 == 0:
                print(f"Imported {imported} records...")
        except Exception as e:
            errors.append(f"Row {i + 2}: {str(e)}")
    
    print(f"\n✅ Import completed!")
    print(f"   Imported: {imported} records")
    print(f"   Skipped (duplicates): {skipped} records")
    print(f"   Errors: {len(errors)}")
    if errors[:5]:
        print("   First errors:", errors[:5])
    print(f"   Total in table: {StagingApplicantQualificationDetail.objects.count()}")

if __name__ == '__main__':
    import_data()
