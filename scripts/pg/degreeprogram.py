# Usage: /home/gaurav/.pyenv/versions/pup-umis/bin/python scripts/pg/degreeprogram.py
import os
import sys
import django
import openpyxl
from django.conf import settings

def setup_django():
    # Add the project root to sys.path
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    sys.path.append(project_root)
    
    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "pup_umis_backend.settings")
    django.setup()

def run():
    setup_django()
    
    from pg.models import PGDegree, PGProgram, PGDepartment
    
    # File path
    file_path = os.path.join(settings.BASE_DIR, 'courses_data/pg/degreeandprogram.xlsx')
    
    print(f"Reading file: {file_path}")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
    except FileNotFoundError:
        print(f"Error: File not found at {file_path}")
        return
    except Exception as e:
        print(f"Error opening Excel file: {e}")
        return

    # Skip Header (Row 1)
    # Row 2 onwards
    max_row = sheet.max_row
    print(f"Processing rows 2 to {max_row}...")

    for row_idx in range(2, max_row + 1):
        # Column 1: Degree Short Name
        degree_short = sheet.cell(row=row_idx, column=1).value
        # Column 2: Degree Name
        degree_name = sheet.cell(row=row_idx, column=2).value
        # Column 3: Program Short Name
        prog_short = sheet.cell(row=row_idx, column=3).value
        # Column 4: Program Name
        prog_name = sheet.cell(row=row_idx, column=4).value
        
        if not degree_name or not prog_name:
            continue

        degree_name = str(degree_name).strip()
        degree_short = str(degree_short).strip() if degree_short else ''
        prog_name = str(prog_name).strip()
        prog_short = str(prog_short).strip() if prog_short else ''

        # 1. Create/Get Degree
        degree, d_created = PGDegree.objects.get_or_create(
            name=degree_name,
            defaults={
                'short_name': degree_short
            }
        )
        if d_created:
            print(f"  Created Degree: {degree_name}")


        # 2. Extract Department & Create/Get Program
        # Extract Department Name from Program Name (last part after ' in ')
        # e.g. "Master of Arts in Economics" -> "Economics"
        department = None
        if " in " in prog_name:
            dept_name_candidate = prog_name.split(" in ")[-1].strip()
        elif prog_name.startswith("Master of "):
            # Fallback for "Master of Commerce" -> "Commerce"
            dept_name_candidate = prog_name.replace("Master of ", "").strip()
        else:
            dept_name_candidate = None
            
        if dept_name_candidate:
            # Find PGDepartment
            department = PGDepartment.objects.filter(name__iexact=dept_name_candidate).first()
            if not department:
                 # Try with brackets handling if needed, or exact match variations
                 print(f"    WARNING: Department '{dept_name_candidate}' not found for program '{prog_name}'")

        # Create/Get Program
        program, p_created = PGProgram.objects.get_or_create(
            name=prog_name,
            defaults={
                'short_name': prog_short,
                'degree': degree,
                'department': department
            }
        )
        
        if p_created:
            print(f"    Created Program: {prog_name}")
        else:
            # Update if exists to ensure links are correct
            updated = False
            if program.degree != degree:
                program.degree = degree
                updated = True
            if program.department != department:
                program.department = department
                updated = True
            if program.short_name != prog_short:
                program.short_name = prog_short
                updated = True
                
            if updated:
                program.save()
                print(f"    Updated Program: {prog_name}")

    print("Done.")

if __name__ == "__main__":
    run()
