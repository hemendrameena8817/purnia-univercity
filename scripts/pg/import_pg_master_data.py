"""
Script to import PG Faculty, Department, Degree, Program, and PGCourseStructure data.
Reads from courses_data/pg/ folder - specifically the Subject Detail_PG.xlsx file.
IMPORTANT: Reads exact data from files - does NOT generate any data automatically.

Usage:
    poetry run python manage.py shell -c "exec(open('scripts/pg/import_pg_master_data.py').read()); import_pg_master_data()"
"""
import os
import re
from openpyxl import load_workbook


def clean_name(name):
    """Clean name."""
    if not name:
        return None
    name = str(name).strip()
    # Remove newlines
    name = name.replace('\n', ' ').replace('\r', '')
    return name if name else None


def get_semester_from_sheet_name(sheet_name):
    """Extract semester number from sheet name."""
    name_upper = sheet_name.upper()
    # Check for IV first (before checking for I, II, III)
    if 'IV SEMESTER' in name_upper or '4TH SEMESTER' in name_upper:
        return 4
    elif 'III SEMESTER' in name_upper or '3RD SEMESTER' in name_upper:
        return 3
    elif 'II SEMESTER' in name_upper or '2ND SEMESTER' in name_upper:
        return 2
    elif 'I SEMESTER' in name_upper or '1ST SEMESTER' in name_upper:
        return 1
    return None


def import_pg_master_data(clear_existing=False, courses_dir=None):
    """Import PG Faculty, Department, Degree, Program, and PGCourseStructure."""
    from university.models import University
    from pg.models import PGFaculty, PGDepartment, PGDegree, PGProgram, PGBatch, PGCourseStructure
    
    # Prompt for file path if not provided
    default_path = '/Users/anuprash/Desktop/projects/pup-umis-backend/courses_data/pg'
    if courses_dir is None:
        courses_dir = input(f"Enter PG courses folder path [{default_path}]: ").strip()
        if not courses_dir:
            courses_dir = default_path
    
    print("="*70)
    print("IMPORTING PG MASTER DATA")
    print("(Faculty, Department, Degree, Program, Batch, PGCourseStructure)")
    print("="*70)
    print(f"Courses folder: {courses_dir}")
    
    # Get university
    university = University.objects.first()
    if not university:
        print("❌ No university found! Please create one first.")
        return
    print(f"University: {university.name}")
    
    # Clear existing data if requested
    if clear_existing:
        print("\n🗑️  Clearing existing PG master data...")
        PGCourseStructure.objects.all().delete()
        PGBatch.objects.all().delete()
        PGProgram.objects.all().delete()
        PGDepartment.objects.all().delete()
        PGDegree.objects.all().delete()
        PGFaculty.objects.all().delete()
        print("   Cleared all PG master data tables")
    
    # Track created records
    departments_created = set()
    batches_created = set()
    courses_created = 0
    
    # Get all xlsx files
    xlsx_files = [f for f in os.listdir(courses_dir) if f.endswith('.xlsx')]
    print(f"\nFound {len(xlsx_files)} xlsx files to process")
    
    # Create default PG Faculty
    print("\n📁 Creating PG Faculty...")
    pg_faculty, created = PGFaculty.objects.get_or_create(
        name='Postgraduate Studies',
        defaults={
            'university': university,
        }
    )
    if created:
        print(f"   ✅ Created Faculty: Postgraduate Studies")
    
    # Create PG Degrees (not in xlsx file, using standard values)
    print("\n📚 Creating PG Degrees...")
    pg_degrees_data = [
        {'name': 'Master of Arts', 'short_name': 'M.A.', 'total_semesters': 4, 'total_years': 2},
        {'name': 'Master of Science', 'short_name': 'M.Sc.', 'total_semesters': 4, 'total_years': 2},
        {'name': 'Master of Commerce', 'short_name': 'M.Com.', 'total_semesters': 4, 'total_years': 2},
    ]
    
    degrees_created = 0
    for deg_data in pg_degrees_data:
        degree, created = PGDegree.objects.get_or_create(
            name=deg_data['name'],
            defaults={
                'short_name': deg_data['short_name'],
                'total_semesters': deg_data['total_semesters'],
                'total_years': deg_data['total_years']
            }
        )
        if created:
            degrees_created += 1
            print(f"   ✅ Created Degree: {deg_data['name']} ({deg_data['short_name']})")
    
    for xlsx_file in xlsx_files:
        xlsx_path = os.path.join(courses_dir, xlsx_file)
        print(f"\n📄 Processing: {xlsx_file}")
        
        wb = load_workbook(xlsx_path, read_only=True)
        
        # Track which semesters have been processed to avoid duplicates
        processed_semesters = set()
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            # Only process semester sheets
            semester = get_semester_from_sheet_name(sheet_name)
            if semester is None:
                print(f"   ⏭️  Skipping sheet: {sheet_name} (not a semester sheet)")
                continue
            
            # Skip if this semester was already processed (avoid duplicates like "PG I SEMESTER" and "PG I SEMESTER (2)")
            if semester in processed_semesters:
                print(f"   ⏭️  Skipping sheet: {sheet_name} (Semester {semester} already processed)")
                continue
            
            processed_semesters.add(semester)
            
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            
            print(f"\n   📋 Sheet: {sheet_name} (Semester {semester}) - {len(rows)} rows")
            
            # Extract batch from first row (e.g., "PG-Ist Semester December,2025" or "PG 2nd SEMESTER JUNE-2022")
            current_batch = None
            if rows and rows[0]:
                first_row_text = str(rows[0][0] or rows[0][1] or '').strip()
                # Try to extract batch like "December,2025" or "JUNE-2022" or just year
                batch_match = re.search(r'(\w+[,\-]?\s*\d{4})', first_row_text)
                if batch_match:
                    batch_name = batch_match.group(1).strip()
                    if batch_name not in batches_created:
                        current_batch, created = PGBatch.objects.get_or_create(
                            name=batch_name
                        )
                        if created:
                            batches_created.add(batch_name)
                            print(f"      📦 Created Batch: {batch_name}")
                    else:
                        current_batch = PGBatch.objects.filter(name=batch_name).first()
            
            # Detect column structure from header row
            # Different sheets have different structures
            subject_col = None
            paper_col = None
            paper_name_col = None
            paper_code_col = None
            max_marks_col = None
            min_marks_col = None
            credit_col = None
            course_col = None  # For "Course" column like "M.Sc."
            
            # Find header row and column positions
            header_row_idx = None
            for i, row in enumerate(rows[:5]):
                if not row:
                    continue
                for j, cell in enumerate(row):
                    if cell:
                        cell_str = str(cell).lower().strip()
                        if cell_str == 'subject' or cell_str == 'subjects':
                            subject_col = j
                            header_row_idx = i
                        elif cell_str == 'course':
                            course_col = j
                            header_row_idx = i
                        elif 'paper name' in cell_str or cell_str == 'title' or cell_str == 'papers':
                            paper_name_col = j
                        elif 'paper code' in cell_str or 'course code' in cell_str:
                            paper_code_col = j
                        elif cell_str == 'paper' or cell_str == 'paper ':
                            paper_col = j
                        elif 'max' in cell_str and 'mark' in cell_str:
                            max_marks_col = j
                        elif 'min' in cell_str and 'mark' in cell_str:
                            min_marks_col = j
                        elif 'credit' in cell_str:
                            credit_col = j
            
            # Log detected columns
            print(f"      Columns - Subject: {subject_col}, Course: {course_col}, Paper: {paper_col}, Name: {paper_name_col}, Code: {paper_code_col}")
            
            if header_row_idx is None:
                print(f"      ⚠️  Could not find header row, skipping...")
                continue
            
            # Determine data start row (after header and any sub-header rows)
            data_start = header_row_idx + 1
            if data_start < len(rows):
                # Check if next row is a sub-header (like ESE/CIA row)
                next_row = rows[data_start]
                if next_row and all(cell is None or str(cell).strip() in ['ESE', 'CIA', ''] for cell in next_row[:10]):
                    data_start += 1
            
            # Process data rows
            current_subject = None
            sheet_courses = 0
            
            for row_idx, row in enumerate(rows[data_start:], start=data_start):
                if not any(row):
                    continue
                
                # Get subject (department)
                # Check both subject_col and course_col for subject name
                subject_name = None
                if subject_col is not None and len(row) > subject_col and row[subject_col]:
                    val = clean_name(row[subject_col])
                    # Skip if it's a header or note
                    if val and len(val) > 1 and 'Subject' not in val and 'PG' not in val:
                        subject_name = val
                elif course_col is not None and len(row) > course_col and row[course_col]:
                    # For sheets where Course column has the degree type (M.Sc., M.A.)
                    pass  # We get subject from subject_col
                
                # For PG II SEMESTER, subject is in column 1 (after Course column)
                if subject_name is None and subject_col is None and len(row) > 1 and row[1]:
                    val = clean_name(row[1])
                    if val and len(val) > 1 and 'Subject' not in val:
                        subject_name = val
                
                # Update current subject only if we got a new one
                if subject_name:
                    current_subject = subject_name
                
                if not current_subject:
                    continue
                
                # Get paper code
                paper_code = None
                if paper_code_col is not None and len(row) > paper_code_col:
                    paper_code = clean_name(row[paper_code_col])
                elif paper_col is not None and len(row) > paper_col:
                    paper_code = clean_name(row[paper_col])
                
                # Get paper name
                paper_name = None
                if paper_name_col is not None and len(row) > paper_name_col:
                    paper_name = clean_name(row[paper_name_col])
                
                if not paper_name:
                    continue
                
                # Skip header-like rows
                if 'Paper Name' in paper_name or 'Title' in paper_name or 'Papers' == paper_name:
                    continue
                
                # Create department
                if current_subject not in departments_created:
                    dept, created = PGDepartment.objects.get_or_create(
                        name=current_subject,
                        faculty=pg_faculty
                    )
                    if created:
                        departments_created.add(current_subject)
                        print(f"      ✅ Created Department: {current_subject}")
                else:
                    dept = PGDepartment.objects.filter(name=current_subject, faculty=pg_faculty).first()
                
                if not dept:
                    continue
                
                # Determine course type from paper code (read exactly from file)
                course_type = None
                if paper_code:
                    code_upper = str(paper_code).upper()
                    # Extract course type prefix
                    match = re.match(r'^([A-Z]+)', code_upper)
                    if match:
                        course_type = match.group(1)
                        # Normalize common types
                        if course_type in ['AECC', 'AEC']:
                            course_type = 'AECC'
                        elif course_type == 'GE':
                            course_type = 'GE'
                        elif course_type == 'EC':
                            course_type = 'EC'
                        elif course_type == 'SEC':
                            course_type = 'SEC'
                        elif course_type == 'CC':
                            course_type = 'CC'
                
                # Get marks info if available
                max_marks = None
                min_marks = None
                max_credit = None
                
                if max_marks_col is not None and len(row) > max_marks_col:
                    try:
                        val = row[max_marks_col]
                        if val is not None:
                            max_marks = int(float(val))
                    except (ValueError, TypeError):
                        pass
                
                if min_marks_col is not None and len(row) > min_marks_col:
                    try:
                        val = row[min_marks_col]
                        if val is not None:
                            min_marks = float(val)
                    except (ValueError, TypeError):
                        pass
                
                if credit_col is not None and len(row) > credit_col:
                    try:
                        val = row[credit_col]
                        if val is not None and isinstance(val, (int, float)) and val <= 10:
                            max_credit = int(val)
                    except (ValueError, TypeError):
                        pass
                
                # Create PGCourseStructure
                # code is from file, label will be set by another script (CIA Theory, ESE Practical, etc.)
                cs, created = PGCourseStructure.objects.get_or_create(
                    name=paper_name,
                    department=dept,
                    code=paper_code,  # From file (can be None if not in file)
                    semester=semester,
                    defaults={
                        'course_type': course_type,
                        # label will be set by another script (CIA Theory, ESE Practical, etc.)
                        'max_marks': max_marks,
                        'min_mark': int(min_marks) if min_marks else None,
                        'max_credit': max_credit,
                        'batch': current_batch,
                    }
                )
                if created:
                    courses_created += 1
                    sheet_courses += 1
            
            print(f"      📊 Courses created from this sheet: {sheet_courses}")
        
        wb.close()
    
    print("\n" + "="*70)
    print("IMPORT SUMMARY")
    print("="*70)
    print(f"   Faculties created: {1 if pg_faculty else 0}")
    print(f"   Departments created: {len(departments_created)}")
    print(f"   Batches created: {len(batches_created)}")
    print(f"   Degrees created: {degrees_created}")
    print(f"   PGCourseStructures created: {courses_created}")
    print(f"\n   Total PGFaculty: {PGFaculty.objects.count()}")
    print(f"   Total PGDepartment: {PGDepartment.objects.count()}")
    print(f"   Total PGBatch: {PGBatch.objects.count()}")
    print(f"   Total PGDegree: {PGDegree.objects.count()}")
    print(f"   Total PGProgram: {PGProgram.objects.count()}")
    print(f"   Total PGCourseStructure: {PGCourseStructure.objects.count()}")


if __name__ == '__main__':
    import_pg_master_data()
