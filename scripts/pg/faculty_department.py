# Usage: /home/gaurav/.pyenv/versions/pup-umis/bin/python scripts/pg/faculty_department.py
import os
import sys
import django
import openpyxl
from django.conf import settings

def setup_django():
    # Add the project root to sys.path
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    sys.path.append(project_root)
    
    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "pup_umis_backend.settings")
    django.setup()

def run():
    setup_django()
    
    from pg.models import PGFaculty as Faculty, PGDepartment as Department
    from university.models import University
    
    
    # File path
    file_path = os.path.join(settings.BASE_DIR, 'courses_data/pg/subjects_by_faculty.xlsx')
    
    print(f"Reading file: {file_path}")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
    except FileNotFoundError:
        print(f"Error: File not found at {file_path}")
        return
    except Exception as e:
        print(f"Error opening Excel file: {e}")
        return

    # Ensure University exists
    university, created = University.objects.get_or_create(
        name="Purnea University",
        defaults={
            'short_name': 'PU',
            'address': 'Purnia, Bihar'
        }
    )
    if created:
        print("Created University: Purnea University")
    else:
        print("Using existing University: Purnea University")

    # Get Headers (Row 1)
    headers = [cell.value for cell in sheet[1]]
    
    # Analyze columns
    # User says: 
    # Row 1: All faculties
    # Columns under headers: Departments
    # 4th Column (Index 3): "Group" (No faculty)
    
    # We will process each column
    max_col = sheet.max_column
    max_row = sheet.max_row
    
    print(f"Processing {max_col} columns and approx {max_row} rows...")

    for col_idx in range(1, max_col + 1):
        # 1-based index in openpyxl, but python lists are 0-based
        # index 3 means 4th column.
        # col_idx 1 corresponds to headers[0]
        
        header_val = headers[col_idx - 1] if (col_idx - 1) < len(headers) else None
        
        faculty = None
        
        if not header_val:
            continue
        faculty_name = str(header_val).strip()
        print(f"Column {col_idx}: Faculty '{faculty_name}'")

        if not faculty_name:
            continue
            
        # Get or Create Faculty
        # Get or Create Faculty
        # Name is unique globally, so we lookup by name only.
        faculty, created = Faculty.objects.get_or_create(
            name=faculty_name,
            defaults={
                'university': university,
                'short_name': faculty_name[:10] if faculty_name else ''
            }
        )
        if created:
            print(f"  Created Faculty: {faculty.name}")
        
        # Iterate rows for this column
        # Row 2 onwards
        for row_idx in range(2, max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            
            if cell_value:
                department_name = str(cell_value).strip()
                
                if not department_name or department_name.lower() == 'none':
                    continue
                
                # Create/Update Department
                dept, d_created = Department.objects.get_or_create(
                    name=department_name,
                    faculty=faculty,
                    defaults={
                        'code': department_name[:3].upper()
                    }
                )
                
                if d_created:
                    print(f"    Created Department: {department_name}")
                # else:
                #     print(f"    Existing Department: {department_name}")

    print("Done.")

if __name__ == "__main__":
    run()
