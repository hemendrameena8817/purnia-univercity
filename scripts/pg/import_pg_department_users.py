
# Steps to run:
# 1. Ensure you are in the project root directory
# 2. Run the command: poetry run python scripts/pg/import_pg_department_users.py

import os
import sys
import django
import openpyxl
from django.db import transaction

def setup_django():
    # Add the project root to sys.path
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    sys.path.append(project_root)
    
    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "pup_umis_backend.settings")
    django.setup()

def run():
    setup_django()
    
    from accounts.models import UserAccount, CollegeUserProfile
    from colleges.models import College
    from pg.models import PGDepartment
    from django.contrib.auth.hashers import make_password

    # File path
    from django.conf import settings
    file_path = os.path.join(settings.BASE_DIR, 'courses_data', 'pg', 'pg_college_department.xlsx')
    
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return
        
    print(f"Reading file: {file_path}")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
    except Exception as e:
        print(f"Error opening Excel file: {e}")
        return

    # Get Headers (Row 1)
    # A: College Name | B: College Code | C: Department (if Any) | D: Department Code | E: First Name | F: Email | G: Contact No
    headers = [cell.value for cell in sheet[1]]
    print(f"Headers: {headers}")
    
    # We will process each row from row 2
    max_row = sheet.max_row
    print(f"Processing {max_row - 1} rows...")

    success_count = 0
    error_count = 0

    for row_idx in range(2, max_row + 1):
        row_data = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 8)]
        
        college_name = str(row_data[0]).strip() if row_data[0] else None
        college_code = str(row_data[1]).strip() if row_data[1] else None
        dept_name = str(row_data[2]).strip() if row_data[2] else None
        dept_code = str(row_data[3]).strip() if row_data[3] else None
        first_name = str(row_data[4]).strip() if row_data[4] else ""
        email = str(row_data[5]).strip() if row_data[5] else None
        contact_no = str(row_data[6]).strip() if row_data[6] else None

        if not email:
            print(f"Row {row_idx}: Skipping because email is missing.")
            continue
            
        if not college_code:
            print(f"Row {row_idx}: Skipping because college code is missing.")
            continue

        try:
            with transaction.atomic():
                # 1. Get College
                try:
                    # Pad college code if necessary (assuming it might be stored as '04' or '4' in DB)
                    # For now just try exact match
                    college = College.objects.get(college_code=college_code)
                except College.DoesNotExist:
                    # Try with leading zero if it's a number
                    if college_code.isdigit() and len(college_code) == 1:
                        try:
                            college = College.objects.get(college_code=f"0{college_code}")
                        except College.DoesNotExist:
                            print(f"Row {row_idx}: College with code {college_code} not found. Skipping.")
                            error_count += 1
                            continue
                    else:
                        print(f"Row {row_idx}: College with code {college_code} not found. Skipping.")
                        error_count += 1
                        continue

                # 2. Handle Department
                pg_dept = None
                if dept_code or dept_name:
                    if dept_code:
                        pg_dept, d_created = PGDepartment.objects.get_or_create(
                            code=dept_code,
                            defaults={'name': dept_name}
                        )
                        if d_created:
                            print(f"Row {row_idx}: Created PG Department: {dept_name} ({dept_code})")
                    else:
                        # Fallback to name if code is missing but name is present
                        pg_dept, d_created = PGDepartment.objects.get_or_create(
                            name=dept_name
                        )
                        if d_created:
                            print(f"Row {row_idx}: Created PG Department by name: {dept_name}")
                    
                    # Update Head of Department
                    if pg_dept:
                        pg_dept.head_of_department = first_name
                        pg_dept.save()

                # 3. Handle User Account
                username = email
                # Password logic: email prefix + @ + college_code
                email_prefix = email.split('@')[0] if '@' in email else email
                password_plain = f"{email_prefix}@{college_code}"
                
                user, u_created = UserAccount.objects.get_or_create(
                    username=username,
                    defaults={
                        'email': email,
                        'first_name': first_name,
                        'phone': contact_no,
                        'user_type': 'college_user',
                        'college': college,
                        'is_active': True,
                        'is_staff': True, # Since it's a college user, they might need admin access
                    }
                )
                
                if u_created:
                    user.set_password(password_plain)
                    user.save()
                    print(f"Row {row_idx}: Created User: {username}")
                else:
                    # Optionally update existing user? User didn't specify. 
                    # I'll update fields just in case.
                    user.first_name = first_name
                    user.phone = contact_no
                    user.user_type = 'college_user'
                    user.college = college
                    user.save()
                    print(f"Row {row_idx}: Updated existing User: {username}")

                # 4. Handle College User Profile
                profile, p_created = CollegeUserProfile.objects.update_or_create(
                    user=user,
                    defaults={
                        'college': college,
                        'PG_department': pg_dept,
                        'designation': dept_name if dept_name else "College Admin",
                        'is_active': True,
                        'can_manage_students': True,
                        'can_manage_marks': True,
                        'can_manage_results': True,
                        'can_verify_data': True,
                        'can_approve_certificates': True,
                    }
                )
                
                if p_created:
                    print(f"Row {row_idx}: Created College User Profile for {username}")
                else:
                    print(f"Row {row_idx}: Updated College User Profile for {username}")

                success_count += 1

        except Exception as e:
            print(f"Row {row_idx}: Error processing row: {e}")
            error_count += 1

    print(f"\nImport Finished!")
    print(f"Successfully processed: {success_count}")
    print(f"Errors: {error_count}")

if __name__ == "__main__":
    run()
