"""
Import script for applicant_reg_master.xlsx
Run from Django shell:
    exec(open('scripts/import_stagingapplicantregmaster.py').read())
"""
import os
from openpyxl import load_workbook
from staging.models import ApplicantRegMaster

XLSX_PATH = 'old_data/applicant_reg_master.xlsx'

def import_data():
    if not os.path.exists(XLSX_PATH):
        print(f"Error: XLSX file not found at {XLSX_PATH}")
        return
    
    existing = ApplicantRegMaster.objects.count()
    print(f"Existing records: {existing}")
    
    wb = load_workbook(XLSX_PATH, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows:
        print("Empty file!")
        return

    # Skip header
    header = rows[0]
    data_rows = rows[1:]
    
    print(f"Found {len(data_rows)} records to import")
    
    imported = 0
    errors = []
    
    # Bulk create in chunks is faster, but let's do simple loop first for safety
    for i, row in enumerate(data_rows):
        try:
            ApplicantRegMaster.objects.create(
                    applied_date=str(row[0]) if row[0] is not None else None,
                    applied_program=str(row[1]) if row[1] is not None else None,
                    communication_flag=str(row[2]) if row[2] is not None else None,
                    created_by=str(row[3]) if row[3] is not None else None,
                    created_on=str(row[4]) if row[4] is not None else None,
                    dob=str(row[5]) if row[5] is not None else None,
                    dob1=str(row[6]) if row[6] is not None else None,
                    email_id=str(row[7]) if row[7] is not None else None,
                    first_name=str(row[8]) if row[8] is not None else None,
                    csv_id=str(row[9]) if row[9] is not None else None,
                    institute_code=str(row[10]) if row[10] is not None else None,
                    last_name=str(row[11]) if row[11] is not None else None,
                    last_updated=str(row[12]) if row[12] is not None else None,
                    mid_name=str(row[13]) if row[13] is not None else None,
                    mobile=str(row[14]) if row[14] is not None else None,
                    mode=str(row[15]) if row[15] is not None else None,
                    pin=str(row[16]) if row[16] is not None else None,
                    reg_mode=str(row[17]) if row[17] is not None else None,
                    reg_status=str(row[18]) if row[18] is not None else None,
                    reg_user_id=str(row[19]) if row[19] is not None else None,
                    scrutiny_remark=str(row[20]) if row[20] is not None else None,
                    scrutiny_status=str(row[21]) if row[21] is not None else None,
                    status=str(row[22]) if row[22] is not None else None,
                    updated_by=str(row[23]) if row[23] is not None else None,
                    updated_on=str(row[24]) if row[24] is not None else None,
            )
            imported += 1
            if imported % 1000 == 0:
                print(f"Imported {imported} records...")
        except Exception as e:
            errors.append(f"Row {i + 2}: {str(e)}")
    
    print(f"\n✅ Import completed!")
    print(f"   Imported: {imported} records")
    print(f"   Errors: {len(errors)}")
    if errors[:5]:
        print("   First errors:", errors[:5])
    print(f"   Total in table: {ApplicantRegMaster.objects.count()}")

if __name__ == '__main__':
    import_data()
