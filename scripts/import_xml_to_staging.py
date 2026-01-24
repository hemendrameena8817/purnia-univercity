"""
Script to import all XML-converted XLSX files into staging tables.
Uses dynamic field mapping to match XLSX columns to model fields.
Deletes existing data before import.

Usage:
    poetry run python manage.py shell -c "exec(open('scripts/import_xml_to_staging.py').read()); import_all()"
"""
from openpyxl import load_workbook
import os


def clean_value(value):
    """Clean a cell value for database insertion."""
    if value is None:
        return None
    val_str = str(value).strip()
    if val_str in ('', 'None', 'NULL', '[NULL]', '\\N'):
        return None
    return val_str


def get_model_fields(model_class):
    """Get all field names from a Django model (excluding auto fields)."""
    skip_fields = {'id', 'uid', 'is_migrated', 'migration_notes', 'imported_at'}
    return [f.name for f in model_class._meta.fields if f.name not in skip_fields]


def import_xlsx_to_model(xlsx_path, model_class, id_field='csv_id'):
    """
    Generic function to import XLSX data into a Django model.
    Dynamically maps XLSX columns to model fields.
    """
    if not os.path.exists(xlsx_path):
        print(f"❌ File not found: {xlsx_path}")
        return 0
    
    print(f"\n📂 Processing: {xlsx_path}")
    print(f"   Model: {model_class.__name__}")
    
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).lower() if h else f'col_{i}' for i, h in enumerate(next(rows))]
    print(f"   XLSX Columns: {len(headers)}")
    
    # Get model fields
    model_fields = get_model_fields(model_class)
    print(f"   Model fields: {len(model_fields)}")
    
    # Create field mapping (XLSX column -> model field)
    field_mapping = {}
    for header in headers:
        normalized = header.lower().strip()
        
        # Direct match
        if normalized in model_fields:
            field_mapping[header] = normalized
        # Handle 'id' -> id_field mapping
        elif normalized == 'id' and id_field in model_fields:
            field_mapping[header] = id_field
        # Handle special cases
        elif normalized == 'last_update' and 'last_updated' in model_fields:
            field_mapping[header] = 'last_updated'
        elif normalized == 'last_updated' and 'last_update' in model_fields:
            field_mapping[header] = 'last_update'
        elif normalized == 'update_on' and 'updated_on' in model_fields:
            field_mapping[header] = 'updated_on'
        elif normalized == 'updated_on' and 'update_on' in model_fields:
            field_mapping[header] = 'update_on'
        # Handle 3rd prefix (like 3RDCENTER2017)
        elif normalized.startswith('3rd') and normalized[3:] in model_fields:
            field_mapping[header] = normalized[3:]
        elif normalized.startswith('3rd'):
            # Try variations
            for mf in model_fields:
                if normalized.replace('3rd', '') == mf or normalized.replace('3rd', '').replace('_', '') == mf.replace('_', ''):
                    field_mapping[header] = mf
                    break
    
    print(f"   Mapped fields: {len(field_mapping)}")
    unmapped = [h for h in headers if h not in field_mapping]
    if unmapped:
        print(f"   Unmapped: {unmapped[:5]}{'...' if len(unmapped) > 5 else ''}")
    
    # Delete existing records
    existing = model_class.objects.count()
    print(f"   Deleting {existing} existing records...")
    model_class.objects.all().delete()
    
    # Import data
    imported = 0
    errors = []
    for row in rows:
        try:
            row_dict = dict(zip(headers, row))
            
            # Build kwargs for create
            kwargs = {}
            for xlsx_col, model_field in field_mapping.items():
                kwargs[model_field] = clean_value(row_dict.get(xlsx_col))
            
            model_class.objects.create(**kwargs)
            imported += 1
            
            if imported % 10000 == 0:
                print(f"   ...{imported} records")
        except Exception as e:
            errors.append(str(e))
            if len(errors) > 10:
                break
    
    wb.close()
    print(f"   ✅ Imported: {imported}")
    if errors:
        print(f"   ⚠️ Errors: {len(errors)}")
        print(f"      First error: {errors[0][:100]}")
    return imported


def import_all():
    """Import all XML XLSX files into staging tables."""
    from staging.models import (
        StagingInstituteMaster,
        DisciplineMaster, 
        SubjectMaster,
        PaperSubjectMapping,
        CourseDisciplineSemPaperMapping,
        RegisteredApplicantMaster
    )
    
    print("="*60)
    print("IMPORTING ALL XML DATA TO STAGING TABLES")
    print("="*60)
    
    results = {}
    
    # Import each table
    results['institute_master'] = import_xlsx_to_model(
        'old_data/institute_master_xml.xlsx',
        StagingInstituteMaster,
        id_field='institute_id'
    )
    
    results['discipline_master'] = import_xlsx_to_model(
        'old_data/discipline_master_xml.xlsx',
        DisciplineMaster,
        id_field='csv_id'
    )
    
    results['subject_master'] = import_xlsx_to_model(
        'old_data/subject_master_xml.xlsx',
        SubjectMaster,
        id_field='csv_id'
    )
    
    results['paper_subject_mapping'] = import_xlsx_to_model(
        'old_data/paper_subject_mapping_xml.xlsx',
        PaperSubjectMapping,
        id_field='csv_id'
    )
    
    results['course_discipline_sem_paper_mapping'] = import_xlsx_to_model(
        'old_data/course_discipline_sem_paper_mapping_xml.xlsx',
        CourseDisciplineSemPaperMapping,
        id_field='csv_id'
    )
    
    results['registered_applicant_master'] = import_xlsx_to_model(
        'old_data/registered_applicant_master_xml.xlsx',
        RegisteredApplicantMaster,
        id_field='csv_id'
    )
    
    print("\n" + "="*60)
    print("IMPORT SUMMARY")
    print("="*60)
    for table, count in results.items():
        print(f"   {table}: {count} records")
    print(f"\n   TOTAL: {sum(results.values())} records imported")
    print("="*60)


if __name__ == '__main__':
    import_all()
