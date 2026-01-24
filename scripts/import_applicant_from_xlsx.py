"""
Script to import applicant_master from XLSX into StagingApplicantMaster table.
Deletes existing records before import.

Usage:
    poetry run python manage.py shell -c "exec(open('scripts/import_applicant_from_xlsx.py').read()); import_from_xlsx()"
"""
from openpyxl import load_workbook
import os

def clean_value(value):
    if value is None or str(value).strip() in ('', 'None', 'NULL', '\\N'):
        return None
    return str(value).strip()


def import_from_xlsx():
    from staging.models import StagingApplicantMaster
    
    XLSX_PATH = 'old_data/applicant_master_xml.xlsx'
    
    if not os.path.exists(XLSX_PATH):
        print(f"Error: File not found: {XLSX_PATH}")
        return
    
    print(f"Loading XLSX: {XLSX_PATH}")
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    
    # Get headers from first row
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).lower() if h else f'col_{i}' for i, h in enumerate(next(rows))]
    print(f"Found {len(headers)} columns")
    print(f"Columns: {headers[:10]}...")  # Show first 10
    
    # Delete existing records
    existing_count = StagingApplicantMaster.objects.count()
    print(f"\nDeleting {existing_count} existing records...")
    StagingApplicantMaster.objects.all().delete()
    print("✅ Existing records deleted")
    
    # Import data
    print("\nImporting data...")
    imported = 0
    errors = []
    
    for row in rows:
        try:
            row_dict = dict(zip(headers, row))
            
            StagingApplicantMaster.objects.create(
                csv_id=clean_value(row_dict.get('id')),
                reg_user_id=clean_value(row_dict.get('reg_user_id')),
                center=clean_value(row_dict.get('center')),
                applied_program=clean_value(row_dict.get('applied_program')),
                first_name=clean_value(row_dict.get('first_name')),
                mid_name=clean_value(row_dict.get('mid_name')),
                last_name=clean_value(row_dict.get('last_name')),
                full_name=clean_value(row_dict.get('full_name')),
                applied_class=clean_value(row_dict.get('applied_class')),
                exam_center_code=clean_value(row_dict.get('exam_center_code')),
                gender=clean_value(row_dict.get('gender')),
                nationality=clean_value(row_dict.get('nationality')),
                dob=clean_value(row_dict.get('dob')),
                dob_in_word=clean_value(row_dict.get('dob_in_word')),
                blood_group=clean_value(row_dict.get('blood_group')),
                caste=clean_value(row_dict.get('caste')),
                second_language=clean_value(row_dict.get('second_language')),
                univ_regn_no=clean_value(row_dict.get('univ_regn_no')),
                category=clean_value(row_dict.get('category')),
                is_general=clean_value(row_dict.get('is_general')),
                is_physically_challanged=clean_value(row_dict.get('is_physically_challanged')),
                instruction_mode=clean_value(row_dict.get('instruction_mode')),
                last_grade=clean_value(row_dict.get('last_grade')),
                last_board=clean_value(row_dict.get('last_board')),
                present_status=clean_value(row_dict.get('present_status')),
                employer_address=clean_value(row_dict.get('employer_address')),
                comm_address_ref_id=clean_value(row_dict.get('comm_address_ref_id')),
                perm_address_ref_id=clean_value(row_dict.get('perm_address_ref_id')),
                institute_code=clean_value(row_dict.get('institute_code')),
                created_on=clean_value(row_dict.get('created_on')),
                updated_by=clean_value(row_dict.get('updated_by')),
                updated_on=clean_value(row_dict.get('updated_on')),
                religion=clean_value(row_dict.get('religion')),
                applicant_email=clean_value(row_dict.get('applicant_email')),
                applicant_landline=clean_value(row_dict.get('applicant_landline')),
                applicant_mobile=clean_value(row_dict.get('applicant_mobile')),
                highest_qualification=clean_value(row_dict.get('highest_qualification')),
                secured_mark=clean_value(row_dict.get('secured_mark')),
                guardian_name=clean_value(row_dict.get('guardian_name')),
                marital_status=clean_value(row_dict.get('marital_status')),
                created_by=clean_value(row_dict.get('created_by')),
                record_status=clean_value(row_dict.get('record_status')),
                last_updated=clean_value(row_dict.get('last_updated')),
                discipline_code=clean_value(row_dict.get('discipline_code')),
                aadhar_no=clean_value(row_dict.get('aadhar_no')),
                medium=clean_value(row_dict.get('medium')),
                applied=clean_value(row_dict.get('applied')),
                applied_details=clean_value(row_dict.get('applied_details')),
                application_status=clean_value(row_dict.get('application_status')),
                differently_abled=clean_value(row_dict.get('differently_abled')),
            )
            imported += 1
            if imported % 5000 == 0:
                print(f"Imported {imported} records...")
        except Exception as e:
            errors.append(f"Row {imported + 1}: {str(e)}")
            if len(errors) > 10:
                break
    
    wb.close()
    
    print(f"\n✅ Import completed!")
    print(f"   Imported: {imported} records")
    print(f"   Errors: {len(errors)}")
    if errors[:5]:
        print("   First errors:", errors[:5])
    print(f"   Total in table: {StagingApplicantMaster.objects.count()}")


if __name__ == '__main__':
    import_from_xlsx()
