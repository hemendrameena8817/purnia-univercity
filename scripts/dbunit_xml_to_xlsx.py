"""
Script to convert DBUnit XML files to XLSX (Excel) format.

DBUnit XML format example:
<dataset>
    <table_name column1="value1" column2="value2" />
    <table_name column1="value3" column2="value4" />
</dataset>

Usage:
    poetry run python scripts/dbunit_xml_to_xlsx.py data.xml
    poetry run python scripts/dbunit_xml_to_xlsx.py data.xml output.xlsx
"""
import xml.etree.ElementTree as ET
import sys
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def parse_dbunit_xml(xml_path):
    """Parse DBUnit XML and return dict of tables with their rows"""
    tables = defaultdict(list)
    
    tree = ET.parse(xml_path)
    root = tree.getroot()
    
    for element in root:
        table_name = element.tag
        row_data = dict(element.attrib)
        tables[table_name].append(row_data)
    
    return dict(tables)


def dbunit_xml_to_xlsx(xml_path, xlsx_path=None):
    """Convert DBUnit XML file to XLSX with each table as a sheet"""
    if not os.path.exists(xml_path):
        print(f"Error: File not found: {xml_path}")
        return False
    
    # Default output path
    if xlsx_path is None:
        xlsx_path = os.path.splitext(xml_path)[0] + '.xlsx'
    
    print(f"Converting: {xml_path}")
    print(f"Output: {xlsx_path}")
    
    # Parse XML
    print("Parsing XML...")
    tables = parse_dbunit_xml(xml_path)
    print(f"Found {len(tables)} tables")
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    total_rows = 0
    for table_name, rows in tables.items():
        if not rows:
            continue
            
        # Create sheet (max 31 chars for sheet name)
        sheet_name = table_name[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        # Get all columns from all rows
        all_columns = set()
        for row in rows:
            all_columns.update(row.keys())
        columns = sorted(all_columns)
        
        # Write header
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write data
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, '')
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        total_rows += len(rows)
        print(f"  - {table_name}: {len(rows)} rows, {len(columns)} columns")
        
        # Auto-adjust column widths
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(col_name)
            for row_idx in range(2, min(102, len(rows) + 2)):  # Check first 100 rows
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
    
    # Save
    wb.save(xlsx_path)
    
    file_size = os.path.getsize(xlsx_path) / 1024  # KB
    print(f"\n✅ Conversion complete!")
    print(f"   Tables: {len(tables)}")
    print(f"   Total rows: {total_rows}")
    print(f"   Size: {file_size:.2f} KB")
    print(f"   Saved to: {xlsx_path}")
    
    return True


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python scripts/dbunit_xml_to_xlsx.py <xml_file> [output_file.xlsx]")
        sys.exit(1)
    
    xml_path = sys.argv[1]
    xlsx_path = sys.argv[2] if len(sys.argv) > 2 else None
    dbunit_xml_to_xlsx(xml_path, xlsx_path)
