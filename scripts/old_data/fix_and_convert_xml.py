"""
Script to fix and convert DBUnit XML files that have parsing issues.
Uses regex-based parsing to handle malformed XML.
"""
import re
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def extract_records_regex(xml_path):
    """Extract records from DBUnit XML using regex (handles malformed XML)."""
    print(f"Reading file: {xml_path}")
    
    with open(xml_path, 'r', encoding='utf-8', errors='replace') as f:
        content = f.read()
    
    # Find all element tags (table records)
    # Pattern: <TABLE_NAME attr1="val1" attr2="val2" .../>
    pattern = r'<(\w+)\s+([^>]+)/>'
    matches = re.findall(pattern, content)
    
    if not matches:
        print("No records found!")
        return {}
    
    tables = {}
    for match in matches:
        table_name = match[0]
        attrs_str = match[1]
        
        # Parse attributes
        # Pattern: ATTR_NAME="value"
        attr_pattern = r'(\w+)="([^"]*)"'
        attrs = dict(re.findall(attr_pattern, attrs_str))
        
        if table_name not in tables:
            tables[table_name] = {'columns': set(), 'rows': []}
        
        tables[table_name]['columns'].update(attrs.keys())
        tables[table_name]['rows'].append(attrs)
    
    return tables


def convert_to_xlsx(tables, xlsx_path):
    """Convert extracted tables to XLSX format."""
    wb = Workbook()
    wb.remove(wb.active)
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    total_rows = 0
    
    for table_name, data in tables.items():
        ws = wb.create_sheet(title=table_name[:31])  # Excel sheet name limit
        
        columns = sorted(data['columns'])
        rows = data['rows']
        
        # Write header
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
        
        # Write data
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, '')
                # Clean [NULL] values
                if value == '[NULL]':
                    value = ''
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Adjust column widths
        for col_idx, col_name in enumerate(columns, 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(col_name)
            for row in rows[:100]:  # Sample first 100 rows
                val = str(row.get(col_name, ''))
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        
        total_rows += len(rows)
        print(f"  - {table_name}: {len(rows)} rows, {len(columns)} columns")
    
    wb.save(xlsx_path)
    file_size = os.path.getsize(xlsx_path) / 1024
    
    print(f"\n✅ Conversion complete!")
    print(f"   Tables: {len(tables)}")
    print(f"   Total rows: {total_rows}")
    print(f"   Size: {file_size:.2f} KB")
    print(f"   Saved to: {xlsx_path}")
    
    return total_rows


def main():
    if len(sys.argv) < 3:
        print("Usage: python fix_and_convert_xml.py <input_xml> <output_xlsx>")
        sys.exit(1)
    
    xml_path = sys.argv[1]
    xlsx_path = sys.argv[2]
    
    print(f"Converting (robust mode): {xml_path}")
    print(f"Output: {xlsx_path}")
    
    tables = extract_records_regex(xml_path)
    
    if tables:
        convert_to_xlsx(tables, xlsx_path)
    else:
        print("No data to convert!")
        sys.exit(1)


if __name__ == '__main__':
    main()
