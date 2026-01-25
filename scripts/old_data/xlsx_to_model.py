"""
Generic XLSX to Django Model Generator

This script reads an XLSX file, extracts column names from the first sheet's header, and:
1. Generates a Django model class code
2. Generates an admin class code
3. Generates an import script

Usage:
    poetry run python scripts/old_data/xlsx_to_model.py old_data/file.xlsx ModelName
"""
import sys
import os
import re
from openpyxl import load_workbook


def sanitize_field_name(column_name):
    """Convert column name to valid Python/Django field name"""
    s = str(column_name).lower()
    # Replace special chars with underscore
    name = re.sub(r'[^a-z0-9_]', '_', s)
    # Remove consecutive underscores
    name = re.sub(r'_+', '_', name)
    # Remove leading/trailing underscores
    name = name.strip('_')
    # If starts with digit, prefix with underscore
    if name and name[0].isdigit():
        name = f'col_{name}'
    return name or 'unknown_field'


def get_xlsx_columns(xlsx_path):
    """Read XLSX and return column names from first row of active sheet"""
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = []
    
    # Read first row
    first_row = next(ws.iter_rows(values_only=True))
    for cell in first_row:
        if cell:
            headers.append(str(cell).strip())
            
    wb.close()
    return headers


def generate_model_code(table_name, columns):
    """Generate Django model code from column names"""
    class_name = f"Staging{table_name}"
    
    model_code = f'''import uuid
from django.db import models


class {class_name}(models.Model):
    """
    Staging table for {table_name} XLSX data.
    All fields are nullable strings to accept raw data.
    """
    uid = models.UUIDField(default=uuid.uuid4, unique=True, editable=False)
    
'''
    
    for col in columns:
        field_name = sanitize_field_name(col)
        # Use TextField for potentially long fields, CharField for others
        if any(word in field_name for word in ['address', 'description', 'details', 'notes', 'content']):
            model_code += f"    {field_name} = models.TextField(null=True, blank=True, help_text='{col}')\n"
        else:
            model_code += f"    {field_name} = models.TextField(null=True, blank=True, help_text='{col}')\n"
    
    model_code += '''
    # Meta fields
    is_migrated = models.BooleanField(default=False, help_text="Has this record been migrated?")
    migration_notes = models.TextField(null=True, blank=True)
    imported_at = models.DateTimeField(auto_now_add=True)
    
    class Meta:
        verbose_name = 'Staging - ''' + table_name + ''''
        verbose_name_plural = 'Staging - ''' + table_name + ''''
        
    def __str__(self):
        return f"{self.uid}"
'''
    
    return model_code, class_name


def generate_admin_code(class_name, columns):
    """Generate Django admin code"""
    # Pick first 5 meaningful columns for list_display
    display_cols = [sanitize_field_name(c) for c in columns[:5]]
    display_cols = [c for c in display_cols if c] # Filter empty
    display_cols.append('is_migrated')
    display_cols.append('imported_at')
    
    admin_code = f'''
@admin.register({class_name})
class {class_name}Admin(admin.ModelAdmin):
    list_display = {tuple(display_cols)}
    list_filter = ('is_migrated',)
    search_fields = {tuple(display_cols[:3])}
    readonly_fields = ('uid', 'imported_at')
    list_editable = ('is_migrated',)
'''
    return admin_code


def generate_import_script(xlsx_path, class_name, columns):
    """Generate import script code"""
    field_mappings = []
    for col_idx, col in enumerate(columns):
        field_name = sanitize_field_name(col)
        # Assuming row maps to columns by index
        field_mappings.append(f"                    {field_name}=str(row[{col_idx}]) if row[{col_idx}] is not None else None,")
    
    script_code = f'''"""
Import script for {os.path.basename(xlsx_path)}
Run from Django shell:
    exec(open('scripts/import_{class_name.lower()}.py').read())
"""
import os
from openpyxl import load_workbook
from staging.models import {class_name}

XLSX_PATH = '{xlsx_path}'

def import_data():
    if not os.path.exists(XLSX_PATH):
        print(f"Error: XLSX file not found at {{XLSX_PATH}}")
        return
    
    existing = {class_name}.objects.count()
    print(f"Existing records: {{existing}}")
    
    wb = load_workbook(XLSX_PATH, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows:
        print("Empty file!")
        return

    # Skip header
    header = rows[0]
    data_rows = rows[1:]
    
    print(f"Found {{len(data_rows)}} records to import")
    
    imported = 0
    errors = []
    
    # Bulk create in chunks is faster, but let's do simple loop first for safety
    for i, row in enumerate(data_rows):
        try:
            {class_name}.objects.create(
{chr(10).join(field_mappings)}
            )
            imported += 1
            if imported % 1000 == 0:
                print(f"Imported {{imported}} records...")
        except Exception as e:
            errors.append(f"Row {{i + 2}}: {{str(e)}}")
    
    print(f"\\n✅ Import completed!")
    print(f"   Imported: {{imported}} records")
    print(f"   Errors: {{len(errors)}}")
    if errors[:5]:
        print("   First errors:", errors[:5])
    print(f"   Total in table: {{{class_name}.objects.count()}}")

if __name__ == '__main__':
    import_data()
'''
    return script_code


def main():
    if len(sys.argv) < 3:
        print("Usage: python xlsx_to_model.py <xlsx_path> <ModelName>")
        sys.exit(1)
        
    xlsx_path = sys.argv[1]
    model_name = sys.argv[2]
    
    print(f"\n📂 Reading XLSX: {xlsx_path}")
    columns = get_xlsx_columns(xlsx_path)
    print(f"   Found {len(columns)} columns: {columns}")
    
    # Generate model code
    model_code, class_name = generate_model_code(model_name, columns)
    
    # Generate admin code
    admin_code = generate_admin_code(class_name, columns)
    
    # Generate import script
    import_script = generate_import_script(xlsx_path, class_name, columns)
    
    print(f"\n{'='*60}")
    print("MODEL CODE (add to staging/models.py):")
    print('='*60)
    print(model_code)
    
    print(f"\n{'='*60}")
    print("ADMIN CODE (add to staging/admin.py):")
    print('='*60)
    print(admin_code)
    
    # Save import script
    script_path = f"scripts/import_{class_name.lower()}.py"
    with open(script_path, 'w') as f:
        f.write(import_script)
    print(f"\n✅ Import script saved to: {script_path}")
    
    
if __name__ == '__main__':
    main()
