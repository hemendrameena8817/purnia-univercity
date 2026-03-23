import os
import shutil
import uuid
from collections import defaultdict

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

from bca_hons_year.models import BCAHonsStudentCourseAssessment, BCAHonsStudentProfile
from bca_hons_year.models import BCAHonsCommonCourseStructure, BCAHonsCourseStructure
from django.db.models import Q
from openpyxl.drawing.image import Image as XLImage
from bca_hons_year.utils.tr.grading import determine_overall_result, get_hons_classification
from mba_sem.utils.tr.pdf_converter import convert_excel_to_pdf

# ─────────────────────────────────────────────────────────────────
# TEMPLATE LAYOUT (Matched to BCA 77 Column structure)
# ─────────────────────────────────────────────────────────────────

COL_COMP_P1 = 7
COL_COMP_P2 = 8
COL_COMP_TOT = 9

COL_HONS_TH_P1 = 10
COL_HONS_TH_P2 = 11
COL_HONS_TH_P9 = 12
COL_HONS_TH_P10 = 13
COL_HONS_TH_P11 = 14
COL_HONS_TH_P12 = 15
COL_HONS_TH_P3_TOT = 16
COL_HONS_TH_GRAND_TOT = 17

COL_HONS_PR_P1 = 18
COL_HONS_PR_P2 = 19
COL_HONS_PR_P9 = 20
COL_HONS_PR_P10 = 21
COL_HONS_PR_P11 = 22
COL_HONS_PR_P3_TOT = 23
COL_HONS_PR_GRAND_TOT = 24

COL_HONS_OVERALL_TOT = 25

def get_sub_cols(idx):
    # Subs 1 through 6
    start = 26 + (idx * 8)
    return {
        "name": start,
        "th_1": start + 1,
        "th_2": start + 2,
        "th_tot": start + 3,
        "pr_1": start + 4,
        "pr_2": start + 5,
        "pr_tot": start + 6,
        "grand_tot": start + 7
    }

COL_GES = 74
COL_AGGREGATE = 75
COL_RESULT = 76
COL_REMARKS = 77

DATA_START_ROW = 15
STUDENTS_PER_PAGE = 13
TEMPLATE_CAPACITY = 13

# ─────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────

def _get_marks(student_id, code, student_map, mark_type='ESE', course_name_hint=None):
    best_res = {"marks": 0.0, "found": False, "absent": False}
    search_labels = []
    if mark_type == 'ESE':
        search_labels = ['ESE', 'THEORY', 'Theory']
    else:
        search_labels = ['CIA', 'IA', 'INTERNAL', 'Internal Assessment', 'PRACTICAL', 'Practical']

    found_any = False
    all_absent = True
    for rec in student_map.get(student_id, []):
        if str(rec.paper_code or "").strip().upper() == str(code or "").strip().upper():
            if course_name_hint and course_name_hint.upper() not in str(rec.course_name or "").upper():
                continue
            
            label = str(rec.label or "").upper()
            if any(s.upper() in label for s in search_labels):
                is_absent = bool(rec.ind_is_absent)
                marks_raw = rec.ind_final_marks_obtained if rec.ind_final_marks_obtained is not None else rec.ind_marks_obtained
                
                if marks_raw is not None:
                    found_any = True
                    if not is_absent: all_absent = False
                    marks = float(marks_raw)
                    if not best_res["found"] or marks >= best_res["marks"]:
                        best_res["marks"] = marks; best_res["found"] = True; best_res["absent"] = is_absent
                elif is_absent:
                    found_any = True
                    # If it is just an absent record but no marks, we still count as found
                    if not best_res["found"]:
                         best_res["marks"] = 0.0; best_res["found"] = True; best_res["absent"] = True
    
    if found_any and not all_absent: 
        best_res["absent"] = False
    return best_res

def _get_previous_year_honours_marks_total(student_id, year_str, student_map, honours_codes, mark_type='ESE', course_name_hint=None):
    total = 0.0
    found_any = False
    all_absent = True
    for code in honours_codes:
        m = _get_marks(student_id, code, student_map, mark_type=mark_type, course_name_hint=course_name_hint)
        if m["found"]:
            found_any = True
            if not m["absent"]:
                all_absent = False
                total += m["marks"]
    if not found_any: 
        all_absent = False
    return total, all_absent, found_any

def fmt_tot(val, is_ab, found=True):
    if not found: return None
    return "AB" if is_ab else val

def fill_students(ws, students, student_map, honours_map, sub_map, sub_names, appeared_codes, target_year):
    row = DATA_START_ROW
    counts = {"total": len(students), "dist": 0, "first": 0, "second": 0, "fail": 0, "pending": 0, "absent": 0, "expelled": 0}

    # Identify paper codes for Part 3
    # Based on the image, the 4 Hons papers in Part 3 are 9th, 10th, 11th, 12th
    target_year = str(target_year)
    p3_h_codes = honours_map.get("3", [])
    code_9 = p3_h_codes[0] if len(p3_h_codes) > 0 else "BCAHN301"
    code_10 = p3_h_codes[1] if len(p3_h_codes) > 1 else "BCAHN302"
    code_11 = p3_h_codes[2] if len(p3_h_codes) > 2 else "BCAHN303"
    code_12 = p3_h_codes[3] if len(p3_h_codes) > 3 else "BCAHN304"
    
    ges_codes = list(BCAHonsCommonCourseStructure.objects.filter(year="3", paper_type="COMPOSITION").values_list("code", flat=True))
    ges_code = ges_codes[0] if ges_codes else "BCAHN305"

    rbh_codes_y1 = list(BCAHonsCommonCourseStructure.objects.filter(year="1", paper_type="COMPOSITION").values_list("code", flat=True))
    rbh_code_y1 = rbh_codes_y1[0] if rbh_codes_y1 else "BCAHN107"
    rbh_codes_y2 = list(BCAHonsCommonCourseStructure.objects.filter(year="2", paper_type="COMPOSITION").values_list("code", flat=True))
    rbh_code_y2 = rbh_codes_y2[0] if rbh_codes_y2 else "BCAHN207"

    for student in students:
        ws.cell(row=row, column=1).value = student.roll_no
        ws.cell(row=row, column=2).value = student.get_full_name()
        ws.cell(row=row, column=3).value = f"{student.registration_no} \n {student.registration_year}" if getattr(student, 'registration_year', None) else student.registration_no

        ws.cell(row=row, column=4).value = f"{student.roll_no} / {student.session_str}"
        ws.cell(row=row, column=5).value = f"{student.roll_no} / {student.session_str}"
        ws.cell(row=row, column=6).value = student.course.name if student.course else "BCA HONS"

        # Composition Subject (Part 1, Part 2)
        m_comp_1 = _get_marks(student.id, rbh_code_y1, student_map, mark_type='ESE')
        m_comp_2 = _get_marks(student.id, rbh_code_y2, student_map, mark_type='ESE')
        c1_val = fmt_tot(m_comp_1["marks"], m_comp_1["absent"], m_comp_1["found"])
        c2_val = fmt_tot(m_comp_2["marks"], m_comp_2["absent"], m_comp_2["found"])
        ws.cell(row=row, column=COL_COMP_P1).value = c1_val
        ws.cell(row=row, column=COL_COMP_P2).value = c2_val
        tot_comp = (m_comp_1["marks"] if not m_comp_1["absent"] else 0) + (m_comp_2["marks"] if not m_comp_2["absent"] else 0)
        tot_comp_ab = m_comp_1["absent"] and m_comp_2["absent"]
        tot_comp_found = m_comp_1["found"] or m_comp_2["found"]
        ws.cell(row=row, column=COL_COMP_TOT).value = fmt_tot(tot_comp, tot_comp_ab, tot_comp_found)

        # Honours Subject (Theory)
        h1_th, h1_th_ab, h1_th_found = _get_previous_year_honours_marks_total(student.id, "1", student_map, honours_map.get("1", []), mark_type='ESE')
        h2_th, h2_th_ab, h2_th_found = _get_previous_year_honours_marks_total(student.id, "2", student_map, honours_map.get("2", []), mark_type='ESE')
        ws.cell(row=row, column=COL_HONS_TH_P1).value = fmt_tot(h1_th, h1_th_ab, h1_th_found)
        ws.cell(row=row, column=COL_HONS_TH_P2).value = fmt_tot(h2_th, h2_th_ab, h2_th_found)

        m9_th = _get_marks(student.id, code_9, student_map, mark_type='ESE')
        m10_th = _get_marks(student.id, code_10, student_map, mark_type='ESE')
        m11_th = _get_marks(student.id, code_11, student_map, mark_type='ESE')
        m12_th = _get_marks(student.id, code_12, student_map, mark_type='ESE')
        
        ws.cell(row=row, column=COL_HONS_TH_P9).value = fmt_tot(m9_th["marks"], m9_th["absent"], m9_th["found"])
        ws.cell(row=row, column=COL_HONS_TH_P10).value = fmt_tot(m10_th["marks"], m10_th["absent"], m10_th["found"])
        ws.cell(row=row, column=COL_HONS_TH_P11).value = fmt_tot(m11_th["marks"], m11_th["absent"], m11_th["found"])
        ws.cell(row=row, column=COL_HONS_TH_P12).value = fmt_tot(m12_th["marks"], m12_th["absent"], m12_th["found"])

        p3_th_tot = (m9_th["marks"] if not m9_th["absent"] else 0) + (m10_th["marks"] if not m10_th["absent"] else 0) + (m11_th["marks"] if not m11_th["absent"] else 0) + (m12_th["marks"] if not m12_th["absent"] else 0)
        p3_th_tot_ab = m9_th["absent"] and m10_th["absent"] and m11_th["absent"] and m12_th["absent"]
        p3_th_found = m9_th["found"] or m10_th["found"] or m11_th["found"] or m12_th["found"]
        ws.cell(row=row, column=COL_HONS_TH_P3_TOT).value = fmt_tot(p3_th_tot, p3_th_tot_ab, p3_th_found)

        grand_th_tot = (h1_th if not h1_th_ab else 0) + (h2_th if not h2_th_ab else 0) + (p3_th_tot if not p3_th_tot_ab else 0)
        grand_th_ab = h1_th_ab and h2_th_ab and p3_th_tot_ab
        grand_th_found = h1_th_found or h2_th_found or p3_th_found
        ws.cell(row=row, column=COL_HONS_TH_GRAND_TOT).value = fmt_tot(grand_th_tot, grand_th_ab, grand_th_found)

        # Honours Subject (Practical) -- BCA has practicals as per standard CIA structure
        h1_pr, h1_pr_ab, h1_pr_found = _get_previous_year_honours_marks_total(student.id, "1", student_map, honours_map.get("1", []), mark_type='CIA')
        h2_pr, h2_pr_ab, h2_pr_found = _get_previous_year_honours_marks_total(student.id, "2", student_map, honours_map.get("2", []), mark_type='CIA')
        ws.cell(row=row, column=COL_HONS_PR_P1).value = fmt_tot(h1_pr, h1_pr_ab, h1_pr_found)
        ws.cell(row=row, column=COL_HONS_PR_P2).value = fmt_tot(h2_pr, h2_pr_ab, h2_pr_found)

        m9_pr = _get_marks(student.id, code_9, student_map, mark_type='CIA')
        m10_pr = _get_marks(student.id, code_10, student_map, mark_type='CIA')
        m11_pr = _get_marks(student.id, code_11, student_map, mark_type='CIA')
        # Based on image, 12th paper does not have practical
        
        ws.cell(row=row, column=COL_HONS_PR_P9).value = fmt_tot(m9_pr["marks"], m9_pr["absent"], m9_pr["found"])
        ws.cell(row=row, column=COL_HONS_PR_P10).value = fmt_tot(m10_pr["marks"], m10_pr["absent"], m10_pr["found"])
        ws.cell(row=row, column=COL_HONS_PR_P11).value = fmt_tot(m11_pr["marks"], m11_pr["absent"], m11_pr["found"])

        p3_pr_tot = (m9_pr["marks"] if not m9_pr["absent"] else 0) + (m10_pr["marks"] if not m10_pr["absent"] else 0) + (m11_pr["marks"] if not m11_pr["absent"] else 0)
        p3_pr_tot_ab = m9_pr["absent"] and m10_pr["absent"] and m11_pr["absent"]
        p3_pr_found = m9_pr["found"] or m10_pr["found"] or m11_pr["found"]
        ws.cell(row=row, column=COL_HONS_PR_P3_TOT).value = fmt_tot(p3_pr_tot, p3_pr_tot_ab, p3_pr_found)

        grand_pr_tot = (h1_pr if not h1_pr_ab else 0) + (h2_pr if not h2_pr_ab else 0) + (p3_pr_tot if not p3_pr_tot_ab else 0)
        grand_pr_ab = h1_pr_ab and h2_pr_ab and p3_pr_tot_ab
        grand_pr_found = h1_pr_found or h2_pr_found or p3_pr_found
        ws.cell(row=row, column=COL_HONS_PR_GRAND_TOT).value = fmt_tot(grand_pr_tot, grand_pr_ab, grand_pr_found)

        # Honours Overall Total
        hons_overall = grand_th_tot + grand_pr_tot
        ws.cell(row=row, column=COL_HONS_OVERALL_TOT).value = fmt_tot(hons_overall, grand_th_ab and grand_pr_ab, grand_th_found or grand_pr_found)

        # Subsidiary Subjects (Up to 6)
        subsidiary_tot = 0
        for idx in range(6):
            c_info = get_sub_cols(idx)
            sub_key = str(idx + 1)
            ws.cell(row=row, column=c_info["name"]).value = sub_names.get(sub_key, "")

            # Assume each subsidiary paper corresponds to p1 in yr1 and p2 in yr2
            p1_codes = sub_map.get(sub_key, {}).get("p1", [])
            p2_codes = sub_map.get(sub_key, {}).get("p2", [])

            # Theory / ESE marks (Subject 1 Yr 1, Subject 2 Yr 2)
            n1_full = sub_map.get(sub_key, {}).get("n1_name", "")
            n2_full = sub_map.get(sub_key, {}).get("n2_name", "")
            
            def get_hint(name):
                uname = str(name).upper()
                if "MATHEMATICS" in uname: return "MATH"
                if "PHYSICS" in uname: return "PHY"
                if "ECONOMICS" in uname: return "ECON"
                if "ENGLISH" in uname: return "ENG"
                if "BUSINESS" in uname: return "BUSI"
                if "MONEY" in uname: return "MONEY"
                if "FINANCIAL" in uname: return "FINAN"
                if "PLANNING" in uname: return "PLAN"
                return name

            h1 = get_hint(n1_full)
            h2 = get_hint(n2_full)
            
            s1_th, s1_th_ab, s1_th_found = _get_previous_year_honours_marks_total(student.id, "1", student_map, p1_codes, mark_type='ESE', course_name_hint=h1)
            s2_th, s2_th_ab, s2_th_found = _get_previous_year_honours_marks_total(student.id, "2", student_map, p2_codes, mark_type='ESE', course_name_hint=h2)
            ws.cell(row=row, column=c_info["th_1"]).value = fmt_tot(s1_th, s1_th_ab, s1_th_found)
            ws.cell(row=row, column=c_info["th_2"]).value = fmt_tot(s2_th, s2_th_ab, s2_th_found)
            ws.cell(row=row, column=c_info["th_tot"]).value = fmt_tot(s1_th + s2_th, s1_th_ab and s2_th_ab, s1_th_found or s2_th_found)

            # Practical / CIA marks
            has_pr1 = sub_map.get(sub_key, {}).get("has_pr1", True)
            has_pr2 = sub_map.get(sub_key, {}).get("has_pr2", True)

            if has_pr1:
                s1_pr, s1_pr_ab, s1_pr_found = _get_previous_year_honours_marks_total(student.id, "1", student_map, p1_codes, mark_type='CIA', course_name_hint=h1)
            else:
                s1_pr, s1_pr_ab, s1_pr_found = 0, False, False

            if has_pr2:
                s2_pr, s2_pr_ab, s2_pr_found = _get_previous_year_honours_marks_total(student.id, "2", student_map, p2_codes, mark_type='CIA', course_name_hint=h2)
            else:
                s2_pr, s2_pr_ab, s2_pr_found = 0, False, False

            ws.cell(row=row, column=c_info["pr_1"]).value = fmt_tot(s1_pr, s1_pr_ab, s1_pr_found)
            ws.cell(row=row, column=c_info["pr_2"]).value = fmt_tot(s2_pr, s2_pr_ab, s2_pr_found)
            ws.cell(row=row, column=c_info["pr_tot"]).value = fmt_tot(s1_pr + s2_pr, s1_pr_ab and s2_pr_ab, s1_pr_found or s2_pr_found)

            sub_grand = (s1_th if not s1_th_ab else 0) + (s2_th if not s2_th_ab else 0) + (s1_pr if not s1_pr_ab else 0) + (s2_pr if not s2_pr_ab else 0)
            sub_grand_found = s1_th_found or s2_th_found or s1_pr_found or s2_pr_found
            ws.cell(row=row, column=c_info["grand_tot"]).value = fmt_tot(sub_grand, s1_th_ab and s2_th_ab and s1_pr_ab and s2_pr_ab, sub_grand_found)
            subsidiary_tot += sub_grand

        # GES
        m_ges = _get_marks(student.id, ges_code, student_map, mark_type='ESE')
        ges_tot = m_ges["marks"] if not m_ges["absent"] else 0
        ws.cell(row=row, column=COL_GES).value = fmt_tot(m_ges["marks"], m_ges["absent"], m_ges["found"])

        # Aggregate Total
        aggregate_sum = hons_overall + subsidiary_tot + ges_tot
        aggregate_found = grand_th_found or grand_pr_found or sub_grand_found or m_ges["found"]
        ws.cell(row=row, column=COL_AGGREGATE).value = fmt_tot(aggregate_sum, False, aggregate_found)

        # Calculate Full Result logic using the adapted determine_overall_result
        hons_data = [] # Just pass dummy passes that trigger correct result parsing
        sub_data = []  
        # Let's mock a simple bypass since true calculation relies on max marks lookup 
        # which isn't currently populated in db, but we replicate the basic TR flow

        # Simplistic mapping (Placeholder until database is filled):
        class_str = "PENDING"
        if aggregate_sum > 0:
            pct = (hons_overall / 800) * 100
            if pct >= 75: class_str = "1st Class With Distinction"
            elif pct >= 60: class_str = "1st Class"
            elif pct >= 45: class_str = "2nd Class"
            else: class_str = "FAIL"
            
        ws.cell(row=row, column=COL_RESULT).value = class_str

        if "Distinction" in class_str: counts["dist"] += 1
        elif "1st Class" in class_str: counts["first"] += 1
        elif "2nd Class" in class_str: counts["second"] += 1
        elif "FAIL" in class_str: counts["fail"] += 1

        row += 1

    return counts

def fill_marks_header(ws, honours_map, sub_map):
    # Full Marks Row: 12, Pass Marks Row: 13
    fm_row, pm_row = 12, 13
    
    # Composition (Part 1, Part 2, Total)
    ws.cell(row=fm_row, column=COL_COMP_P1).value = 100; ws.cell(row=pm_row, column=COL_COMP_P1).value = 33
    ws.cell(row=fm_row, column=COL_COMP_P2).value = 100; ws.cell(row=pm_row, column=COL_COMP_P2).value = 33
    ws.cell(row=fm_row, column=COL_COMP_TOT).value = 200; ws.cell(row=pm_row, column=COL_COMP_TOT).value = 66
    
    # Honours Theory
    h1_count = len(honours_map.get("1", []))
    h2_count = len(honours_map.get("2", []))
    ws.cell(row=fm_row, column=COL_HONS_TH_P1).value = h1_count * 100; ws.cell(row=pm_row, column=COL_HONS_TH_P1).value = h1_count * 45
    ws.cell(row=fm_row, column=COL_HONS_TH_P2).value = h2_count * 100; ws.cell(row=pm_row, column=COL_HONS_TH_P2).value = h2_count * 45
    
    # Part 3 individual papers (9, 10, 11, 12)
    for c in [COL_HONS_TH_P9, COL_HONS_TH_P10, COL_HONS_TH_P11, COL_HONS_TH_P12]:
        ws.cell(row=fm_row, column=c).value = 100; ws.cell(row=pm_row, column=c).value = 45
    
    ws.cell(row=fm_row, column=COL_HONS_TH_P3_TOT).value = 400; ws.cell(row=pm_row, column=COL_HONS_TH_P3_TOT).value = 180
    
    grand_th_fm = (h1_count + h2_count + 4) * 100
    ws.cell(row=fm_row, column=COL_HONS_TH_GRAND_TOT).value = grand_th_fm; ws.cell(row=pm_row, column=COL_HONS_TH_GRAND_TOT).value = int(grand_th_fm * 0.45)
    
    # Honours Practical (Assuming 50 FM / 23 PM per practical)
    ws.cell(row=fm_row, column=COL_HONS_PR_P1).value = h1_count * 50; ws.cell(row=pm_row, column=COL_HONS_PR_P1).value = h1_count * 23
    ws.cell(row=fm_row, column=COL_HONS_PR_P2).value = h2_count * 50; ws.cell(row=pm_row, column=COL_HONS_PR_P2).value = h2_count * 23
    
    for c in [COL_HONS_PR_P9, COL_HONS_PR_P10, COL_HONS_PR_P11]:
        ws.cell(row=fm_row, column=c).value = 50; ws.cell(row=pm_row, column=c).value = 23
    
    ws.cell(row=fm_row, column=COL_HONS_PR_P3_TOT).value = 150; ws.cell(row=pm_row, column=COL_HONS_PR_P3_TOT).value = 69
    
    grand_pr_fm = (h1_count + h2_count + 3) * 50
    ws.cell(row=fm_row, column=COL_HONS_PR_GRAND_TOT).value = grand_pr_fm; ws.cell(row=pm_row, column=COL_HONS_PR_GRAND_TOT).value = int(grand_pr_fm * 0.46)
    
    ws.cell(row=fm_row, column=COL_HONS_OVERALL_TOT).value = grand_th_fm + grand_pr_fm
    ws.cell(row=pm_row, column=COL_HONS_OVERALL_TOT).value = int((grand_th_fm + grand_pr_fm) * 0.45)

    # Subsidiaries
    for idx in range(6):
        c_info = get_sub_cols(idx)
        sub_key = str(idx + 1)
        config = sub_map.get(sub_key, {})
        has_pr1 = config.get("has_pr1", True)
        has_pr2 = config.get("has_pr2", True)
        
        # Paper codes in Row 11
        p1_list = config.get("p1", [])
        p2_list = config.get("p2", [])
        c1 = next((x for x in p1_list if "1" in x), p1_list[0]) if p1_list else ""
        c2 = next((x for x in p2_list if "2" in x), p2_list[0]) if p2_list else ""
        ws.cell(row=11, column=c_info["th_1"]).value = c1
        ws.cell(row=11, column=c_info["th_2"]).value = c2
        if has_pr1: ws.cell(row=11, column=c_info["pr_1"]).value = c1
        if has_pr2: ws.cell(row=11, column=c_info["pr_2"]).value = c2

        # Physics/Science Rule: 75/25 split for practical subjects
        fm1, pm1 = (75, 25) if has_pr1 else (100, 33)
        fm2, pm2 = (75, 25) if has_pr2 else (100, 33)
        
        ws.cell(row=fm_row, column=c_info["th_1"]).value = fm1; ws.cell(row=pm_row, column=c_info["th_1"]).value = pm1
        ws.cell(row=fm_row, column=c_info["th_2"]).value = fm2; ws.cell(row=pm_row, column=c_info["th_2"]).value = pm2
        ws.cell(row=fm_row, column=c_info["th_tot"]).value = fm1 + fm2; ws.cell(row=pm_row, column=c_info["th_tot"]).value = pm1 + pm2
        
        # Practical FM/PM (25/9 if science, else 50/17 placeholder)
        pfm1, ppm1 = (25, 9) if has_pr1 else (0, 0)
        pfm2, ppm2 = (25, 9) if has_pr2 else (0, 0)
        
        ws.cell(row=fm_row, column=c_info["pr_1"]).value = pfm1 if pfm1 > 0 else None
        ws.cell(row=pm_row, column=c_info["pr_1"]).value = ppm1 if ppm1 > 0 else None
        ws.cell(row=fm_row, column=c_info["pr_2"]).value = pfm2 if pfm2 > 0 else None
        ws.cell(row=pm_row, column=c_info["pr_2"]).value = ppm2 if ppm2 > 0 else None
        
        pr_fm_tot = pfm1 + pfm2
        pr_pm_tot = ppm1 + ppm2
        ws.cell(row=fm_row, column=c_info["pr_tot"]).value = pr_fm_tot if pr_fm_tot > 0 else None
        ws.cell(row=pm_row, column=c_info["pr_tot"]).value = pr_pm_tot if pr_pm_tot > 0 else None
        
        grand_fm_tot = fm1 + fm2 + pfm1 + pfm2
        grand_pm_tot = pm1 + pm2 + ppm1 + ppm2
        ws.cell(row=fm_row, column=c_info["grand_tot"]).value = grand_fm_tot if grand_fm_tot > 0 else None
        ws.cell(row=pm_row, column=c_info["grand_tot"]).value = grand_pm_tot if grand_pm_tot > 0 else None

    # GES
    ws.cell(row=fm_row, column=COL_GES).value = 100; ws.cell(row=pm_row, column=COL_GES).value = 33
    
    # Aggregate (Hons + GES + Composition + Subs)
    ws.cell(row=fm_row, column=COL_AGGREGATE).value = 3050 
    ws.cell(row=pm_row, column=COL_AGGREGATE).value = 1154


def fill_footer(ws, counts, row):
    ws.cell(row=row,     column=1).value = f"No. of Students : {counts['total']}"
    ws.cell(row=row + 1, column=1).value = f"1st Class With Distinction : {counts['dist']}"
    ws.cell(row=row + 2, column=1).value = f"Fail : {counts['fail']}"
    ws.cell(row=row + 3, column=1).value = f"Result Pending : {counts['pending']}"

    ws.cell(row=row + 1, column=5).value = f"1st Class : {counts['first']}"
    ws.cell(row=row + 2, column=5).value = f"Absent : {counts['absent']}"

    ws.cell(row=row + 2, column=9).value = f"Expelled : {counts['expelled']}"
    ws.cell(row=row + 3, column=9).value = f"2nd Class : {counts['second']}"

def fill_header(ws, exam_name, college, center_name=None, subject_name=None):
    ws["AQ4"] = exam_name or "BCA Part III Examination"
    ws["BN3"] = f"Subject : {subject_name or 'B.C.A.'}"
    ws["BN4"] = f"Centre : {center_name}"
    ws["BN5"] = f"Dept./College : {college.name}"

def add_logo(ws):
    # Absolute path to ensure it works in all environments
    logo_path = os.path.join(settings.BASE_DIR, "static", "images", "purnea-logo.png")
    if os.path.exists(logo_path):
        img = XLImage(logo_path)
        # 1 inch is approx 96 pixels, setting it to 75x75 for a standard fit
        img.width, img.height = 100, 100
        # Ensure row 1 is high enough for the logo (approx 75-80 points for 100px)
        ws.row_dimensions[1].height = 80
        # Specifically anchor to AL1 as requested
        ws.add_image(img, "AL1")
    else:
        # Debugging fallback if file is missing (though our check said it exists)
        ws["AL1"] = "[LOGO MISSING]"

class BCAHons3rdYearResultGenerator:
    def __init__(self, students, college, batch_uid=None, year="3", exam_name=None):
        self.students = list(students)
        self.college = college
        self.batch_uid = batch_uid
        self.year = str(year)
        self.exam_name = exam_name

    def generate(self):
        if not self.students: return None
        template_path = os.path.join(settings.BASE_DIR, "bca_hons_year", "static", "tr", "BCA_Hons_3rd_year.xlsx")
        temp_excel = os.path.join(os.environ.get("TEMP", "/tmp"), f"bcah3_{uuid.uuid4().hex}.xlsx")
        shutil.copy(template_path, temp_excel)
        wb = load_workbook(temp_excel)
        master = wb.active
        master.title = "MASTER"

        qs = BCAHonsStudentCourseAssessment.objects.filter(student__in=self.students, year__in=["1", "2", "3"]).order_by('-id')
        if self.batch_uid: qs = qs.filter(batch__uid=self.batch_uid)
        
        student_map = defaultdict(list)
        for obj in qs: student_map[obj.student.id].append(obj)
        print(f"=================Student Map======================")
        print(f"Unique Students with Assessment Records: {len(student_map)}")
        print(f"=================Student Map======================")
        honours_map = {
            "1": list(BCAHonsCommonCourseStructure.objects.filter(year="1", paper_type="HONOURS").order_by('code').values_list("code", flat=True)),
            "2": list(BCAHonsCommonCourseStructure.objects.filter(year="2", paper_type="HONOURS").order_by('code').values_list("code", flat=True)),
            "3": list(BCAHonsCommonCourseStructure.objects.filter(year="3", paper_type="HONOURS").order_by('code').values_list("code", flat=True)),
        }

        # Map subsidiaries based on specific user request for slots Z to BN
        # Parameters: (P1 Name, P2 Name, P1 Has Practical, P2 Has Practical)
        subs_config = [
            ("Physics", "Economics", True, False),
            ("Mathematics", "English", True, False),
            ("Economics", "Mathematics", False, True),
            ("English", "Physics", False, True),
            ("Business Organisation", "Money & Banking", False, False),
            ("Financial Accounting", "Planning and Economic Development of India", False, False)
        ]
        
        sub_map = {}
        sub_names = {}
        
        for idx, (n1, n2, pr1_flag, pr2_flag) in enumerate(subs_config):
            # 1. Search for codes matching these subject names in common course structure
            # Strictly filter by SUBSIDIARY to avoid clashing with Honours papers (101-104)
            p1_codes = list(BCAHonsCommonCourseStructure.objects.filter(year="1", paper_type="SUBSIDIARY", course_name__icontains=n1).values_list("code", flat=True))
            p2_codes = list(BCAHonsCommonCourseStructure.objects.filter(year="2", paper_type="SUBSIDIARY", course_name__icontains=n2).values_list("code", flat=True))
            
            # 2. Add fallbacks including the ambiguous codes 105, 106, 205, 206
            # We rely on our new HINT system to distinguish between them
            fallback_p1 = ["BCAHN1" + "".join(filter(str.isalpha, n1.upper()))[:2]]
            if "Phy" in n1 or "Math" in n1 or "Eco" in n1 or "Eng" in n1:
                fallback_p1 += ["BCAHN105", "BCAHN106"]
            
            fallback_p2 = ["BCAHN2" + "".join(filter(str.isalpha, n2.upper()))[:2]]
            if "Eco" in n2 or "Eng" in n2 or "Math" in n2 or "Phy" in n2:
                fallback_p2 += ["BCAHN201", "BCAHN205", "BCAHN206"]

            sub_map[str(idx+1)] = {
                "n1_name": n1,
                "n2_name": n2,
                "p1": list(set(p1_codes + fallback_p1)), 
                "p2": list(set(p2_codes + fallback_p2)),
                "has_pr1": pr1_flag,
                "has_pr2": pr2_flag
            }
            sub_names[str(idx+1)] = f"{n1} / {n2}"

        appeared_codes = set(qs.values_list("paper_code", flat=True))
        
        print(f"\n--- DEBUG: STUDENT DATA CHECK ---")
        print(f"Total students in Profile for this College/Batch: {len(self.students)}")
        
        students_no_records = [s for s in self.students if s.id not in student_map]
        print(f"Students with NO Assessment records at all ({len(students_no_records)}):")
        if students_no_records:
            print(f"Roll Nos: {', '.join([str(s.roll_no) for s in students_no_records])}")
        
        target_year_students = []
        missing_target_year = []
        for s in [s for s in self.students if s.id in student_map]:
            if any(rec.year == self.year for rec in student_map[s.id]):
                target_year_students.append(s)
            else:
                missing_target_year.append(s)
        
        print(f"Students with records for other years but MISSING Year {self.year} ({len(missing_target_year)}):")
        if missing_target_year:
            print(f"Roll Nos: {', '.join([str(s.roll_no) + " " + str(s.registration_no) for s in missing_target_year])}")
            
        self.students = target_year_students
        print(f"Final Count for TR (Students with Year {self.year} data): {len(self.students)}")
        print(f"---------------------------------\n")

        if not self.students: return None

        pages = [self.students[i : i+STUDENTS_PER_PAGE] for i in range(0, len(self.students), STUDENTS_PER_PAGE)]
        for idx, chunk in enumerate(pages):
            ws = wb.copy_worksheet(master)
            ws.title = f"Page_{idx+1}"
            
            # Add Logo and Headers
            # Fill Part-I and Part-II Full/Pass Marks
            fill_marks_header(ws, honours_map, sub_map)
            
            counts = fill_students(ws, chunk, student_map, honours_map, sub_map, sub_names, appeared_codes, self.year)
            
            # Hide unused student rows so the grid doesn't show empty boxes
            unused_count = TEMPLATE_CAPACITY - len(chunk)
            if unused_count > 0:
                for empty_row in range(DATA_START_ROW + len(chunk), DATA_START_ROW + TEMPLATE_CAPACITY):
                    ws.row_dimensions[empty_row].hidden = True
                
            from bca_hons_year.models import BCAHonsExamCenterMapping, BCAHonsExam
            exam_obj = BCAHonsExam.objects.filter(year=self.year).last()
            center_name = "-"
            if exam_obj:
                mapping = BCAHonsExamCenterMapping.objects.filter(exam=exam_obj, attached_colleges=self.college).first()
                if mapping: center_name = mapping.center.name
            
            subject_name = self.students[0].course.name if (self.students and self.students[0].course) else "BCA HONS"

            fill_header(ws, self.exam_name, self.college, center_name=center_name, subject_name=subject_name)
            
            # Properly fixed footer logic: 
            # We want the footer at physical Row 45. 
            # If we hide N rows above it, we must write to Row (45 + N) to keep it at the same physical spot.
            target_physical_row = 34
            footer_row = target_physical_row + unused_count
            fill_footer(ws, counts, footer_row)
            
            # Explicitly add logo last to ensure visibility
            add_logo(ws)
            
            ws.page_setup.paperSize = 8  # A3
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_margins.left = 0
            ws.page_margins.right = 0
            ws.page_margins.top = 0.2
            ws.page_margins.bottom = 0.5
            ws.print_area = ws.dimensions
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
            ws.page_setup.scale = None

        if "MASTER" in wb.sheetnames: wb.remove(wb["MASTER"])
        wb.save(temp_excel)
        return convert_excel_to_pdf(temp_excel)

def generate_bca_hons_3rd_year_tr_pdf(students, college, batch_uid=None, year="3", exam_name=None):
    generator = BCAHons3rdYearResultGenerator(students, college, batch_uid, year, exam_name)
    return generator.generate()

def generate_static_bca_hons_3rd_year_tr_pdf():
    template_path = os.path.join(settings.BASE_DIR, "bca_hons_year", "static", "tr", "BCA_Hons_3rd_year_static.xlsx")
    if not os.path.exists(template_path): return None
    temp_dir = os.environ.get("TEMP", "/tmp")
    temp_excel = os.path.join(temp_dir, f"BCA_Hons_3rd_year_{uuid.uuid4().hex}.xlsx")
    shutil.copy(template_path, temp_excel)
    wb = load_workbook(temp_excel)
    for ws in wb.worksheets:
        ws.page_setup.paperSize = 8
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_margins.left = 0.6
        ws.page_margins.right = 0.6
        ws.page_margins.top = 1
        ws.page_margins.bottom = 0.2
        ws.print_area = ws.dimensions
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.scale = None
    wb.save(temp_excel)
    return convert_excel_to_pdf(temp_excel)
