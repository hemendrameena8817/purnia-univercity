from io import BytesIO
import os
import base64
from django.template.loader import render_to_string
from weasyprint import HTML
from django.conf import settings
from django.db.models import Q
from collections import OrderedDict
from ..models import (
    UGStudentProfile, ExamRegistration, StudentCourseAssessment, 
    UGExamCenterMapping, UGExamSchedule, CourseStructure
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def wordwrap_br(text, words_per_line=4):
    """Split a long string into chunks of N words joined by <br>."""
    if not text: return ""
    words = text.split()
    lines = []
    for i in range(0, len(words), words_per_line):
        lines.append(" ".join(words[i:i + words_per_line]))
    return "<br>".join(lines)

def get_sem_integer(sem_str):
    """Helper to convert 'Semester-I', '1st', or '1' to integer 1."""
    if not sem_str: return None
    sem_str = str(sem_str).strip().upper()
    
    # 1. Check if it's already a digit
    if sem_str.isdigit(): return int(sem_str)
    
    # 2. Handle '1ST', '2ND', '3RD', '4TH' etc.
    import re
    digit_match = re.match(r'^(\d+)', sem_str)
    if digit_match:
        return int(digit_match.group(1))

    # 3. Simple Roman Numeral mapping
    roman_map = {'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5, 'VI': 6, 'VII': 7, 'VIII': 8}
    for rom, val in roman_map.items():
        if sem_str.endswith(f"-{rom}") or sem_str.endswith(f" {rom}") or sem_str == rom:
            return val
    return None

def get_base64_image(image_field):
    if not image_field:
        return ""
    try:
        with open(image_field.path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')
    except Exception:
        return ""

def _get_roll_sheet_data(exam, college, department_uid=None):
    """
    Common helper to fetch and process data for both PDF and Excel roll sheets.
    """
    sem_int = get_sem_integer(exam.semester)
    session_str = str(exam.session or "").strip()
    
    # 1. Fetch Exam Registrations using the foreign key
    filters = {
        'student__college': college,
        'exam': exam,
        'status': 'REGISTERED'  # Only show confirmed students
    }
    if department_uid:
        # Use department's UID directly if provided
        filters['student__major_course__uid'] = department_uid

    registrations = ExamRegistration.objects.filter(
        **filters
    ).select_related(
        'student', 'student__batch', 'student__program', 'student__major_course'
    ).prefetch_related('assessment').order_by('student__roll_no', 'student__registration_no')

    if not registrations.exists():
        return None

    # 2. Get Exam Center
    center_name = "-"
    center_mapping = UGExamCenterMapping.objects.filter(
        exam=exam,
        attached_colleges=college
    ).select_related('center').first()
    
    if center_mapping and center_mapping.center:
        center_name = center_mapping.center.name

    # 3. Define Header Categories
    category_order = ['MJC', 'MIC', 'SEC', 'VAC', 'MDC', 'AEC']
    
    college_registrations = ExamRegistration.objects.filter(
        student__college=college,
        sem=sem_int,
        session__iexact=session_str,
    )
    
    found_types = StudentCourseAssessment.objects.filter(
        exam_registrations__in=college_registrations,
        label='ESE-Theory'
    ).values_list('course_type', flat=True).distinct()
    
    found_types = [t.strip().upper() for t in found_types if t]
    active_headers = [cat for cat in category_order if cat in found_types]
    others = sorted([t for t in found_types if t not in category_order])
    active_headers.extend(others)

    # 4. Build Student Row Data
    student_data = []
    for reg in registrations:
        student = reg.student
        assessments = reg.assessment.filter(label='ESE-Theory')
        name_lookup = { (ass.course_type or "").strip().upper(): ass.course_name for ass in assessments }
        row_subjects = [name_lookup.get(h, "-") for h in active_headers]

        student_data.append({
            "name": f"{student.first_name} {student.last_name or ''}".strip().upper(),
            "roll_no": student.roll_no or "-",
            "registration_no": student.registration_no or "-",
            "department_name": student.major_course.name if student.major_course else "-",
            "subjects_marked": row_subjects,
        })
 
    batch_val = ""
    # Removed session and batch completely per request
    batch_name = ""

    return {
        "exam": exam,
        "college_name": college.name,
        "college_code": college.college_code or college.center_code or "-",
        "center_name": center_name,
        "batch_name": batch_name,
        "year": exam.exam_month_year or "-",
        "session": batch_name,
        "active_headers": active_headers,
        "student_data": student_data,
        "registrations": registrations, # Kept for backward compat/advanced use
    }


def generate_ug_roll_sheet_pdf(exam, college, department_uid=None):
    """
    Generates Exam Roll Sheet PDF for UG.
    """
    data = _get_roll_sheet_data(exam, college, department_uid)
    if not data:
        return None

    # PDF-specific subject formatting
    subjects = [{"course_name_html": h, "course_type": h} for h in data['active_headers']]

    # Signature
    base_static_path = os.path.join(settings.BASE_DIR, 'ug', 'static', 'ug', 'images')
    controller_sig_path = os.path.join(base_static_path, 'controller-of-examination-signature.png')
    controller_sig_b64 = ""
    if os.path.exists(controller_sig_path):
        with open(controller_sig_path, 'rb') as f:
            controller_sig_b64 = base64.b64encode(f.read()).decode('utf-8')

    context = {
        **data,
        "subjects": subjects,
        "controller_signature": controller_sig_b64,
    }

    html_string = render_to_string('ug/roll_sheet.html', context)
    buffer = BytesIO()
    HTML(string=html_string, base_url=settings.BASE_DIR).write_pdf(target=buffer)
    
    return buffer.getvalue()


def generate_ug_roll_sheet_excel(exam, college, department_uid=None):
    """
    Generates Exam Roll Sheet in Excel format (.xlsx) for UG.
    """
    data = _get_roll_sheet_data(exam, college, department_uid)
    if not data:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Roll Sheet"
    
    header_font = Font(bold=True, size=12)
    sub_header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # Header section
    ws.merge_cells('A1:J1')
    ws['A1'] = f"PURNEA UNIVERSITY, PURNIA"
    ws['A1'].font = header_font
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:J2')
    ws['A2'] = f"Exam Roll Sheet - {exam.name or 'UG Exam'}"
    ws['A2'].alignment = center_align

    cur_row = 4
    ws.cell(row=cur_row, column=1, value="COLLEGE:").font = sub_header_font
    ws.cell(row=cur_row, column=2, value=data['college_name'])
    ws.cell(row=cur_row, column=5, value="CODE:").font = sub_header_font
    ws.cell(row=cur_row, column=6, value=data['college_code'])
    
    cur_row += 1
    ws.cell(row=cur_row, column=1, value="SESSION:").font = sub_header_font
    ws.cell(row=cur_row, column=2, value=data['session'])
    ws.cell(row=cur_row, column=5, value="YEAR:").font = sub_header_font
    ws.cell(row=cur_row, column=6, value=data['year'])
    
    cur_row += 2
    # Table Header
    headers = ["#", "Student Name", "Roll No", "Reg No", "Department"] + data['active_headers']
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=cur_row, column=col_num, value=header_title)
        cell.font = sub_header_font
        cell.border = thin_border
        cell.alignment = center_align

    # Student Data
    data_row = cur_row + 1
    for idx, student in enumerate(data['student_data'], 1):
        row_values = [
            idx,
            student['name'],
            student['roll_no'],
            student['registration_no'],
            student['department_name'],
        ]
        row_values.extend(student['subjects_marked'])
        
        for col_num, value in enumerate(row_values, 1):
            cell = ws.cell(row=data_row, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = left_align if col_num == 2 else center_align
        
        data_row += 1

    # Adjust column widths
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    for col_idx in range(6, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=cur_row, column=col_idx).column_letter].width = 25

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
