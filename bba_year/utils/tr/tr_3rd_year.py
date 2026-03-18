import os
import math
import shutil
import uuid
from collections import defaultdict

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font

from bba_year.models import BBAStudentCourseAssessment, BBAStudentProfile
from bba_year.models import BBACommonCourseStructure, BBACourseStructure
from django.db.models import Q
from bba_year.utils.tr.grading import determine_overall_result, get_hons_classification
from mba_sem.utils.tr.pdf_converter import convert_excel_to_pdf

# ─────────────────────────────────────────────────────────────────
# TEMPLATE LAYOUT (Matched to MBA dynamic logic, restricted to A-M)
# ─────────────────────────────────────────────────────────────────

# Each block: which row-6 cell holds the code, which column the block starts.
# For BBA Part III, columns I (9), J (10), K (11), L (12) are the subject papers.
# Headers for CIA (Internal Assessment)
CIA_P1_COL_15 = 15
CIA_P2_COL_16 = 16
CIA_TOTAL_P3_COL_21 = 21
CIA_GRAND_TOTAL_COL_22 = 22
OVERALL_TOTAL_COL_23 = 23

# Each block maps Ninth, Tenth, Eleventh, Twelveth papers.
# ESE (Theory) columns: 9, 10, 11, 12.
# CIA (Internal) columns: 17, 18, 19, 20.
SUBJECT_BLOCKS = [
    {"code": "BBA301", "col": 9,  "cia_col": 17},
    {"code": "BBA302", "col": 10, "cia_col": 18},
    {"code": "BBA303", "col": 11, "cia_col": 19},
    {"code": "BBA304", "col": 12, "cia_col": 20},
]

# Column 13 (M) is the Total Marks in Part III.
TOTAL_COL_M = 13
# Column 14 (N) is the Grand Total (Part I + Part II + Part III)
GRAND_TOTAL_COL_N = 14

# Subsidiary Subjects
SUB1_NAME_COL  = 24
SUB1_P1_COL    = 25
SUB1_P2_COL    = 26
SUB1_TOTAL_COL = 27

SUB2_NAME_COL  = 28
SUB2_P1_COL    = 29
SUB2_P2_COL    = 30
SUB2_TOTAL_COL = 31

AGGREGATE_COL_32 = 32
RESULT_COL_33 = 33
REMARKS_COL_34 = 34

DATA_START_ROW    = 13
STUDENTS_PER_PAGE = 5
TEMPLATE_CAPACITY = 5

# ─────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────

def _get_marks(student_id, code, student_map, mark_type='ESE'):
    """Return the BEST ESE or CIA marks for a student + paper code among all attempts."""
    best_res = {"marks": 0.0, "found": False, "absent": False}
    search_labels = []
    if mark_type == 'ESE':
        search_labels = ['ESE', 'THEORY', 'Theory']
    else:
        search_labels = ['CIA', 'IA', 'INTERNAL', 'Internal Assessment']

    found_any = False
    all_absent = True

    for rec in student_map.get(student_id, []):
        if rec.paper_code == code:
            label = str(rec.label or "").upper()
            if any(s.upper() in label for s in search_labels):
                found_any = True
                is_absent = bool(rec.ind_is_absent)
                if not is_absent:
                    all_absent = False
                
                marks = 0.0
                if not is_absent:
                    marks = float(
                        rec.ind_final_marks_obtained
                        if rec.ind_final_marks_obtained is not None
                        else (rec.ind_marks_obtained or 0)
                    )
                
                # Keep the best marks found so far
                if not best_res["found"] or marks >= best_res["marks"]:
                    best_res["marks"] = marks
                    best_res["found"] = True
                    best_res["absent"] = is_absent

    # If they were present in AT LEAST one attempt, we don't show them as AB
    if not all_absent:
        best_res["absent"] = False
        
    return best_res


def _get_previous_year_honours_marks_total(student_id, year_str, student_map, honours_codes, mark_type='ESE'):
    """Calculate the total marks and check if the student was absent in ALL papers for that year/type."""
    total = 0.0
    found_any = False
    all_absent = True
    for code in honours_codes:
        m = _get_marks(student_id, code, student_map, mark_type=mark_type)
        if m["found"]:
            found_any = True
            if not m["absent"]:
                all_absent = False
                total += m["marks"]
    
    # If no records found, we don't treat it as "fully absent", just 0
    if not found_any:
        all_absent = False

    return total, all_absent


def _get_year_meta_generic(year_str, codes, appeared_codes, mark_type='ESE', paper_type=None):
    """Generic fetch Max/Pass for a specific year and paper list."""
    filter_q = Q(year=str(year_str), code__in=codes)
    if paper_type:
        filter_q &= Q(paper_type=paper_type)
        
    valid_codes = set(BBACommonCourseStructure.objects.filter(filter_q).values_list('code', flat=True))
    valid_codes = valid_codes.intersection(appeared_codes)
    
    max_m = 0.0
    pass_m = 0.0
    
    label_filter = Q()
    if mark_type == 'ESE':
        label_filter = Q(label='ESE') | Q(label='THEORY') | Q(label='Theory')
    else:
        label_filter = Q(label='CIA') | Q(label='IA') | Q(label='INTERNAL') | Q(label__icontains='Internal')

    qs = BBACourseStructure.objects.filter(year=str(year_str), course_code__in=valid_codes).filter(label_filter)
    for obj in qs:
        m_val = float(obj.max_marks or 0)
        # Honours need 45%, Subsidiaries need 35%
        default_pct = 0.45 if paper_type == "HONOURS" else 0.35
        p_val = float(obj.min_marks if obj.min_marks else (m_val * default_pct))
        max_m += m_val
        pass_m += p_val
    return max_m, pass_m


def _get_year_honours_meta(year_str, codes, appeared_codes, mark_type='ESE'):
    return _get_year_meta_generic(year_str, codes, appeared_codes, mark_type, paper_type='HONOURS')


# ─────────────────────────────────────────────────────────────────
# DATA FILLER
# ─────────────────────────────────────────────────────────────────

# Logic moved to grading.py

def fill_students(ws, students, student_map, honours_map, sub_map, sub_names, appeared_codes, target_year):
    """
    Fill student data into columns A to M (1 to 13).
    
    Layout:
        A (1) : Roll No
        B (2) : Name of Candidate
        C (3) : Registration No
        D (4) : Roll No & Year of Part I
        E (5) : Roll No & Year of Part II
        F (6) : Name of the Honours Subject
        G (7) : Part I Marks Obtained
        H (8) : Part II Marks Obtained
        I (9) : Ninth Paper
        J (10): Tenth Paper
        K (11): Eleventh Paper
        L (12): Twelfth Paper
        M (13): Total Marks in Part III
    """

    # Fill row 10 (Max) and 11 (Pass) Header (Theory Section)
    h1_max, h1_pass = _get_year_honours_meta("1", honours_map.get("1", []), appeared_codes, mark_type='ESE')
    h2_max, h2_pass = _get_year_honours_meta("2", honours_map.get("2", []), appeared_codes, mark_type='ESE')
    
    ws.cell(row=10, column=7).value = h1_max
    ws.cell(row=11, column=7).value = h1_pass
    ws.cell(row=10, column=8).value = h2_max
    ws.cell(row=11, column=8).value = h2_pass

    p3_total_max = 0
    p3_total_pass = 0
    for blk in SUBJECT_BLOCKS:
        p_max, p_pass = _get_year_honours_meta("3", [blk["code"]], appeared_codes, mark_type='ESE')
        ws.cell(row=10, column=blk["col"]).value = p_max
        ws.cell(row=11, column=blk["col"]).value = p_pass
        p3_total_max += p_max
        p3_total_pass += p_pass

    ws.cell(row=10, column=TOTAL_COL_M).value = p3_total_max
    ws.cell(row=11, column=TOTAL_COL_M).value = p3_total_pass
    
    grand_ese_max = h1_max + h2_max + p3_total_max
    grand_ese_pass = h1_pass + h2_pass + p3_total_pass
    ws.cell(row=10, column=GRAND_TOTAL_COL_N).value = grand_ese_max
    ws.cell(row=11, column=GRAND_TOTAL_COL_N).value = grand_ese_pass

    # ── Fill row 10/11 for Internal Assessment (CIA Section) ─────
    c1_max, c1_pass = _get_year_honours_meta("1", honours_map.get("1", []), appeared_codes, mark_type='CIA')
    c2_max, c2_pass = _get_year_honours_meta("2", honours_map.get("2", []), appeared_codes, mark_type='CIA')
    
    ws.cell(row=10, column=CIA_P1_COL_15).value = c1_max
    ws.cell(row=11, column=CIA_P1_COL_15).value = c1_pass
    ws.cell(row=10, column=CIA_P2_COL_16).value = c2_max
    ws.cell(row=11, column=CIA_P2_COL_16).value = c2_pass

    p3_cia_total_max = 0
    p3_cia_total_pass = 0
    for blk in SUBJECT_BLOCKS:
        pc_max, pc_pass = _get_year_honours_meta("3", [blk["code"]], appeared_codes, mark_type='CIA')
        ws.cell(row=10, column=blk["cia_col"]).value = pc_max
        ws.cell(row=11, column=blk["cia_col"]).value = pc_pass
        p3_cia_total_max += pc_max
        p3_cia_total_pass += pc_pass

    ws.cell(row=10, column=CIA_TOTAL_P3_COL_21).value = p3_cia_total_max
    ws.cell(row=11, column=CIA_TOTAL_P3_COL_21).value = p3_cia_total_pass
    
    grand_cia_max = c1_max + c2_max + p3_cia_total_max
    grand_cia_pass = c1_pass + c2_pass + p3_cia_total_pass
    ws.cell(row=10, column=CIA_GRAND_TOTAL_COL_22).value = grand_cia_max
    ws.cell(row=11, column=CIA_GRAND_TOTAL_COL_22).value = grand_cia_pass

    # Overall Honours Combined (23) Max/Pass
    ws.cell(row=10, column=OVERALL_TOTAL_COL_23).value = grand_ese_max + grand_cia_max
    ws.cell(row=11, column=OVERALL_TOTAL_COL_23).value = grand_ese_pass + grand_cia_pass

    # ── Fill row 10/11 for Subsidiary Subjects ────────────────────
    # SUB 1
    s1_p1_codes = sub_map.get("1", {}).get("p1", [])
    s1_p2_codes = sub_map.get("1", {}).get("p2", [])
    s1_p1_max, s1_p1_pass = _get_year_meta_generic("1", s1_p1_codes, appeared_codes)
    s1_p2_max, s1_p2_pass = _get_year_meta_generic("2", s1_p2_codes, appeared_codes)
    
    ws.cell(row=10, column=SUB1_P1_COL).value = s1_p1_max
    ws.cell(row=11, column=SUB1_P1_COL).value = s1_p1_pass
    ws.cell(row=10, column=SUB1_P2_COL).value = s1_p2_max
    ws.cell(row=11, column=SUB1_P2_COL).value = s1_p2_pass
    ws.cell(row=10, column=SUB1_TOTAL_COL).value = s1_p1_max + s1_p2_max
    ws.cell(row=11, column=SUB1_TOTAL_COL).value = s1_p1_pass + s1_p2_pass

    # SUB 2
    s2_p1_codes = sub_map.get("2", {}).get("p1", [])
    s2_p2_codes = sub_map.get("2", {}).get("p2", [])
    s2_p1_max, s2_p1_pass = _get_year_meta_generic("1", s2_p1_codes, appeared_codes)
    s2_p2_max, s2_p2_pass = _get_year_meta_generic("2", s2_p2_codes, appeared_codes)
    
    ws.cell(row=10, column=SUB2_P1_COL).value = s2_p1_max
    ws.cell(row=11, column=SUB2_P1_COL).value = s2_p1_pass
    ws.cell(row=10, column=SUB2_P2_COL).value = s2_p2_max
    ws.cell(row=11, column=SUB2_P2_COL).value = s2_p2_pass
    ws.cell(row=10, column=SUB2_TOTAL_COL).value = s2_p1_max + s2_p2_max
    ws.cell(row=11, column=SUB2_TOTAL_COL).value = s2_p1_pass + s2_p2_pass
    
    # Aggregate Max/Pass (Col 32)
    ws.cell(row=10, column=AGGREGATE_COL_32).value = (grand_ese_max + grand_cia_max + 
                                                     s1_p1_max + s1_p2_max + 
                                                     s2_p1_max + s2_p2_max)
    ws.cell(row=11, column=AGGREGATE_COL_32).value = (grand_ese_pass + grand_cia_pass + 
                                                     s1_p1_pass + s1_p2_pass + 
                                                     s2_p1_pass + s2_p2_pass)
    
    # Header Names for Subsidiaries (Row 10 or 7 as per screenshot?)
    # Usually we fill it in the student row if they vary, but if constant:
    # ws.cell(row=10, column=SUB1_NAME_COL).value = sub_names.get("1", "")
    # ws.cell(row=10, column=SUB2_NAME_COL).value = sub_names.get("2", "")

    row = DATA_START_ROW
    counts = {
        "total": len(students),
        "dist": 0, "first": 0, "second": 0, "fail": 0, 
        "pending": 0, "absent": 0, "expelled": 0
    }

    for student in students:
        print(f"{student = }")
        # ── Basic Info (A-F) ───────────────────────────────────────
        ws.cell(row=row, column=1).value = student.roll_no
        ws.cell(row=row, column=2).value = student.get_full_name()
        ws.cell(row=row, column=3).value = student.registration_no

        # Roll No / Session placeholders
        ws.cell(row=row, column=4).value = f"{student.roll_no} / {student.session_str}"
        ws.cell(row=row, column=5).value = f"{student.roll_no} / {student.session_str}"
        
        # Honours Subject
        ws.cell(row=row, column=6).value = student.course.name if student.course else ""

        # Helper for showing totals
        def fmt_tot(val, is_ab): return "AB" if is_ab else val

        # Previous Honours ESE (7, 8)
        p1_ese, p1_ese_ab = _get_previous_year_honours_marks_total(student.id, "1", student_map, honours_map.get("1", []), mark_type='ESE')
        p2_ese, p2_ese_ab = _get_previous_year_honours_marks_total(student.id, "2", student_map, honours_map.get("2", []), mark_type='ESE')
        
        ws.cell(row=row, column=7).value = fmt_tot(p1_ese if p1_ese > 0 or p1_ese_ab else (student.part_1_marks or 0), p1_ese_ab)
        ws.cell(row=row, column=8).value = fmt_tot(p2_ese if p2_ese > 0 or p2_ese_ab else (student.part_2_marks or 0), p2_ese_ab)

        # Part III Theory (9-12) and Total (13)
        total_p3_ese = 0.0
        p3_ese_ab = True
        total_p3_cia = 0.0
        p3_cia_ab = True

        for blk in SUBJECT_BLOCKS:
            code = blk["code"]
            
            # ESE portion
            m_ese = _get_marks(student.id, code, student_map, mark_type='ESE')
            val_ese = "AB" if m_ese["absent"] else m_ese["marks"]
            ws.cell(row=row, column=blk["col"]).value = val_ese
            if not m_ese["absent"]:
                total_p3_ese += m_ese["marks"]
                p3_ese_ab = False
            
            # CIA portion
            m_cia = _get_marks(student.id, code, student_map, mark_type='CIA')
            val_cia = "AB" if m_cia["absent"] else m_cia["marks"]
            ws.cell(row=row, column=blk["cia_col"]).value = val_cia
            if not m_cia["absent"]:
                total_p3_cia += m_cia["marks"]
                p3_cia_ab = False

        ws.cell(row=row, column=TOTAL_COL_M).value = fmt_tot(total_p3_ese, p3_ese_ab)
        
        # Grand Total ESE (Part I + II + III)
        grand_ese_val = (p1_ese if not p1_ese_ab else 0) + (p2_ese if not p2_ese_ab else 0) + (total_p3_ese if not p3_ese_ab else 0)
        grand_ese_ab = p1_ese_ab and p2_ese_ab and p3_ese_ab
        ws.cell(row=row, column=GRAND_TOTAL_COL_N).value = fmt_tot(grand_ese_val, grand_ese_ab)

        # Internal Assessment CIA (15-22)
        p1_cia, p1_cia_ab = _get_previous_year_honours_marks_total(student.id, "1", student_map, honours_map.get("1", []), mark_type='CIA')
        p2_cia, p2_cia_ab = _get_previous_year_honours_marks_total(student.id, "2", student_map, honours_map.get("2", []), mark_type='CIA')
        ws.cell(row=row, column=CIA_P1_COL_15).value = fmt_tot(p1_cia, p1_cia_ab)
        ws.cell(row=row, column=CIA_P2_COL_16).value = fmt_tot(p2_cia, p2_cia_ab)
        ws.cell(row=row, column=CIA_TOTAL_P3_COL_21).value = fmt_tot(total_p3_cia, p3_cia_ab)
        
        grand_cia_val = (p1_cia if not p1_cia_ab else 0) + (p2_cia if not p2_cia_ab else 0) + (total_p3_cia if not p3_cia_ab else 0)
        grand_cia_ab = p1_cia_ab and p2_cia_ab and p3_cia_ab
        ws.cell(row=row, column=CIA_GRAND_TOTAL_COL_22).value = fmt_tot(grand_cia_val, grand_cia_ab)

        # Overall Total (23) (Grand ESE + Grand CIA)
        total_hons = grand_ese_val + grand_cia_val
        total_hons_ab = grand_ese_ab and grand_cia_ab
        ws.cell(row=row, column=OVERALL_TOTAL_COL_23).value = fmt_tot(total_hons, total_hons_ab)

        # ── Subsidiary Subjects ──────────────────────────────────────
        # Sub 1 (Part I = Year 1, Part II = Year 2)
        ws.cell(row=row, column=SUB1_NAME_COL).value = sub_names.get("1", "")
        s1_p1_m, s1_p1_ab = _get_previous_year_honours_marks_total(student.id, "1", student_map, sub_map.get("1", {}).get("p1", []), mark_type='ESE')
        s1_p2_m, s1_p2_ab = _get_previous_year_honours_marks_total(student.id, "2", student_map, sub_map.get("1", {}).get("p2", []), mark_type='ESE')
        ws.cell(row=row, column=SUB1_P1_COL).value = fmt_tot(s1_p1_m, s1_p1_ab)
        ws.cell(row=row, column=SUB1_P2_COL).value = fmt_tot(s1_p2_m, s1_p2_ab)
        ws.cell(row=row, column=SUB1_TOTAL_COL).value = fmt_tot(s1_p1_m + s1_p2_m, s1_p1_ab and s1_p2_ab)

        # Sub 2
        ws.cell(row=row, column=SUB2_NAME_COL).value = sub_names.get("2", "")
        s2_p1_m, s2_p1_ab = _get_previous_year_honours_marks_total(student.id, "1", student_map, sub_map.get("2", {}).get("p1", []), mark_type='ESE')
        s2_p2_m, s2_p2_ab = _get_previous_year_honours_marks_total(student.id, "2", student_map, sub_map.get("2", {}).get("p2", []), mark_type='ESE')
        ws.cell(row=row, column=SUB2_P1_COL).value = fmt_tot(s2_p1_m, s2_p1_ab)
        ws.cell(row=row, column=SUB2_P2_COL).value = fmt_tot(s2_p2_m, s2_p2_ab)
        ws.cell(row=row, column=SUB2_TOTAL_COL).value = fmt_tot(s2_p1_m + s2_p2_m, s2_p1_ab and s2_p2_ab)

        # Final Aggregate Total (32)
        total_agg = total_hons + (s1_p1_m + s1_p2_m) + (s2_p1_m + s2_p2_m)
        ws.cell(row=row, column=AGGREGATE_COL_32).value = total_agg

        # --- Calculate Results using the new grading.py logic (Years up to current) ---
        hons_marks_data = [] # List of (total_marks, max_marks) for Honours papers
        sub_marks_data = []  # List of (total_marks, max_marks) for Subsidiary papers

        # 1. Collect Honours papers from years 1 to target_year
        years_list = [str(y) for y in range(1, int(target_year) + 1)]
        for yr in years_list:
            for code in honours_map.get(yr, []):
                pc_max_ese, _ = _get_year_honours_meta(yr, [code], appeared_codes, mark_type='ESE')
                pc_max_cia, _ = _get_year_honours_meta(yr, [code], appeared_codes, mark_type='CIA')
                
                if (pc_max_ese + pc_max_cia) > 0:
                    m_ese = _get_marks(student.id, code, student_map, mark_type='ESE')
                    m_cia = _get_marks(student.id, code, student_map, mark_type='CIA')
                    hons_marks_data.append((
                        m_ese["marks"] if not m_ese["absent"] else 0,
                        pc_max_ese,
                        m_cia["marks"] if not m_cia["absent"] else 0,
                        pc_max_cia
                    ))

        # 2. Collect Subsidiary papers from years 1 to target_year (Subs usually only years 1 & 2)
        for sub_key in ["1", "2"]:
            # Part I Subsidiary (p1)
            if "1" in years_list:
                for code in sub_map.get(sub_key, {}).get("p1", []):
                    # For Subsidiary, usually CIA is 0 on the TR template, so we pass 0/0 for CIA here.
                    m_max, _ = _get_year_meta_generic("1", [code], appeared_codes)
                    if m_max > 0:
                        m_val = _get_marks(student.id, code, student_map, mark_type='ESE')
                        sub_marks_data.append((
                            m_val["marks"] if not m_val["absent"] else 0,
                            m_max,
                            0, 0 # CIA marks/max for subsidiaries
                        ))
            
            # Part II Subsidiary (p2)
            if "2" in years_list:
                for code in sub_map.get(sub_key, {}).get("p2", []):
                    m_max, _ = _get_year_meta_generic("2", [code], appeared_codes)
                    if m_max > 0:
                        m_val = _get_marks(student.id, code, student_map, mark_type='ESE')
                        sub_marks_data.append((
                            m_val["marks"] if not m_val["absent"] else 0,
                            m_max,
                            0, 0 # CIA marks/max
                        ))

        overall_res = determine_overall_result(hons_marks_data, sub_marks_data)
        
        # Determine Final Degree Classification (Class) based on Honours aggregate (Final year only)
        total_hons_max = grand_ese_max + grand_cia_max
        class_str = ""
        if target_year == "3":
            class_str = get_hons_classification(total_hons, total_hons_max)
        
        # Display Division/Status in Column 33
        if target_year == "3" and overall_res in ["Pass with Hons.", "PASS"]:
            result_to_show = class_str
        else:
            result_to_show = overall_res
            
        ws.cell(row=row, column=RESULT_COL_33).value = result_to_show
        ws.cell(row=row, column=REMARKS_COL_34).value = ""
        
        # Count Results for footer
        if "Distinction" in class_str: counts["dist"] += 1
        elif "1st Class" in class_str: counts["first"] += 1
        elif "2nd Class" in class_str: counts["second"] += 1
        elif "FAIL" in class_str or overall_res == "FAIL": counts["fail"] += 1

        row += 1

    return counts

def fill_footer(ws, counts, row):
    """Write summary stats at a dynamic row."""
    ws.cell(row=row,     column=1).value = f"No. of Students : {counts['total']}"
    ws.cell(row=row + 1, column=1).value = f"1st Class With Distinction : {counts['dist']}"
    ws.cell(row=row + 2, column=1).value = f"Fail : {counts['fail']}"
    ws.cell(row=row + 3, column=1).value = f"Result Pending : {counts['pending']}"

    ws.cell(row=row + 1, column=5).value = f"1st Class : {counts['first']}"
    ws.cell(row=row + 2, column=5).value = f"Absent : {counts['absent']}"

    ws.cell(row=row + 2, column=9).value = f"Expelled : {counts['expelled']}"
    ws.cell(row=row + 3, column=9).value = f"2nd Class : {counts['second']}"

def _add_logo(ws, college=None):
    """Add the logo to the worksheet. Prefers college logo, fallbacks to university logo."""
    logo_path = None
    
    if college and college.logo:
        try:
            if os.path.exists(college.logo.path):
                logo_path = college.logo.path
        except Exception:
            pass
            
    if not logo_path or not os.path.exists(logo_path):
        logo_path = os.path.join(settings.BASE_DIR, "static/images/purnea-logo.png")

    if os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width = 900
        img.height = 900
        img.anchor = "R1"
        ws.add_image(img)

def fill_header(ws, exam_name, college, center_name=None, subject_name=None):
    print("header_val (A5) = ", exam_name)
    ws["A5"] = exam_name or "BBA Part III Examination"
    
    # Dynamic headers as per user request
    ws["AB3"] = f"Subject : {subject_name or 'BBA HONS'}"
    ws["AB4"] = f"Centre : {center_name}"
    ws["AB5"] = f"Dept./College : {college.name}"
    
    # Also keeping the BC fallback if needed, or remove if AB is the new standard
    # ws["BC4"] = f"Dept./College : {college.name}"

class BBA3rdYearResultGenerator:
    """
    Fills BBA Year TR for columns A through M.
    """
    def __init__(self, students, college, batch_uid=None, year="3", exam_name=None):
        self.students = list(students)
        self.college = college
        self.batch_uid = batch_uid
        self.year = str(year)
        self.exam_name = exam_name

    def generate(self):
        if not self.students: return None
        
        template_path = os.path.join(settings.BASE_DIR, "bba_year", "static", "tr", "BBA_3rd_year.xlsx")
        temp_excel = os.path.join(os.environ.get("TEMP", "/tmp"), f"bba3_{uuid.uuid4().hex}.xlsx")
        shutil.copy(template_path, temp_excel)

        wb = load_workbook(temp_excel)
        master = wb.active
        master.title = "MASTER"

        # Fetch assessment records for all relevant years (1, 2, and 3)
        # Ordered by -id so that _get_marks picks the latest attempt (Back marks)
        qs = BBAStudentCourseAssessment.objects.filter(
            student__in=self.students, 
            year__in=["1", "2", "3"]
        ).order_by('-id')
        if self.batch_uid: qs = qs.filter(batch__uid=self.batch_uid)
        
        student_map = defaultdict(list)
        for obj in qs:
            student_map[obj.student.id].append(obj)
            
        # Honours codes for ALL parts
        honours_map = {
            "1": list(BBACommonCourseStructure.objects.filter(year="1", paper_type="HONOURS").values_list("code", flat=True)),
            "2": list(BBACommonCourseStructure.objects.filter(year="2", paper_type="HONOURS").values_list("code", flat=True)),
            "3": list(BBACommonCourseStructure.objects.filter(year="3", paper_type="HONOURS").values_list("code", flat=True)),
        }

        # Subsidiary Mapping
        # Sub 1 = Year 1 Paper 1 + Year 2 Paper 1 (usually mapped by code order)
        # Sub 2 = Year 1 Paper 2 + Year 2 Paper 2
        y1_subs = sorted(list(BBACommonCourseStructure.objects.filter(year="1", paper_type="SUBSIDIARY").values_list("code", flat=True)))
        y2_subs = sorted(list(BBACommonCourseStructure.objects.filter(year="2", paper_type="SUBSIDIARY").values_list("code", flat=True)))
        
        sub_map = {
            "1": {"p1": [y1_subs[0]] if len(y1_subs) > 0 else [], "p2": [y2_subs[0]] if len(y2_subs) > 0 else []},
            "2": {"p1": [y1_subs[1]] if len(y1_subs) > 1 else [], "p2": [y2_subs[1]] if len(y2_subs) > 1 else []},
        }

        # Names for Sub 1 & 2
        def get_sub_names_combined(codes):
            if not codes: return ""
            names = list(BBACommonCourseStructure.objects.filter(code__in=codes).values_list('course_name', flat=True))
            # Remove duplicates and join
            unique_names = []
            for n in names:
                if n not in unique_names: unique_names.append(n)
            return " / ".join(unique_names)

        sub_names = {
            "1": get_sub_names_combined(sub_map["1"]["p1"] + sub_map["1"]["p2"]),
            "2": get_sub_names_combined(sub_map["2"]["p1"] + sub_map["2"]["p2"]),
        }
        
        # Determine codes actually given in exam by any student in this set
        appeared_codes = set(qs.values_list("paper_code", flat=True))

        # Filter out students who have NO assessment records (implements "if the data is not available the row is not show")
        self.students = [s for s in self.students if any(rec.year == self.year for rec in student_map.get(s.id, []))]
        
        if not self.students: return None

        pages = [self.students[i : i+STUDENTS_PER_PAGE] for i in range(0, len(self.students), STUDENTS_PER_PAGE)]
        for idx, chunk in enumerate(pages):
            ws = wb.copy_worksheet(master)
            ws.title = f"Page_{idx+1}"
            
            counts = fill_students(ws, chunk, student_map, honours_map, sub_map, sub_names, appeared_codes, self.year)
            
            # Delete empty rows to keep the layout tight
            if len(chunk) < TEMPLATE_CAPACITY:
                ws.delete_rows(DATA_START_ROW + len(chunk), TEMPLATE_CAPACITY - len(chunk))
                
            # Fetch Center & Subject
            from bba_year.models import BBAExamCenterMapping, BBAExam
            exam_obj = BBAExam.objects.filter(year=self.year).last()
            center_name = ""
            if exam_obj:
                mapping = BBAExamCenterMapping.objects.filter(exam=exam_obj, attached_colleges=self.college).first()
                if mapping: center_name = mapping.center.name
            
            subject_name = self.students[0].course.name if (self.students and self.students[0].course) else "BBA HONS"

            fill_header(ws, self.exam_name, self.college, center_name=center_name, subject_name=subject_name)
            _add_logo(ws, self.college)
            
            # Dynamic footer placement
            footer_row = DATA_START_ROW + len(chunk) + 3
            fill_footer(ws, counts, footer_row)
            
            # Setup A3 layout
            ws.page_setup.paperSize = 8  # A3
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.print_options.horizontalCentered = True
            ws.print_options.verticalCentered = False
            ws.print_area = "A1:AH40"

            # Set narrow margins to fix the right side space
            ws.page_margins.left = 0.2
            ws.page_margins.right = 0.2
            ws.page_margins.top = 0.2
            ws.page_margins.bottom = 0.2

            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = False
            ws.page_setup.scale = None

        if "MASTER" in wb.sheetnames: wb.remove(wb["MASTER"])
        wb.save(temp_excel)
        return convert_excel_to_pdf(temp_excel)

# Entry point for dynamic generation
def generate_bba_3rd_year_tr_pdf(students, college, batch_uid=None, year="3", exam_name=None):
    generator = BBA3rdYearResultGenerator(students, college, batch_uid, year, exam_name)
    return generator.generate()
