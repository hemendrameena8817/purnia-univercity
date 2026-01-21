"""
Generate comprehensive Excel template for legacy data import.
Simplified version - Covers Colleges, Courses, Students, Marks, Results, and Certificates.

Usage:
    poetry run python data_templates/generate_template.py
"""

import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Please install openpyxl: poetry add openpyxl")
    sys.exit(1)


def create_header_style():
    """Create styles for headers"""
    return {
        'font': Font(bold=True, color='FFFFFF', size=11),
        'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def add_sheet(wb, sheet_name, headers, sample_data=None):
    """Add a sheet with headers and optional sample data"""
    ws = wb.create_sheet(title=sheet_name)
    header_style = create_header_style()
    
    # Add headers
    for col, (header, required, data_type, example, description) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_style['font']
        cell.fill = header_style['fill']
        cell.alignment = header_style['alignment']
        cell.border = header_style['border']
        
        # Set column width based on header length
        ws.column_dimensions[get_column_letter(col)].width = max(len(header) + 5, 15)
    
    # Add sample data if provided
    if sample_data:
        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    return ws


def generate_template():
    """Generate the comprehensive Excel template"""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # ========================================
    # MASTER DATA SHEETS
    # ========================================
    
    # Sheet 1: Colleges/Institutes
    add_sheet(wb, '1_Colleges', [
        ('name', True, 'Text (255)', 'ABC College', 'Full college name'),
        ('short_name', True, 'Text (100)', 'ABCC', 'Short name'),
        ('college_code', True, 'Text (50)', 'COL001', 'Unique college code'),
        ('address', True, 'Text', 'Main Road, Purnea', 'Full address'),
        ('principal', True, 'Text (255)', 'Dr. Shyam Das', 'Principal name'),
        ('contact_no', False, 'Text (15)', '9876543211', 'Contact number'),
        ('email', False, 'Email', 'abc@college.edu', 'College email'),
        ('founded', False, 'Date', '1990-07-01', 'YYYY-MM-DD format'),
        ('website', False, 'URL', 'https://abccollege.edu', 'Website'),
    ], [
        ('ABC College of Arts and Science', 'ABCC', 'COL001', 'Main Road, Purnea', 'Dr. Shyam Das', '9876543211', 'abc@college.edu', '1990-07-01', 'https://abccollege.edu'),
        ('XYZ Degree College', 'XYZDC', 'COL002', 'Station Road, Purnea', 'Dr. Ram Prasad', '9876543222', 'xyz@college.edu', '1985-06-01', ''),
    ])
    
    # Sheet 2: Faculties
    add_sheet(wb, '2_Faculties', [
        ('name', True, 'Text (255)', 'Faculty of Science', 'Full faculty name'),
        ('short_name', False, 'Text (100)', 'FoS', 'Short name'),
        ('description', False, 'Text', 'Science division', 'Description'),
    ], [
        ('Faculty of Science', 'FoS', 'Science and Technology Education'),
        ('Faculty of Commerce', 'FoC', 'Commerce and Business Education'),
        ('Faculty of Humanities', 'FoH', 'Arts and Humanities Education'),
    ])
    
    # Sheet 3: Departments
    add_sheet(wb, '3_Departments', [
        ('name', True, 'Text (255)', 'History', 'Full department name'),
        ('code', True, 'Text (50)', 'CS', 'Unique department code'),
        ('faculty_name', True, 'Text', 'Faculty of Science', 'Reference to faculty'),
    ], [
        ('History', 'CS', 'Faculty of Science'),
        ('Economics', 'COM', 'Faculty of Commerce'),
        ('Physics', 'ENG', 'Faculty of Humanities'),
        ('Mathematics', 'MATH', 'Faculty of Science'),
    ])
    
    # Sheet 4: Degrees
    add_sheet(wb, '4_Degrees', [
        ('name', True, 'Text (255)', 'Bachelor of Computer Applications', 'Full degree name'),
        ('degree_level', True, 'Text', 'UG', 'UG (Undergraduate) or PG (Postgraduate)'),
        ('total_semesters', True, 'Integer', '6', 'Total number of semesters'),
        ('total_years', True, 'Integer', '3', 'Duration in years'),
    ], [
        ('Bachelor of Computer Applications', 'UG', 6, 3),
        ('Bachelor of Commerce', 'UG', 6, 3),
        ('Bachelor of Arts', 'UG', 6, 3),
        ('Bachelor of Business Administration', 'UG', 6, 3),
        ('Master of Business Administration', 'PG', 4, 2),
        ('Master of Computer Applications', 'PG', 4, 2),
    ])
    
    # Sheet 5: Programs
    add_sheet(wb, '5_Programs', [
        ('name', True, 'Text (255)', 'BCA (Honours)', 'Full program name'),
        ('short_name', True, 'Text (50)', 'BCA', 'Short name'),
        ('degree_name', True, 'Text', 'Bachelor of Computer Applications', 'Reference to degree'),
        ('department_code', False, 'Text', 'CS', 'Reference to department'),
    ], [
        ('BCA (Honours)', 'BCA', 'Bachelor of Computer Applications', 'CS'),
        ('B.Com (Honours)', 'BCOM', 'Bachelor of Commerce', 'COM'),
        ('BA (Honours) English', 'BA-ENG', 'Bachelor of Arts', 'ENG'),
        ('BBA', 'BBA', 'Bachelor of Business Administration', 'COM'),
        ('MBA', 'MBA', 'Master of Business Administration', 'COM'),
        ('MCA', 'MCA', 'Master of Computer Applications', 'CS'),
    ])
    
    # Sheet 6: Batches
    add_sheet(wb, '6_Batches', [
        ('name', True, 'Text (50)', '2024-2028', 'Batch name (admission-graduation)'),
        ('start_year', True, 'Integer', '2024', 'Admission year'),
        ('end_year', True, 'Integer', '2028', 'Expected graduation year'),
        ('is_active', True, 'Boolean', 'TRUE', 'Is batch currently active?'),
    ], [
        ('2024-2028', 2024, 2028, 'TRUE'),
        ('2023-2027', 2023, 2027, 'TRUE'),
        ('2022-2026', 2022, 2026, 'TRUE'),
        ('2021-2025', 2021, 2025, 'TRUE'),
        ('2020-2024', 2020, 2024, 'FALSE'),
        ('2019-2023', 2019, 2023, 'FALSE'),
    ])
    
    # Sheet 7: Sessions
    add_sheet(wb, '7_Sessions', [
        ('name', True, 'Text (50)', '2024-2025', 'Academic session'),
        ('start_date', True, 'Date', '2024-07-01', 'Session start date'),
        ('end_date', True, 'Date', '2025-06-30', 'Session end date'),
        ('is_current', True, 'Boolean', 'TRUE', 'Is current session?'),
    ], [
        ('2024-2025', '2024-07-01', '2025-06-30', 'TRUE'),
        ('2023-2024', '2023-07-01', '2024-06-30', 'FALSE'),
        ('2022-2023', '2022-07-01', '2023-06-30', 'FALSE'),
    ])
    
    # Sheet 8: Exam Centres
    add_sheet(wb, '8_ExamCentres', [
        ('centre_code', True, 'Text (50)', 'EC001', 'Unique exam centre code'),
        ('centre_name', True, 'Text (255)', 'ABC College Exam Centre', 'Exam centre name'),
        ('address', True, 'Text', 'Main Road, Purnea', 'Full address'),
        ('city', True, 'Text (100)', 'Purnea', 'City'),
        ('district', True, 'Text (100)', 'Purnea', 'District'),
        ('state', True, 'Text (100)', 'Bihar', 'State'),
        ('pincode', False, 'Text (10)', '854301', 'PIN code'),
        ('college_code', False, 'Text', 'COL001', 'Reference to college (if centre is at a college)'),
        ('contact_person', False, 'Text (255)', 'Dr. A.K. Singh', 'Centre in-charge name'),
        ('contact_phone', False, 'Text (15)', '9876543210', 'Contact phone'),
        ('contact_email', False, 'Email', 'ecabc@college.edu', 'Contact email'),
        ('seating_capacity', False, 'Integer', '500', 'Total seating capacity'),
        ('is_active', True, 'Boolean', 'TRUE', 'Is centre active?'),
    ], [
        ('EC001', 'ABC College Exam Centre', 'Main Road, Purnea', 'Purnea', 'Purnea', 'Bihar', '854301', 'COL001', 'Dr. A.K. Singh', '9876543210', 'ecabc@college.edu', 500, 'TRUE'),
        ('EC002', 'XYZ Degree College Exam Centre', 'Station Road, Purnea', 'Purnea', 'Purnea', 'Bihar', '854302', 'COL002', 'Dr. B.K. Verma', '9876543211', 'ecxyz@college.edu', 400, 'TRUE'),
        ('EC003', 'Government School Exam Centre', 'Civil Lines, Purnea', 'Purnea', 'Purnea', 'Bihar', '854301', '', 'Mr. R.K. Jha', '9876543212', '', 300, 'TRUE'),
    ])
    
    # ========================================
    # COURSE DATA (Simplified - All info in one sheet)
    # ========================================
    
    # Sheet 9: Courses/Subjects (with all course structure info)
    add_sheet(wb, '9_Courses', [
        ('name', True, 'Text (255)', 'Introduction to Programming', 'Course/Subject name'),
        ('code', True, 'Text (50)', 'BCA101', 'Unique course code'),
        ('course_slot', True, 'Text (20)', 'MJC-1', 'Course slot (MJC-1, MNC-2, SEC-1, etc.)'),
        ('semester', True, 'Integer', '1', 'Semester number'),
        ('credits', True, 'Integer', '4', 'Course credits'),
        ('cia_marks', True, 'Integer', '25', 'CIA/Internal marks'),
        ('final_marks', True, 'Integer', '75', 'Final exam marks'),
        ('total_marks', True, 'Integer', '100', 'Total marks'),
        ('description', False, 'Text', 'Basics of programming', 'Description'),
        ('department_code', False, 'Text', 'CS', 'Reference to department'),
        ('program_short_name', True, 'Text', 'BCA', 'Reference to program'),
        ('is_elective', False, 'Boolean', 'FALSE', 'Is this an elective?'),
        ('is_active', False, 'Boolean', 'TRUE', 'Is course active?'),
    ], [
        ('Introduction to Programming', 'BCA101', 'MJC-1', 1, 4, 25, 75, 100, 'Fundamentals of programming using C', 'CS', 'BCA', 'FALSE', 'TRUE'),
        ('Digital Electronics', 'BCA102', 'MJC-2', 1, 4, 25, 75, 100, 'Basics of digital systems', 'CS', 'BCA', 'FALSE', 'TRUE'),
        ('Mathematics-I', 'BCA103', 'MNC-1', 1, 4, 25, 75, 100, 'Calculus and Linear Algebra', 'MATH', 'BCA', 'FALSE', 'TRUE'),
        ('Communication Skills', 'AEC101', 'AEC-1', 1, 2, 15, 35, 50, 'English communication', 'ENG', 'BCA', 'FALSE', 'TRUE'),
        ('Environmental Studies', 'VAC101', 'VAC-1', 1, 2, 15, 35, 50, 'Environment awareness', '', 'BCA', 'FALSE', 'TRUE'),
        ('Data Structures', 'BCA201', 'MJC-3', 2, 4, 25, 75, 100, 'Data structures and algorithms', 'CS', 'BCA', 'FALSE', 'TRUE'),
        ('Computer Networks', 'BCA202', 'MJC-4', 2, 4, 25, 75, 100, 'Networking fundamentals', 'CS', 'BCA', 'FALSE', 'TRUE'),
        ('Financial Accounting', 'BCOM101', 'MJC-1', 1, 4, 25, 75, 100, 'Fundamentals of accounting', 'COM', 'BCOM', 'FALSE', 'TRUE'),
        ('Business Economics', 'BCOM102', 'MJC-2', 1, 4, 25, 75, 100, 'Micro economics for business', 'COM', 'BCOM', 'FALSE', 'TRUE'),
    ])
    
    # ========================================
    # STUDENT DATA
    # ========================================
    
    # Sheet 10: Students
    add_sheet(wb, '10_Students', [
        ('first_name', True, 'Text (255)', 'Rahul', 'First name'),
        ('last_name', True, 'Text (255)', 'Kumar', 'Last name'),
        ('email', True, 'Email', 'rahul@email.com', 'Email (for login)'),
        ('phone', False, 'Text (15)', '9876543212', 'Phone number'),
        ('registration_no', True, 'Text (50)', 'PUP2024001', 'Unique registration number'),
        ('roll_no', True, 'Text (50)', '2024BCA001', 'Unique roll number'),
        ('date_of_birth', True, 'Date', '2002-05-15', 'YYYY-MM-DD format'),
        ('gender', True, 'Text', 'Male', 'Male, Female, or Other'),
        ('address', True, 'Text', 'Village ABC, Purnea', 'Full address'),
        ('father_name', True, 'Text (255)', 'Shri Ram Kumar', "Father's name"),
        ('mother_name', True, 'Text (255)', 'Smt. Sita Devi', "Mother's name"),
        ('admission_date', True, 'Date', '2024-08-01', 'Date of admission'),
        ('enrollment_date', True, 'Date', '2024-08-15', 'Date of enrollment'),
        ('batch', True, 'Text (50)', '2024-2028', 'Batch name'),
        ('session', True, 'Text (50)', '2024-2025', 'Current session'),
        ('current_semester', True, 'Integer', '1', 'Current semester'),
        ('status', True, 'Text', 'Active', 'Active, Suspended, or Alumni'),
        ('college_code', True, 'Text', 'COL001', 'Reference to college'),
        ('department_code', False, 'Text', 'CS', 'Reference to department'),
        ('program_short_name', True, 'Text', 'BCA', 'Reference to program'),
    ], [
        ('Rahul', 'Kumar', 'rahul@email.com', '9876543212', 'PUP2024001', '2024BCA001', '2002-05-15', 'Male', 'Village ABC, Purnea', 'Shri Ram Kumar', 'Smt. Sita Devi', '2024-08-01', '2024-08-15', '2024-2028', '2024-2025', 1, 'Active', 'COL001', 'CS', 'BCA'),
        ('Priya', 'Singh', 'priya@email.com', '9876543213', 'PUP2024002', '2024BCA002', '2003-03-20', 'Female', 'Town XYZ, Purnea', 'Shri Vijay Singh', 'Smt. Radha Singh', '2024-08-01', '2024-08-15', '2024-2028', '2024-2025', 1, 'Active', 'COL001', 'CS', 'BCA'),
        ('Amit', 'Sharma', 'amit@email.com', '9876543214', 'PUP2024003', '2024BCOM001', '2002-11-10', 'Male', 'Block PQR, Purnea', 'Shri Suresh Sharma', 'Smt. Meena Sharma', '2024-08-01', '2024-08-15', '2024-2028', '2024-2025', 1, 'Active', 'COL001', 'COM', 'BCOM'),
    ])
    
    # ========================================
    # MARKS DATA (Separate CIA and Final)
    # ========================================
    
    # Sheet 11: CIA Marks (Continuous Internal Assessment)
    add_sheet(wb, '11_CIAMarks', [
        ('registration_no', True, 'Text', 'PUP2024001', 'Student registration number'),
        ('email', True, 'Email', 'rahul@email.com', 'Student email'),
        ('session', True, 'Text', '2024-2025', 'Academic session'),
        ('semester', True, 'Integer', '1', 'Semester number'),
        ('course_code', True, 'Text', 'BCA101', 'Course code'),
        ('course_slot', True, 'Text', 'MJC-1', 'Course slot'),
        ('theory_marks', False, 'Integer', '12', 'Theory CIA marks obtained'),
        ('theory_max', False, 'Integer', '15', 'Theory CIA max marks'),
        ('practical_marks', False, 'Integer', '8', 'Practical CIA marks obtained'),
        ('practical_max', False, 'Integer', '10', 'Practical CIA max marks'),
        ('total_marks_obtained', True, 'Integer', '20', 'Total CIA marks obtained'),
        ('max_marks', True, 'Integer', '25', 'Maximum CIA marks'),
        ('grade_points', False, 'Decimal', '8.0', 'Grade points (0-10)'),
        ('grade_letter', False, 'Text', 'A', 'Grade letter (O/A+/A/B+/B/C/D/F)'),
        ('grade_description', False, 'Text', 'Excellent', 'Grade description'),
        ('status', True, 'Text', 'Completed', 'Completed/Pending/Absent'),
    ], [
        ('PUP2024001', 'rahul@email.com', '2024-2025', 1, 'BCA101', 'MJC-1', 12, 15, 8, 10, 20, 25, 8.0, 'A', 'Excellent', 'Completed'),
        ('PUP2024001', 'rahul@email.com', '2024-2025', 1, 'BCA102', 'MJC-2', 10, 15, 8, 10, 18, 25, 7.0, 'B+', 'Very Good', 'Completed'),
        ('PUP2024001', 'rahul@email.com', '2024-2025', 1, 'BCA103', 'MNC-1', 13, 15, 9, 10, 22, 25, 9.0, 'A+', 'Outstanding', 'Completed'),
        ('PUP2024001', 'rahul@email.com', '2024-2025', 1, 'AEC101', 'AEC-1', 12, 15, 0, 0, 12, 15, 8.0, 'A', 'Excellent', 'Completed'),
        ('PUP2024002', 'priya@email.com', '2024-2025', 1, 'BCA101', 'MJC-1', 14, 15, 9, 10, 23, 25, 9.0, 'A+', 'Outstanding', 'Completed'),
    ])
    
    # Sheet 12: Final Marks (End Semester Examination)
    add_sheet(wb, '12_FinalMarks', [
        ('registration_no', True, 'Text', 'PUP2024001', 'Student registration number'),
        ('email', True, 'Email', 'rahul@email.com', 'Student email'),
        ('session', True, 'Text', '2024-2025', 'Academic session'),
        ('semester', True, 'Integer', '1', 'Semester number'),
        ('exam_type', True, 'Text', 'Regular', 'Regular/BACK'),
        ('course_code', True, 'Text', 'BCA101', 'Course code'),
        ('course_slot', True, 'Text', 'MJC-1', 'Course slot'),
        ('theory_marks', False, 'Integer', '50', 'Theory marks obtained'),
        ('theory_max', False, 'Integer', '60', 'Theory max marks'),
        ('practical_marks', False, 'Integer', '12', 'Practical marks obtained'),
        ('practical_max', False, 'Integer', '15', 'Practical max marks'),
        ('total_final_obtained', True, 'Integer', '60', 'Total final marks obtained'),
        ('total_final_max', True, 'Integer', '75', 'Maximum final marks'),
        ('grade_points', False, 'Decimal', '8.0', 'Grade points (0-10)'),
        ('grade_letter', False, 'Text', 'A', 'Grade letter (O/A+/A/B+/B/C/D/F)'),
        ('grade_description', False, 'Text', 'Excellent', 'Grade description'),
        ('exam_date', False, 'Date', '2024-12-15', 'Exam date'),
        ('status', True, 'Text', 'Pass', 'Pass/Fail/Absent'),
    ], [
        ('PUP2024001', 'rahul@email.com', '2024-2025', 1, 'Regular', 'BCA101', 'MJC-1', 50, 60, 12, 15, 60, 75, 8.0, 'A', 'Excellent', '2024-12-15', 'Pass'),
        ('PUP2024001', 'rahul@email.com', '2024-2025', 1, 'Regular', 'BCA102', 'MJC-2', 45, 60, 10, 15, 54, 75, 7.0, 'B+', 'Very Good', '2024-12-17', 'Pass'),
        ('PUP2024001', 'rahul@email.com', '2024-2025', 1, 'Regular', 'BCA103', 'MNC-1', 55, 60, 14, 15, 66, 75, 9.0, 'A+', 'Outstanding', '2024-12-19', 'Pass'),
        ('PUP2024002', 'priya@email.com', '2024-2025', 1, 'Regular', 'BCA101', 'MJC-1', 58, 60, 14, 15, 67, 75, 9.0, 'A+', 'Outstanding', '2024-12-15', 'Pass'),
        ('PUP2024003', 'amit@email.com', '2024-2025', 2, 'BACK', 'BCOM101', 'MJC-1', 42, 60, 10, 15, 50, 75, 6.0, 'B', 'Good', '2025-06-15', 'Pass'),
    ])
    
    # ========================================
    # RESULTS DATA
    # ========================================
    
    # Sheet 13: Semester Results (Aggregated)
    add_sheet(wb, '13_SemesterResults', [
        ('registration_no', True, 'Text', 'PUP2024001', 'Student registration number'),
        ('email', True, 'Email', 'rahul@email.com', 'Student email'),
        ('session', True, 'Text', '2024-2025', 'Academic session'),
        ('semester', True, 'Integer', '1', 'Semester number'),
        ('exam_type', True, 'Text', 'Regular', 'Regular/BACK'),
        ('total_credits', True, 'Integer', '22', 'Total credits in semester'),
        ('credits_earned', True, 'Integer', '22', 'Credits earned'),
        ('total_cia_obtained', True, 'Integer', '72', 'Total CIA marks'),
        ('total_cia_max', True, 'Integer', '90', 'Max CIA marks'),
        ('total_final_obtained', True, 'Integer', '203', 'Total final marks'),
        ('total_final_max', True, 'Integer', '260', 'Max final marks'),
        ('total_marks_obtained', True, 'Integer', '275', 'Grand total marks'),
        ('total_max_marks', True, 'Integer', '350', 'Grand total max'),
        ('percentage', True, 'Decimal', '78.57', 'Percentage'),
        ('sgpa', True, 'Decimal', '7.85', 'Semester GPA'),
        ('grade_letter', False, 'Text', 'A', 'Overall grade letter'),
        ('grade_description', False, 'Text', 'Excellent', 'Overall grade description'),
        ('result_status', True, 'Text', 'Pass', 'Pass/Fail/Promoted'),
        ('result_date', False, 'Date', '2025-02-15', 'Result declaration date'),
        ('remarks', False, 'Text', '', 'Any remarks'),
    ], [
        ('PUP2024001', 'rahul@email.com', '2024-2025', 1, 'Regular', 22, 22, 72, 90, 203, 260, 275, 350, 78.57, 7.85, 'A', 'Excellent', 'Pass', '2025-02-15', ''),
        ('PUP2024002', 'priya@email.com', '2024-2025', 1, 'Regular', 22, 22, 82, 90, 230, 260, 312, 350, 89.14, 8.90, 'A+', 'Outstanding', 'Pass', '2025-02-15', ''),
        ('PUP2024003', 'amit@email.com', '2024-2025', 1, 'Regular', 22, 18, 55, 90, 165, 260, 220, 350, 62.86, 6.20, 'B', 'Good', 'Pass', '2025-02-15', 'Backlog in 2 subjects'),
        ('PUP2024003', 'amit@email.com', '2024-2025', 2, 'BACK', 22, 22, 60, 90, 180, 260, 240, 350, 68.57, 6.80, 'B+', 'Very Good', 'Pass', '2025-08-15', 'Cleared backlog'),
    ])
    
    # Sheet 14: Overall Results (CGPA/Final)
    add_sheet(wb, '14_OverallResults', [
        ('registration_no', True, 'Text', 'PUP2024001', 'Student registration number'),
        ('email', True, 'Email', 'rahul@email.com', 'Student email'),
        ('program_short_name', True, 'Text', 'BCA', 'Program'),
        ('batch', True, 'Text', '2024-2028', 'Batch'),
        ('total_semesters_completed', True, 'Integer', '6', 'Semesters completed'),
        ('total_credits', True, 'Integer', '132', 'Total credits'),
        ('credits_earned', True, 'Integer', '132', 'Credits earned'),
        ('total_marks_obtained', True, 'Integer', '2850', 'Total marks'),
        ('total_max_marks', True, 'Integer', '3300', 'Maximum marks'),
        ('overall_percentage', True, 'Decimal', '86.36', 'Overall percentage'),
        ('cgpa', True, 'Decimal', '8.50', 'Cumulative GPA'),
        ('grade_letter', True, 'Text', 'A+', 'Final grade letter'),
        ('grade_description', True, 'Text', 'Outstanding', 'Final grade description'),
        ('division', True, 'Text', 'First Division with Distinction', 'Division'),
        ('final_result', True, 'Text', 'Pass', 'Pass/Fail'),
        ('completion_date', False, 'Date', '2027-06-30', 'Course completion date'),
    ], [
        ('PUP2021001', 'rahul@email.com', 'BCA', '2021-2025', 6, 132, 132, 2850, 3300, 86.36, 8.50, 'A+', 'Outstanding', 'First Division with Distinction', 'Pass', '2024-06-30'),
        ('PUP2021002', 'priya@email.com', 'BCA', '2021-2025', 6, 132, 130, 2650, 3300, 80.30, 7.80, 'A', 'Excellent', 'First Division', 'Pass', '2024-06-30'),
    ])
    
    # ========================================
    # TABULATION REGISTER DATA
    # ========================================
    
    # Sheet 15: TR Data (Tabulation Register)
    add_sheet(wb, '15_TabulationRegister', [
        ('registration_no', True, 'Text', 'PUP2024001', 'Student registration number'),
        ('email', True, 'Email', 'rahul@email.com', 'Student email'),
        ('roll_no', True, 'Text', '2024BCA001', 'Roll number'),
        ('student_name', True, 'Text', 'Rahul Kumar', 'Full name'),
        ('father_name', True, 'Text', 'Shri Ram Kumar', "Father's name"),
        ('program_short_name', True, 'Text', 'BCA', 'Program'),
        ('session', True, 'Text', '2024-2025', 'Session'),
        ('semester', True, 'Integer', '3', 'Current semester'),
        ('exam_type', True, 'Text', 'Regular', 'Regular/BACK'),
        ('college_code', True, 'Text', 'COL001', 'College code'),
        # Course 1
        ('course1_code', False, 'Text', 'BCA301', 'Course 1 code'),
        ('course1_cia', False, 'Integer', '20', 'Course 1 CIA marks'),
        ('course1_theory', False, 'Integer', '45', 'Course 1 theory marks'),
        ('course1_practical', False, 'Integer', '15', 'Course 1 practical marks'),
        ('course1_total', False, 'Integer', '80', 'Course 1 total marks'),
        # Course 2
        ('course2_code', False, 'Text', 'BCA302', 'Course 2 code'),
        ('course2_cia', False, 'Integer', '18', 'Course 2 CIA marks'),
        ('course2_theory', False, 'Integer', '42', 'Course 2 theory marks'),
        ('course2_practical', False, 'Integer', '12', 'Course 2 practical marks'),
        ('course2_total', False, 'Integer', '72', 'Course 2 total marks'),
        # Course 3
        ('course3_code', False, 'Text', 'BCA303', 'Course 3 code'),
        ('course3_cia', False, 'Integer', '22', 'Course 3 CIA marks'),
        ('course3_theory', False, 'Integer', '50', 'Course 3 theory marks'),
        ('course3_practical', False, 'Integer', '16', 'Course 3 practical marks'),
        ('course3_total', False, 'Integer', '88', 'Course 3 total marks'),
        # Course 4
        ('course4_code', False, 'Text', 'AEC301', 'Course 4 code'),
        ('course4_cia', False, 'Integer', '12', 'Course 4 CIA marks'),
        ('course4_theory', False, 'Integer', '23', 'Course 4 theory marks'),
        ('course4_practical', False, 'Integer', '0', 'Course 4 practical marks'),
        ('course4_total', False, 'Integer', '35', 'Course 4 total marks'),
        # Course 5
        ('course5_code', False, 'Text', 'SEC301', 'Course 5 code'),
        ('course5_cia', False, 'Integer', '10', 'Course 5 CIA marks'),
        ('course5_theory', False, 'Integer', '15', 'Course 5 theory marks'),
        ('course5_practical', False, 'Integer', '15', 'Course 5 practical marks'),
        ('course5_total', False, 'Integer', '40', 'Course 5 total marks'),
        # Current Semester Totals
        ('current_sem_total_marks', True, 'Integer', '315', 'Current sem total marks'),
        ('current_sem_credits', True, 'Integer', '22', 'Current sem credits'),
        ('current_sem_sgpa', True, 'Decimal', '8.15', 'Current sem SGPA'),
        # Previous Semester Credits (Cumulative)
        ('sem1_credits_earned', False, 'Integer', '22', 'Credits earned in Sem 1'),
        ('sem2_credits_earned', False, 'Integer', '22', 'Credits earned in Sem 2'),
        ('sem3_credits_earned', False, 'Integer', '22', 'Credits earned in Sem 3'),
        ('sem4_credits_earned', False, 'Integer', '22', 'Credits earned in Sem 4'),
        ('sem5_credits_earned', False, 'Integer', '22', 'Credits earned in Sem 5'),
        ('sem6_credits_earned', False, 'Integer', '0', 'Credits earned in Sem 6'),
        # Cumulative
        ('total_credits_earned', True, 'Integer', '66', 'Total cumulative credits'),
        ('cgpa', True, 'Decimal', '7.85', 'Cumulative GPA'),
        # Result
        ('result', True, 'Text', 'Pass', 'Pass/Fail/Promoted'),
        ('remarks', False, 'Text', '', 'Any remarks'),
    ], [
        ('PUP2022001', 'rahul@email.com', '2022BCA001', 'Rahul Kumar', 'Shri Ram Kumar', 'BCA', '2024-2025', 3, 'Regular', 'COL001', 
         'BCA301', 20, 45, 15, 80, 'BCA302', 18, 42, 12, 72, 'BCA303', 22, 50, 16, 88, 'AEC301', 12, 23, 0, 35, 'SEC301', 10, 15, 15, 40, 
         315, 22, 8.15, 22, 22, 22, 0, 0, 0, 66, 7.85, 'Pass', ''),
        ('PUP2022002', 'priya@email.com', '2022BCA002', 'Priya Singh', 'Shri Vijay Singh', 'BCA', '2024-2025', 3, 'Regular', 'COL001', 
         'BCA301', 23, 52, 15, 90, 'BCA302', 21, 48, 15, 84, 'BCA303', 24, 52, 16, 92, 'AEC301', 14, 28, 0, 42, 'SEC301', 12, 18, 15, 45, 
         353, 22, 8.65, 22, 22, 22, 0, 0, 0, 66, 8.40, 'Pass', ''),
        ('PUP2021001', 'amit@email.com', '2021BCA001', 'Amit Sharma', 'Shri Suresh Sharma', 'BCA', '2024-2025', 5, 'Regular', 'COL001', 
         'BCA501', 18, 40, 12, 70, 'BCA502', 16, 38, 10, 64, 'BCA503', 20, 45, 13, 78, 'AEC501', 10, 20, 0, 30, 'SEC501', 8, 12, 12, 32, 
         274, 22, 7.20, 22, 22, 22, 22, 22, 0, 110, 7.50, 'Promoted', 'Backlog in 1 subject'),
    ])
    
    # ========================================
    # CERTIFICATE DATA
    # ========================================
    
    # Sheet 16: Provisional Certificate Data
    add_sheet(wb, '16_ProvisionalCertificate', [
        ('registration_no', True, 'Text', 'PUP2024001', 'Student registration number'),
        ('roll_no', True, 'Text', '2024BCA001', 'Roll number'),
        ('student_name', True, 'Text', 'Rahul Kumar', 'Full name'),
        ('father_name', True, 'Text', 'Shri Ram Kumar', "Father's name"),
        ('mother_name', True, 'Text', 'Smt. Sita Devi', "Mother's name"),
        ('date_of_birth', True, 'Date', '2002-05-15', 'Date of birth'),
        ('program_name', True, 'Text', 'Bachelor of Computer Applications', 'Program full name'),
        ('program_short_name', True, 'Text', 'BCA', 'Program short name'),
        ('college_name', True, 'Text', 'ABC College of Arts and Science', 'College name'),
        ('batch', True, 'Text', '2021-2024', 'Batch'),
        ('passing_session', True, 'Text', '2023-2024', 'Passing session'),
        ('passing_year', True, 'Integer', '2024', 'Passing year'),
        ('cgpa', True, 'Decimal', '8.50', 'CGPA'),
        ('percentage', True, 'Decimal', '86.36', 'Percentage'),
        ('division', True, 'Text', 'First Division with Distinction', 'Division'),
        ('result', True, 'Text', 'Pass', 'Pass/Fail'),
        ('certificate_no', False, 'Text', 'PUP/PROV/2024/001', 'Certificate number'),
        ('issue_date', False, 'Date', '2024-07-15', 'Issue date'),
        ('status', True, 'Text', 'Pending', 'Pending/Verified/Approved/Issued'),
    ], [
        ('PUP2021001', '2021BCA001', 'Rahul Kumar', 'Shri Ram Kumar', 'Smt. Sita Devi', '2002-05-15', 'Bachelor of Computer Applications', 'BCA', 'ABC College of Arts and Science', '2021-2024', '2023-2024', 2024, 8.50, 86.36, 'First Division with Distinction', 'Pass', 'PUP/PROV/2024/001', '2024-07-15', 'Issued'),
        ('PUP2021002', '2021BCA002', 'Priya Singh', 'Shri Vijay Singh', 'Smt. Radha Singh', '2003-03-20', 'Bachelor of Computer Applications', 'BCA', 'ABC College of Arts and Science', '2021-2024', '2023-2024', 2024, 7.80, 80.30, 'First Division', 'Pass', '', '', 'Pending'),
    ])
    
    # Sheet 17: Final/Degree Certificate Data
    add_sheet(wb, '17_DegreeCertificate', [
        ('registration_no', True, 'Text', 'PUP2024001', 'Student registration number'),
        ('roll_no', True, 'Text', '2024BCA001', 'Roll number'),
        ('student_name', True, 'Text', 'Rahul Kumar', 'Full name'),
        ('father_name', True, 'Text', 'Shri Ram Kumar', "Father's name"),
        ('mother_name', True, 'Text', 'Smt. Sita Devi', "Mother's name"),
        ('date_of_birth', True, 'Date', '2002-05-15', 'Date of birth'),
        ('degree_name', True, 'Text', 'Bachelor of Computer Applications', 'Degree name'),
        ('program_name', True, 'Text', 'BCA (Honours)', 'Program name'),
        ('specialization', False, 'Text', 'Computer Science', 'Specialization if any'),
        ('college_name', True, 'Text', 'ABC College of Arts and Science', 'College name'),
        ('batch', True, 'Text', '2021-2024', 'Batch'),
        ('passing_session', True, 'Text', '2023-2024', 'Passing session'),
        ('passing_year', True, 'Integer', '2024', 'Passing year'),
        ('cgpa', True, 'Decimal', '8.50', 'Final CGPA'),
        ('percentage', True, 'Decimal', '86.36', 'Final percentage'),
        ('division', True, 'Text', 'First Division with Distinction', 'Division'),
        ('total_credits', True, 'Integer', '132', 'Total credits earned'),
        ('convocation_date', False, 'Date', '2024-12-15', 'Convocation date'),
        ('certificate_no', False, 'Text', 'PUP/DEG/2024/001', 'Certificate number'),
        ('issue_date', False, 'Date', '2024-12-20', 'Issue date'),
        ('status', True, 'Text', 'Pending', 'Pending/Verified/Approved/Issued'),
    ], [
        ('PUP2021001', '2021BCA001', 'Rahul Kumar', 'Shri Ram Kumar', 'Smt. Sita Devi', '2002-05-15', 'Bachelor of Computer Applications', 'BCA (Honours)', 'Computer Science', 'ABC College of Arts and Science', '2021-2024', '2023-2024', 2024, 8.50, 86.36, 'First Division with Distinction', 132, '2024-12-15', 'PUP/DEG/2024/001', '2024-12-20', 'Issued'),
    ])
    
    # Sheet 18: Migration Certificate Data
    add_sheet(wb, '18_MigrationCertificate', [
        ('registration_no', True, 'Text', 'PUP2024001', 'Student registration number'),
        ('roll_no', True, 'Text', '2024BCA001', 'Roll number'),
        ('student_name', True, 'Text', 'Rahul Kumar', 'Full name'),
        ('father_name', True, 'Text', 'Shri Ram Kumar', "Father's name"),
        ('date_of_birth', True, 'Date', '2002-05-15', 'Date of birth'),
        ('program_name', True, 'Text', 'BCA (Honours)', 'Program name'),
        ('college_name', True, 'Text', 'ABC College of Arts and Science', 'College name'),
        ('passing_year', True, 'Integer', '2024', 'Passing year'),
        ('migrating_to', False, 'Text', 'XYZ University', 'Migrating to university'),
        ('purpose', False, 'Text', 'Higher Studies', 'Purpose of migration'),
        ('certificate_no', False, 'Text', 'PUP/MIG/2024/001', 'Certificate number'),
        ('issue_date', False, 'Date', '2024-08-01', 'Issue date'),
        ('status', True, 'Text', 'Pending', 'Pending/Verified/Approved/Issued'),
    ], [
        ('PUP2021001', '2021BCA001', 'Rahul Kumar', 'Shri Ram Kumar', '2002-05-15', 'BCA (Honours)', 'ABC College of Arts and Science', 2024, 'Delhi University', 'MCA Admission', 'PUP/MIG/2024/001', '2024-08-01', 'Issued'),
    ])
    
    # Save the workbook
    output_path = os.path.join(os.path.dirname(__file__), 'PUP_UMIS_Data_Import_Template.xlsx')
    wb.save(output_path)
    
    print(f"✅ Template created successfully: {output_path}")
    print(f"\n📋 Total Sheets: {len(wb.sheetnames)}")
    print(f"\n📁 Sheets created:")
    for sheet in wb.sheetnames:
        print(f"   - {sheet}")
    
    print(f"\n📊 Categories:")
    print("   🏛️  Master Data: Colleges, Faculties, Departments, Degrees, Programs, Batches, Sessions, ExamCentres")
    print("   📚 Course Data: Courses (with slot, credits, marks info)")
    print("   👨‍🎓 Student Data: Students")
    print("   📝 Marks & Results: CIAMarks, FinalMarks, SemesterResults, OverallResults, TabulationRegister")
    print("   📜 Certificates: ProvisionalCertificate, DegreeCertificate, MigrationCertificate")
    
    return output_path


if __name__ == '__main__':
    generate_template()
