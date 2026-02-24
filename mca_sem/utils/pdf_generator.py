def generate_mca_admit_card_pdf(student, exam):
    from weasyprint import HTML, CSS
    import io, base64, os
    from django.conf import settings
    from django.template.loader import get_template
    from mca_sem.models import MCAExamCenterMapping, MCAExamSchedule
    from pup_umis_backend.utils.file_utils import image_to_base64
    import logging

    logger = logging.getLogger(__name__)

    # Get exam center mapping
    exam_center = None
    if student.college:
        mapping = MCAExamCenterMapping.objects.filter(
            exam=exam,
            attached_colleges=student.college
        ).first()
        if mapping:
            exam_center = mapping.center
    
    # Get exam schedules
    schedules = MCAExamSchedule.objects.filter(
        exam=exam
    ).select_related('common_course_structure')


    # Prepare context for template
    context = {
        "exam": exam,
        "student": student,
        "center_mapping": mapping,
        "center_name": exam_center.name if exam_center else "-",
        "center_code": exam_center.center_code if exam_center else "-",
        "status": student.get_status_display() if student else "-",
        "schedules": schedules,
        "university_logo": image_to_base64(os.path.join(settings.MEDIA_ROOT, "common/purnea-logo.png")),
        "watermark_logo": image_to_base64(os.path.join(settings.MEDIA_ROOT, "common/purnea-logo.png")),
        "student_photo": image_to_base64(student.profile_image.path if student.profile_image else None),
        "student_sig": image_to_base64(student.signature.path if student.signature else None),
        "controller_signature": image_to_base64(os.path.join(settings.MEDIA_ROOT, "common/controller-of-examination-signature.png")),
    }

    # Render HTML template
    html_string = get_template("mca_sem/admit_card.html").render(context)

    try:
        # Generate PDF using WeasyPrint
        pdf_file = HTML(string=html_string, base_url=settings.MEDIA_ROOT).write_pdf()
        
        logger.info(f"PDF generated successfully using WeasyPrint, size: {len(pdf_file)} bytes")
        return pdf_file
        
    except Exception as e:
        logger.error(f"PDF generation failed with WeasyPrint: {str(e)}")
        return None

def generate_mca_roll_sheet_pdf(exam, college):
    from weasyprint import HTML
    from django.conf import settings
    from django.template.loader import get_template
    from mca_sem.models import (
        MCAExamRegistration,
        MCAExamSchedule,
        MCACommonCourseStructure,
        MCAExamCenterMapping
    )
    from pup_umis_backend.utils.file_utils import image_to_base64
    import os
    import logging

    logger = logging.getLogger(__name__)

    # 1. Get all students registered for this exam in this college
    registrations = MCAExamRegistration.objects.filter(
        exam=exam,
        student__college=college
    ).select_related('student', 'student__batch').order_by('student__roll_no', 'student__registration_no')

    if not registrations.exists():
        logger.warning(f"No registrations found for Exam: {exam.name}, College: {college.name}")
        return None

    # Get exam center mapping for this college
    center_name = "-"
    center_mapping = MCAExamCenterMapping.objects.filter(
        exam=exam,
        attached_colleges=college
    ).first()
    if center_mapping and center_mapping.center:
        center_name = center_mapping.center.name

    # 2. Get the subjects (schedules) for this exam
    schedules = MCAExamSchedule.objects.filter(
        exam=exam
    ).select_related('common_course_structure').order_by('exam_date', 'exam_time')

    subjects = []
    all_subject_ids = []
    for s in schedules:
        if s.common_course_structure:
            subjects.append({
                'id': s.common_course_structure.id,
                'code': s.common_course_structure.code,
                'course_name': s.common_course_structure.course_name
            })
            all_subject_ids.append(s.common_course_structure.id)

    # 3. Prepare student data rows
    student_data = []
    for reg in registrations:
        student = reg.student
        
        # Determine subjects student is registered for
        if reg.exam_type == 'REGULAR':
            student_subject_ids = all_subject_ids
        else:
            # For BACKLOG/IMPROVEMENT, get specific subjects
            student_subject_ids = list(reg.subjects.all().values_list('id', flat=True))
        
        student_data.append({
            'name': student.get_full_name(),
            'roll_no': student.roll_no or "-",
            'registration_no': student.registration_no or "-",
            'subject_ids': student_subject_ids
        })

    # 4. Prepare Context
    context = {
        "exam": exam,
        "college": college,
        "course_name": "MCA",
        "batch_name": registrations[0].student.batch.name if registrations[0].student.batch else "-",
        "session_name": exam.session or "-",
        "college_name": college.name or "-",
        "center_name": center_name,
        "semester": f"{exam.semester}" if exam.semester else "-",
        "syllabus": "-",
        "subjects": subjects,
        "student_data": student_data,
        "controller_signature": image_to_base64(os.path.join(settings.MEDIA_ROOT, "common/controller-of-examination-signature.png")),
    }

    # Render HTML template
    html_string = get_template("mca_sem/roll_sheet.html").render(context)

    try:
        # Generate PDF using WeasyPrint
        pdf_file = HTML(string=html_string, base_url=settings.MEDIA_ROOT).write_pdf()
        return pdf_file
    except Exception as e:
        logger.error(f"Roll Sheet PDF generation failed: {str(e)}")
        return None


def generate_mca_attendance_sheet_pdf(exam, college):
    from weasyprint import HTML
    from django.conf import settings
    from django.template.loader import get_template
    from mca_sem.models import (
        MCAExamRegistration,
        MCAExamSchedule,
        MCAExamCenterMapping
    )
    from pup_umis_backend.utils.file_utils import image_to_base64, generate_barcode_base64
    import os
    import logging

    logger = logging.getLogger(__name__)

    # 1. Get registrations
    registrations = MCAExamRegistration.objects.filter(
        exam=exam,
        student__college=college
    ).select_related('student', 'student__batch', 'student__college').order_by('student__roll_no', 'student__registration_no')

    if not registrations.exists():
        return None

    # 2. Get center info
    center_mapping = MCAExamCenterMapping.objects.filter(
        exam=exam,
        attached_colleges=college
    ).first()
    center_name = f"{center_mapping.center.college_code} - {center_mapping.center.name}" if center_mapping and center_mapping.center else "-"

    # 3. Get all global schedules for this exam (for regular students)
    global_schedules = MCAExamSchedule.objects.filter(
        exam=exam
    ).select_related('common_course_structure').order_by('exam_date', 'exam_time')

    # 4. Prepare student-wise data
    attendance_data = []
    semester_label = f"SEM {exam.semester}" if exam.semester else "-"

    for reg in registrations:
        student = reg.student
        
        # Determine student's schedules
        student_schedules = []
        if reg.exam_type == 'REGULAR':
            target_schedules = global_schedules
        else:
            # For BACKLOG, filter schedules matching their registered backlog subjects
            backlog_ids = list(reg.subjects.all().values_list('id', flat=True))
            target_schedules = global_schedules.filter(common_course_structure_id__in=backlog_ids)

        for s in target_schedules:
            student_schedules.append({
                'date': s.exam_date.strftime('%d-%m-%Y') if s.exam_date else "-",
                'sitting': s.sitting or "-",
                'subject_name': s.common_course_structure.course_name if s.common_course_structure else "-",
                'subject_code': s.common_course_structure.code if s.common_course_structure else "-",
            })

        # Generate Barcode Data
        # Format: Roll: XXX, Reg: YYY, Name: ZZZ, Sem: AAA
        barcode_text = f"Roll:{student.roll_no or ''}, Reg:{student.registration_no or ''}, Name:{student.get_full_name()}, Sem:{semester_label}"
        barcode_base64 = generate_barcode_base64(barcode_text)

        attendance_data.append({
            'name': student.get_full_name(),
            'registration_no': student.registration_no or "-",
            'roll_no': student.roll_no or "-",
            'college_name': student.college.name if student.college else "-",
            'photo': image_to_base64(student.profile_image.path if student.profile_image else None),
            'schedules': student_schedules,
            'barcode': barcode_base64
        })

    # Header like: MCA-1ST-REGULAR-MCA
    exam_header = f"MCA-{exam.semester}"

    context = {
        "attendance_data": attendance_data,
        "center_name": center_name,
        "exam_header": exam_header,
        "university_logo": image_to_base64(os.path.join(settings.MEDIA_ROOT, "common/purnea-logo.png")),
    }

    html_string = get_template("mca_sem/attendance_sheet.html").render(context)

    try:
        return HTML(string=html_string, base_url=settings.MEDIA_ROOT).write_pdf()
    except Exception as e:
        logger.error(f"Attendance Sheet PDF generation failed: {str(e)}")
        return None

def generate_mca_tr_pdf(students, college, exam, batch_uid=None):
    import os
    import uuid
    import shutil
    import subprocess
    import tempfile
    from collections import defaultdict
    from django.conf import settings
    from openpyxl import load_workbook
    from mca_sem.models import MCAStudentAssessment, MCABatch
    import logging

    logger = logging.getLogger(__name__)

    semester = str(exam.semester)

    # Constants based on the TR structure provided in images
    COL_START = 4  # Starting column for marks (ESE is usually around here)
    SUBJECT_WIDTH = 3 # ESE, CIA, Total (as per the physical TR image)
    DATA_START_ROW = 13  # Row where student data starts (after header and subject columns)
    STUDENTS_PER_PAGE = 5

    def normalize(code):
        return (code or "").replace("-", "").replace(" ", "").upper().strip()

    # ===== TEMPLATE LOAD =====
    template_filename = f"MCA_{semester}_TR_FORMATE.xlsx"
    template_path = os.path.join(settings.BASE_DIR, "mca_sem", "static", "TR", template_filename)
    
    if not os.path.exists(template_path):
        template_path = os.path.join(settings.BASE_DIR, "mca_sem", "static", "TR", "MCA_1_TR_FORMATE.xlsx")

    if not os.path.exists(template_path):
        logger.error(f"Template not found: {template_path}")
        return None

    temp_dir = tempfile.gettempdir()
    temp_excel = os.path.join(temp_dir, f"mca_result_{uuid.uuid4().hex}.xlsx")
    shutil.copy(template_path, temp_excel)

    wb = load_workbook(temp_excel)
    master_sheet = wb.active
    master_sheet.title = "MASTER_TEMPLATE"

    students = list(students)
    if not students:
        return None

    # ===== FETCH DATA =====
    # Filter by exam (which already encodes the semester) rather than the
    # non-existent 'semester' field.
    qs = MCAStudentAssessment.objects.filter(
        student__in=students,
        exam=exam
    ).select_related("student", "course_structure")

    if batch_uid:
        qs = qs.filter(batch__uid=batch_uid)

    all_assessments = list(qs)
    student_map = defaultdict(list)
    subject_codes_found = set()

    logger.info(f"Total assessments found: {len(all_assessments)}")

    for obj in all_assessments:
        # course_code lives on the related course_structure, not on the assessment itself
        cs = obj.course_structure
        code = normalize(cs.course_code if cs else None)
        
        # Debug: Log each assessment record
        logger.info(f"Assessment: Student={obj.student.roll_no}, Label={obj.label}, CourseCode={cs.course_code if cs else 'None'}, Normalized={code}, Marks={obj.ind_marks_obtained}")
        
        if not code:
            logger.warning(f"Skipping assessment with no course_code: {obj}")
            continue
            
        subject_codes_found.add(code)
        student_map[obj.student.id].append(obj)

    # We sort codes to ensure consistent column mapping
    SUBJECT_CODES = sorted(list(subject_codes_found))
    
    logger.info(f"Subject codes found: {SUBJECT_CODES}")
    logger.info(f"Students with assessments: {len(student_map)}")

    # ===== PAGINATION =====
    def chunked(lst, size):
        for i in range(0, len(lst), size):
            yield lst[i:i + size]

    pages = list(chunked(students, STUDENTS_PER_PAGE))

    # ===== PAGE LOOP =====
    for page_index, student_chunk in enumerate(pages):
        ws = wb.copy_worksheet(master_sheet)
        ws.title = f"Page_{page_index + 1}"

        # ===== FILL HEADER (EXAM INFO) =====
        # Adjust cell coordinates (B2, B3, etc.) based on your template
        ws['A2'] = f"{exam.name}"
        ws['A3'] = f"Examination {exam.exam_month_year or ''} held in the month of {exam.exam_month_year or ''}"
        ws['A4'] = f"College: {college.college_code} - {college.name}"

        # Footer statistics for this page
        p_total = 0
        p_pass = 0
        p_fail = 0
        p_distinction = 0
        p_first_class = 0
        p_second_class = 0

        # ===== STUDENT LOOP =====
        row_pointer = DATA_START_ROW
        for student in student_chunk:
            # Basic Info
            ws.cell(row=row_pointer, column=1).value = student.roll_no
            ws.cell(row=row_pointer, column=2).value = student.get_full_name()
            ws.cell(row=row_pointer, column=3).value = student.registration_no

            logger.info(f"\n=== Processing Student: {student.roll_no} - {student.get_full_name()} ===")
            logger.info(f"Student has {len(student_map.get(student.id, []))} assessment records")

            grand_total = 0
            max_grand_total = 0
            failed_in_subject = False
            
            col_pointer = COL_START
            
            # Process each subject dynamically based on codes found
            for code in SUBJECT_CODES:
                logger.info(f"  Processing subject code: {code}")
                ese_marks = 0
                cia_marks = 0
                practical_marks = 0
                subject_max = 100 # Default
                
                ese_pass = 28 # 40% of 70
                cia_pass = 12 # 40% of 30
                prac_pass = 40 # 40% of 100
                
                is_practical_paper = False
                
                for rec in student_map.get(student.id, []):
                    rec_cs = rec.course_structure
                    rec_code_normalized = normalize(rec_cs.course_code if rec_cs else None)
                    
                    logger.info(f"    Checking record: Label={rec.label}, CourseCode={rec_cs.course_code if rec_cs else 'None'}, Normalized={rec_code_normalized}")
                    
                    if rec_code_normalized == code:
                        logger.info(f"    ✓ MATCH FOUND! Label={rec.label}, Marks={rec.ind_marks_obtained}")
                        
                        subject_max = float(rec.course_max_marks or 100)
                        rec_label = (rec.label or "").strip()
                        cs_label = (rec_cs.label or "") if rec_cs else ""
                        cs_name  = (rec_cs.course_name or "") if rec_cs else ""
                        if rec_label == "ESE":
                            ese_marks = float(rec.ind_marks_obtained or 0)
                            logger.info(f"      ESE Marks: {ese_marks}")
                            if ese_marks < ese_pass: failed_in_subject = True
                        elif rec_label == "CIA":
                            cia_marks = float(rec.ind_marks_obtained or 0)
                            logger.info(f"      CIA Marks: {cia_marks}")
                            if cia_marks < cia_pass: failed_in_subject = True
                        elif rec_label == "Practical" or "PRACTICAL" in cs_name.upper() or "PRACTICAL" in cs_label.upper():
                            practical_marks = float(rec.ind_marks_obtained or 0)
                            logger.info(f"      Practical Marks: {practical_marks}")
                            is_practical_paper = True
                            if practical_marks < prac_pass: failed_in_subject = True

                subj_total = ese_marks + cia_marks + practical_marks
                logger.info(f"  Subject {code} Total: ESE={ese_marks}, CIA={cia_marks}, Practical={practical_marks}, Total={subj_total}")
                
                grand_total += subj_total
                max_grand_total += subject_max

                # Fill cells based on the 3-column per subject structure (ESE | CIA | Total)
                if not is_practical_paper:
                    ws.cell(row=row_pointer, column=col_pointer).value = ese_marks
                    ws.cell(row=row_pointer, column=col_pointer + 1).value = cia_marks
                    ws.cell(row=row_pointer, column=col_pointer + 2).value = subj_total
                else:
                    # For practical papers, it usually occupies one block
                    ws.cell(row=row_pointer, column=col_pointer).value = practical_marks
                
                col_pointer += SUBJECT_WIDTH

            # ===== FINAL CALCULATION =====
            percentage = (grand_total / max_grand_total * 100) if max_grand_total > 0 else 0
            
            # Pass if all subjects passed AND aggregate >= 45%
            is_passed = (not failed_in_subject) and (percentage >= 45.0)
            
            final_result = "FAIL"
            class_str = "-"
            
            if is_passed:
                final_result = "PASS"
                p_pass += 1
                if percentage >= 75.0:
                    class_str = "Distinction"
                    p_distinction += 1
                elif percentage >= 60.0:
                    class_str = "1st Class"
                    p_first_class += 1
                else:
                    class_str = "2nd Class"
                    p_second_class += 1
            else:
                p_fail += 1

            # Mapping summary to specific columns based on your excel format
            # Using typical columns for Aggregate and Result (change indices if needed)
            ws.cell(row=row_pointer, column=27).value = grand_total
            ws.cell(row=row_pointer, column=28).value = final_result
            ws.cell(row=row_pointer, column=29).value = class_str
            
            p_total += 1
            row_pointer += 1

        # ===== PAGE FOOTER (STATISTICS) =====
        # As per printed TR bottom left:
        ws.cell(row=row_pointer + 2, column=1).value = f"No of Candidate : {p_total}"
        ws.cell(row=row_pointer + 3, column=1).value = f"1st Class with Distinction : {p_distinction}"
        ws.cell(row=row_pointer + 4, column=1).value = f"1st Class : {p_first_class}"
        ws.cell(row=row_pointer + 5, column=1).value = f"Pass : {p_pass}"
        
        ws.cell(row=row_pointer + 2, column=5).value = f"2nd Class : {p_second_class}"
        ws.cell(row=row_pointer + 4, column=5).value = f"Fail : {p_fail}"

    wb.remove(master_sheet)
    wb.save(temp_excel)

    # Convert to PDF using LibreOffice
    def get_soffice_path():
        """Locate the LibreOffice soffice executable on Windows and Unix."""
        import platform
        if platform.system() == "Windows":
            candidates = [
                r"C:\Program Files\LibreOffice\program\soffice.exe",
                r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
            ]
            for path in candidates:
                if os.path.exists(path):
                    return path
        # Fallback: rely on system PATH (works on Linux/Mac)
        return "soffice"

    soffice_path = get_soffice_path()
    try:
        result = subprocess.run(
            [
                soffice_path, "--headless",
                "--convert-to", "pdf:calc_pdf_Export",
                "--outdir", temp_dir,
                temp_excel
            ],
            check=True,
            capture_output=True,
            text=True,
            timeout=120
        )
        logger.info(f"LibreOffice conversion output: {result.stdout}")
    except FileNotFoundError:
        logger.error(f"LibreOffice not found at '{soffice_path}'. Please install LibreOffice and ensure it is accessible.")
        return None
    except subprocess.TimeoutExpired:
        logger.error("LibreOffice conversion timed out after 120 seconds.")
        return None
    except subprocess.CalledProcessError as e:
        logger.error(f"LibreOffice conversion failed (exit {e.returncode}): {e.stderr}")
        return None
    except Exception as e:
        logger.error(f"LibreOffice conversion failed: {e}")
        return None

    pdf_path = temp_excel.replace(".xlsx", ".pdf")
    if not os.path.exists(pdf_path):
        return None

    with open(pdf_path, "rb") as f:
        pdf_bytes = f.read()

    # Cleanup
    if os.path.exists(temp_excel): os.remove(temp_excel)
    if os.path.exists(pdf_path): os.remove(pdf_path)

    return pdf_bytes
