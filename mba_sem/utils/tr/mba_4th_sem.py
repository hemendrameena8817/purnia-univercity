"""
MBA 4th Semester – Data-only result filler
==========================================
The Excel template already contains all static headers, subject names,
full-marks & pass-marks rows.  This module ONLY fills student data.

Subject codes are read from fixed cells in row 6 of the template:
    M6  → col 13  → subj 1  (ESE, CIA, Total, Credit Alloted, Numeric Grade, Credit Earned, GP)
    T6  → col 20  → subj 2  (ESE, CIA, Total, Credit Alloted, Numeric Grade, Credit Earned, GP)
    AA6 → col 27  → subj 3  (ESE, Total, Credit Alloted, Numeric Grade, Credit Earned, GP)  ← MB-403, no CIA
    AG6 → col 33  → subj 4  (ESE, CIA, Total, Credit Alloted, Numeric Grade, Credit Earned, GP)
    AO6 → col 41  → subj 5  (ESE, CIA, Total, Credit Alloted, Numeric Grade, Credit Earned, GP)
    AV6 → col 48  → subj 6  (ESE, CIA, Total, Credit Alloted, Numeric Grade, Credit Earned, GP)

Summary (fixed columns, already labelled in template):
    Col 54  BB  Grand Total Marks Obtained
    Col 55  BC  Total Credit Allotted
    Col 56  BD  Total Credit Earned
    Col 57  BE  GPA (this semester)
    Col 58  BF  CGPA (cumulative 4 semesters)
    Col 59  BG  GPA (for letter-grade reference)
    Col 60  BH  Letter Grade
    Col 61  BI  Description of Grade
    Col 62  BJ  Result Status

Footer (same row logic as original builder):
    DATA_START_ROW(17) + STUDENTS_PER_PAGE(5) - 1 + 4  = row 25
"""

import os
import math
import shutil
import uuid
from collections import defaultdict

from django.conf import settings
from openpyxl import load_workbook

from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font

from mba_sem.models import MBAStudentCourseAssessment
from mba_sem.utils.tr.grading import (
    calculate_numeric_grade,
    calculate_grade_point,
    get_letter_and_description,
)
from mba_sem.utils.tr.pdf_converter import convert_excel_to_pdf


# ─────────────────────────────────────────────────────────────────
# TEMPLATE LAYOUT  (matches MBA_Result_final_1.xlsx row 6 cells)
# ─────────────────────────────────────────────────────────────────

# Each block: which row-6 cell holds the code, which column the block starts,
# and whether the subject has a CIA component.
# Pass-marks are only needed for credit logic (not written to template).
SUBJECT_BLOCKS = [
    {
        "code_cell": "M6",   # Corporate Governance & Business Ethics
        "col":       13,
        "has_cia":   True,
        "ese_pass":  31.5,
        "cia_pass":  13.5,
        "credit":    4,
    },
    {
        "code_cell": "T6",   # Computer Application & MIS
        "col":       20,
        "has_cia":   True,
        "ese_pass":  31.5,
        "cia_pass":  13.5,
        "credit":    4,
    },
    {
        "code_cell": "AA6",  # Comprehensive Viva-Voce  ← ESE only, no CIA
        "col":       27,
        "has_cia":   False,
        "ese_pass":  45.0,
        "cia_pass":  0.0,
        "credit":    4,
    },
    {
        "code_cell": "AG6",  # Business Communication
        "col":       33,
        "has_cia":   True,
        "ese_pass":  31.5,
        "cia_pass":  13.5,
        "credit":    4,
    },
    {
        "code_cell": "AO6",  # Management of Change
        "col":       41,
        "has_cia":   True,
        "ese_pass":  31.5,
        "cia_pass":  13.5,
        "credit":    4,
    },
    {
        "code_cell": "AV6",  # Group Dynamics
        "col":       48,
        "has_cia":   True,
        "ese_pass":  31.5,
        "cia_pass":  13.5,
        "credit":    4,
    },
]

# Column offsets inside a block for ESE+CIA subject:
#  ESE=0, CIA=1, Total=2, Credit Alloted=3, Numeric Grade=4, Credit Earned=5, GP=6
# Column offsets inside a block for ESE-only subject (MB-403):
#  ESE=0, Total=1, Credit Alloted=2, Numeric Grade=3, Credit Earned=4, GP=5

SUMMARY_COLS = {
    "grand_total":    55,   # BC
    "credit_alloted": 56,   # BD
    "credit_earned":  57,   # BE
    "gpa":            58,   # BF
    "cgpa":           59,   # BG
    "gpa2":           60,   # BH
    "letter":         61,   # BI
    "desc":           62,   # BJ
    "status":         63,   # BK
}

DATA_START_ROW    = 17
STUDENTS_PER_PAGE = 5


# ─────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────

def _get_marks(student_id, code, student_map):
    """Return ESE / CIA marks dict for a student + paper code."""
    res = {
        "ese": 0.0, "cia": 0.0,
        "ese_found": False, "cia_found": False,
        "ese_absent": False, "cia_absent": False,
    }
    for rec in student_map.get(student_id, []):
        if rec.paper_code != code:
            continue
        label     = str(rec.label or "").upper()
        is_absent = bool(rec.ind_is_absent)
        marks     = float(
            rec.ind_final_marks_obtained
            if rec.ind_final_marks_obtained is not None
            else (rec.ind_marks_obtained or 0)
        )
        if label.startswith("ESE"):
            res["ese"]        = marks
            res["ese_found"]  = True
            res["ese_absent"] = is_absent
        elif label.startswith("CIA"):
            res["cia"]        = marks
            res["cia_found"]  = True
            res["cia_absent"] = is_absent
    return res


def _calc_subject(marks, block):
    """
    Returns (total, numeric, credit_earned, failed, appeared).
    Identical logic to MBAResultExcelBuilder._determine_subject_result().
    """
    ese_passed = True
    cia_passed = True

    if marks["ese_found"] and not marks["ese_absent"]:
        ese_passed = marks["ese"] >= block["ese_pass"]
    else:
        ese_passed = block["ese_pass"] <= 0

    if block["has_cia"]:
        if marks["cia_found"] and not marks["cia_absent"]:
            cia_passed = marks["cia"] >= block["cia_pass"]
        else:
            cia_passed = block["cia_pass"] <= 0

    total       = marks["ese"] + marks["cia"]
    both_absent = marks["ese_absent"] and (marks["cia_absent"] or not block["has_cia"])
    numeric     = 0 if both_absent else calculate_numeric_grade(total)

    appeared = (marks["ese_found"] and not marks["ese_absent"]) or \
               (block["has_cia"] and marks["cia_found"] and not marks["cia_absent"])

    if ese_passed and cia_passed and numeric >= 5 and appeared:
        credit_earned = float(block["credit"])
        failed        = False
    else:
        credit_earned = 0.0
        failed        = True

    return total, numeric, credit_earned, failed, appeared


def _calc_cgpa(student, current_gp_sum, current_credits, result_status):
    """
    Cumulative GPA calculate using standard formula:
    CGPA = Σ(SemesterGPA × SemesterCredits) / Σ(SemesterCredits)
    
    Numerator   = (GPA1*C1) + (GPA2*C2) + (GPA3*C3) + Current_GP_Sum
    Denominator = C1 + C2 + C3 + Current_Credits
    """
    if result_status in ("Fail", "Absent"):
        return 0.00

    def sf(v):
        try:
            return float(str(v or "0").strip() or "0")
        except Exception:
            return 0.0

    # Numerator contribution from previous semesters
    s1_points = sf(student.sem_1_gpa) * sf(student.sem_1_credit_earned)
    s2_points = sf(student.sem_2_gpa) * sf(student.sem_2_credit_earned)
    s3_points = sf(student.sem_3_gpa) * sf(student.sem_3_credit_earned)
    
    total_points = s1_points + s2_points + s3_points + current_gp_sum
    total_credits = (
        sf(student.sem_1_credit_earned) + 
        sf(student.sem_2_credit_earned) + 
        sf(student.sem_3_credit_earned) + 
        current_credits
    )

    if total_credits == 0:
        return 0.00

    cgpa_raw = total_points / total_credits
    
    # Standard rounding: round off greater or equal to 5 (at 3rd decimal)
    # math.floor(x * 100 + 0.5) / 100 ensures 0.005 -> 0.01 calculation logic
    return math.floor(cgpa_raw * 100 + 0.5) / 100


# ─────────────────────────────────────────────────────────────────
# DATA FILLER  (one worksheet)
# ─────────────────────────────────────────────────────────────────

def fill_students(ws, students, student_map):
    """
    Fill student marks into `ws` (a copy of the master template).
    Reads subject codes from the fixed row-6 cells.
    Writes ONLY data – no structural changes to the sheet.

    Column layout (left side):
        A (1)  : Roll No
        B (2)  : Name of Candidate
        C (3)  : Registration No
        D-F(4-6)  : SEM I  – Roll No | GPA | Credit Earned
        G-I(7-9)  : SEM II – Roll No | GPA | Credit Earned
        J-L(10-12): SEM III– Roll No | GPA | Credit Earned
        M(13)+ : Subject marks blocks
    """

    # 1. Read subject codes from the template row-6 cells
    blocks = []
    for blk in SUBJECT_BLOCKS:
        code = ws[blk["code_cell"]].value
        if not code:
            print(f"[MBA4] WARNING: no code found at {blk['code_cell']}, skipping block")
            continue
        code = str(code).strip()
        blocks.append({**blk, "code": code})
        print(f"[MBA4] Block {blk['code_cell']} → code={code}, col={blk['col']}, has_cia={blk['has_cia']}")

    # 2. Fill one row per student
    row = DATA_START_ROW
    for student in students:

        # ── A: Roll No | B: Name | C: Registration No ──────────────
        ws.cell(row=row, column=1).value = student.roll_no
        ws.cell(row=row, column=2).value = student.get_full_name()
        ws.cell(row=row, column=3).value = student.registration_no

        # ── SEM I  (D=4, E=5, F=6) ─────────────────────────────────
        ws.cell(row=row, column=4).value = student.roll_no
        ws.cell(row=row, column=5).value = student.sem_1_gpa
        ws.cell(row=row, column=6).value = student.sem_1_credit_earned

        # ── SEM II (G=7, H=8, I=9) ─────────────────────────────────
        ws.cell(row=row, column=7).value  = student.roll_no
        ws.cell(row=row, column=8).value  = student.sem_2_gpa
        ws.cell(row=row, column=9).value  = student.sem_2_credit_earned

        # ── SEM III (J=10, K=11, L=12) ─────────────────────────────
        ws.cell(row=row, column=10).value = student.roll_no
        ws.cell(row=row, column=11).value = student.sem_3_gpa
        ws.cell(row=row, column=12).value = student.sem_3_credit_earned

        # ── Subject marks ───────────────────────────────────────────
        total_credit_allotted = 0.0
        total_credit_earned   = 0.0
        total_grade_points    = 0.0
        total_marks_obtained  = 0.0
        any_failed            = False
        student_appeared      = False

        for blk in blocks:
            code  = blk["code"]
            col   = blk["col"]
            marks = _get_marks(student.id, code, student_map)

            total, numeric, credit_earned, subj_failed, appeared = _calc_subject(marks, blk)

            if appeared:    student_appeared = True
            if subj_failed: any_failed       = True

            # GP displayed in the cell = Credit Allotted × Numeric Grade (always)
            gp_display = round(float(numeric) * float(blk["credit"]), 2)

            # GP used for GPA sum = Credit Earned × Numeric Grade (0 when failed)
            gp_for_gpa = calculate_grade_point(numeric, credit_earned)

            total_credit_allotted += float(blk["credit"])
            total_credit_earned   += credit_earned
            total_grade_points    += gp_for_gpa          # GPA only counts earned credits
            total_marks_obtained  += total

            both_absent = marks["ese_absent"] and (
                marks["cia_absent"] if blk["has_cia"] else True
            )

            if blk["has_cia"]:
                # ESE, CIA, Total, Credit Alloted, Numeric Grade, Credit Earned, GP
                ws.cell(row=row, column=col + 0).value = "AB" if marks["ese_absent"] else marks["ese"]
                ws.cell(row=row, column=col + 1).value = "AB" if marks["cia_absent"] else marks["cia"]
                ws.cell(row=row, column=col + 2).value = "AB" if both_absent else total
                ws.cell(row=row, column=col + 3).value = float(blk["credit"])
                ws.cell(row=row, column=col + 4).value = numeric if appeared else 0
                ws.cell(row=row, column=col + 5).value = credit_earned
                ws.cell(row=row, column=col + 6).value = gp_display   # Credit Allotted × Numeric
            else:
                # ESE only (MB-403 Viva): ESE, Total, Credit Alloted, Numeric Grade, Credit Earned, GP
                ws.cell(row=row, column=col + 0).value = "AB" if marks["ese_absent"] else marks["ese"]
                ws.cell(row=row, column=col + 1).value = "AB" if marks["ese_absent"] else total
                ws.cell(row=row, column=col + 2).value = float(blk["credit"])
                ws.cell(row=row, column=col + 3).value = numeric if appeared else 0
                ws.cell(row=row, column=col + 4).value = credit_earned
                ws.cell(row=row, column=col + 5).value = gp_display   # Credit Allotted × Numeric

            print(
                f"[MBA4] {student.roll_no} | {code} | "
                f"ESE={marks['ese']}(AB={marks['ese_absent']}) "
                f"CIA={marks['cia']}(AB={marks['cia_absent']}) "
                f"Total={total} NG={numeric} CE={credit_earned} "
                f"GP(display)={gp_display} GP(gpa)={gp_for_gpa}"
            )

        # ── Result status ───────────────────────────────────────────
        if not student_appeared:
            status = "Absent"
            gpa    = 0.00
        elif any_failed:
            status = "Fail"
            gpa    = 0.00
        else:
            status = "Pass"
            gpa    = (
                round(total_grade_points / total_credit_allotted, 2)
                if total_credit_allotted > 0 else 0.00
            )

        if status == "Pass":
            letter, desc = get_letter_and_description(gpa)
        elif status == "Fail":
            letter, desc = "F",  "Fail"
        else:
            letter, desc = "AB", "Absent"

        cgpa = _calc_cgpa(student, total_grade_points, total_credit_allotted, status)

        # ── Summary block ───────────────────────────────────────────
        ws.cell(row=row, column=SUMMARY_COLS["grand_total"]   ).value = total_marks_obtained
        ws.cell(row=row, column=SUMMARY_COLS["credit_alloted"]).value = total_credit_allotted
        ws.cell(row=row, column=SUMMARY_COLS["credit_earned"] ).value = total_credit_earned
        ws.cell(row=row, column=SUMMARY_COLS["gpa"]           ).value = gpa
        ws.cell(row=row, column=SUMMARY_COLS["cgpa"]          ).value = cgpa
        ws.cell(row=row, column=SUMMARY_COLS["gpa2"]          ).value = gpa
        ws.cell(row=row, column=SUMMARY_COLS["letter"]        ).value = letter
        ws.cell(row=row, column=SUMMARY_COLS["desc"]          ).value = desc
        ws.cell(row=row, column=SUMMARY_COLS["status"]        ).value = status

        print(
            f"[MBA4] {student.roll_no} | {student.get_full_name()} | "
            f"GPA={gpa} CGPA={cgpa} Status={status} Letter={letter}"
        )

        row += 1



def _add_logo(ws):
    """Add the university logo to the top-left of the worksheet."""
    logo_path = os.path.join(settings.BASE_DIR, "static/images/purnea-logo.png")
    if not os.path.exists(logo_path):
        print(f"[MBA4] Logo not found at {logo_path}")
        return

    img = XLImage(logo_path)
    img.width = 110
    img.height = 110
    img.anchor = "W1"
    ws.add_image(img)


def fill_header(ws, exam_name, college, students):
    """Fill the top header with exam name, college name, and subject name."""
    ws["A4"] = exam_name or ""
    ws["BC4"] = f"Dept./College : {college.name}"

    if students and students[0].course:
        ws["BC2"] = f"Subject : {students[0].course.name}"
        ws["BB2"].font = Font(size=11)



def fill_footer(ws, students):
    """Write pass/fail/absent counts and signature lines."""
    footer_row = DATA_START_ROW + STUDENTS_PER_PAGE - 1 + 4   # row 25

    page_pass = page_fail = page_absent = 0
    status_col = SUMMARY_COLS["status"]

    for i in range(len(students)):
        s = ws.cell(row=DATA_START_ROW + i, column=status_col).value
        if   s == "Pass":   page_pass   += 1
        elif s == "Fail":   page_fail   += 1
        elif s == "Absent": page_absent += 1

    # Left block
    ws.cell(row=footer_row,     column=2).value = f"No of Students : {len(students)}"
    ws.cell(row=footer_row + 1, column=2).value = f"Pass : {page_pass}"
    ws.cell(row=footer_row + 2, column=2).value = f"Promoted : 0"

    # Centre block
    ws.cell(row=footer_row,     column=8).value = "Expelled : 0"
    ws.cell(row=footer_row + 1, column=8).value = f"Fail : {page_fail}"
    ws.cell(row=footer_row + 2, column=8).value = "Qualified : 0"

    # Right block
    ws.cell(row=footer_row + 1, column=14).value = f"Absent : {page_absent}"
    ws.cell(row=footer_row + 2, column=14).value = "Result Pending : 0"

    # Signature area
    sign_row = footer_row + 2
    ws.cell(row=sign_row, column=22).value = "Compared By Address"
    ws.cell(row=sign_row, column=38).value = "Full Signature of Tabulator-cum-scrutinizer with date"
    ws.cell(row=sign_row, column=58).value = "Controller of Examination"


# ─────────────────────────────────────────────────────────────────
# GENERATOR  (high-level entry point for MBA 4th semester)
# ─────────────────────────────────────────────────────────────────

class MBA4thSemResultGenerator:
    """
    Usage:
        gen = MBA4thSemResultGenerator(students, college, batch_uid, exam_name)
        pdf_bytes = gen.generate()
    """

    TEMPLATE_PATH = os.path.join(
        settings.BASE_DIR,
        "mba_sem/static/tr/MBA_Result_final_1.xlsx"
    )
    SEMESTER = "4"

    def __init__(self, students, college, batch_uid=None, exam_name=None):
        self.students  = list(students)
        self.college   = college
        self.batch_uid = batch_uid
        self.exam_name = exam_name

    def _build_student_map(self):
        """Single DB query → student_id → list of assessment records."""
        qs = MBAStudentCourseAssessment.objects.filter(
            student__in  = self.students,
            semester     = self.SEMESTER,
            college_code = self.college.college_code,
        )
        if self.batch_uid:
            qs = qs.filter(batch__uid=self.batch_uid)

        student_map = defaultdict(list)
        for obj in qs:
            student_map[obj.student.id].append(obj)

        print(
            f"[MBA4] {qs.count()} assessment records | "
            f"{len(self.students)} students | college={self.college.college_code}"
        )
        return student_map

    def generate(self):
        if not self.students:
            print("[MBA4] No students – aborting.")
            return None

        temp_excel = os.path.join("/tmp", f"mba4_{uuid.uuid4().hex}.xlsx")
        shutil.copy(self.TEMPLATE_PATH, temp_excel)

        wb          = load_workbook(temp_excel)
        master      = wb.active
        master.title = "MASTER_TEMPLATE"

        student_map = self._build_student_map()

        # Split students into pages of 5
        pages = [
            self.students[i : i + STUDENTS_PER_PAGE]
            for i in range(0, len(self.students), STUDENTS_PER_PAGE)
        ]

        for idx, chunk in enumerate(pages):
            ws       = wb.copy_worksheet(master)
            ws.title = f"Page_{idx + 1}"

            fill_students(ws, chunk, student_map)
            fill_header(ws, self.exam_name, self.college, chunk)
            _add_logo(ws)
            fill_footer(ws, chunk)

        # Remove template sheet
        if "MASTER_TEMPLATE" in wb.sheetnames:
            wb.remove(wb["MASTER_TEMPLATE"])

        wb.save(temp_excel)
        print(f"[MBA4] Saved to {temp_excel}")

        return convert_excel_to_pdf(temp_excel)