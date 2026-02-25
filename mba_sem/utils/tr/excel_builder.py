import os
import io
from django.conf import settings
from openpyxl.drawing.image import Image as XLImage
from mba_sem.models import MBACourseStructure
from mba_sem.utils.tr.grading import (
    calculate_numeric_grade,
    calculate_grade_point,
    get_letter_and_description,
)
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter



class MBAResultExcelBuilder:

    COL_START = 13
    DATA_START_ROW = 17
    STUDENTS_PER_PAGE = 5

    def __init__(
        self,
        wb,
        master_sheet,
        students,
        college,
        all_assessments,
        subject_master,
        student_map,
        subject_codes,
        semester,
        exam_name=None
    ):
        self.wb = wb
        self.master_sheet = master_sheet
        self.students = list(students)
        self.college = college
        self.all_assessments = all_assessments
        self.subject_master = subject_master or {}
        self.student_map = student_map or {}
        self.subject_codes = subject_codes or []
        self.semester = str(semester)
        self.exam_name = exam_name
        self.subject_structure = {}

    # ======================================================
    # LOGO
    # ======================================================

    def _add_logo(self, ws):

        logo_path = os.path.join(
            "static/images/",
            "purnea-logo.png"
        )

        if not os.path.exists(logo_path):
            print("Logo not found:", logo_path)
            return

        img = XLImage(logo_path)
        img.width = 110
        img.height = 110
        img.anchor = "W1"

        ws.add_image(img)

    # ======================================================
    # BUILD
    # ========================f==============================

    def build(self):

        pages = [
            self.students[i:i + self.STUDENTS_PER_PAGE]
            for i in range(0, len(self.students), self.STUDENTS_PER_PAGE)
        ]

        for index, student_chunk in enumerate(pages):

            ws = self.wb.copy_worksheet(self.master_sheet)
            ws.title = f"Page_{index+1}"

            self._add_logo(ws)
            self._build_header(ws, student_chunk)
            self._build_subject_header(ws)
            self._build_students(ws, student_chunk)
            self._build_footer(ws, student_chunk)   

        if "MASTER_TEMPLATE" in self.wb.sheetnames:
            self.wb.remove(self.wb["MASTER_TEMPLATE"])

    # ======================================================
    # HEADER
    # ======================================================

    def _build_header(self, ws, student_chunk):
        ws["A4"] = self.exam_name or ""
        ws["BC4"] = f"Dept./College : {self.college.name}"

        if student_chunk and student_chunk[0].course:
            ws["BC2"] = f"Subject : {student_chunk[0].course.name}"
            ws["BB2"].font = Font(size=11)

    # ======================================================
    # SUBJECT HEADER
    # ======================================================

    def _build_subject_header(self, ws):

        col_pointer = self.COL_START

        for code in self.subject_codes:

            structures = MBACourseStructure.objects.filter(
                course_code=code,
                semester=self.semester
            )  

            
            has_ese = False
            has_cia = False
            ese_max = cia_max = 0
            ese_pass = cia_pass = 0
            credit = 0

            for obj in structures:
                label = str(obj.label or "").upper()
                credit = float(obj.credit or 0)

                if label.startswith("ESE"):
                    has_ese = True
                    ese_max += float(obj.max_marks or 0)
                    ese_pass += float(obj.min_marks or 0)

                if label.startswith("CIA"):
                    has_cia = True
                    cia_max += float(obj.max_marks or 0)
                    cia_pass += float(obj.min_marks or 0)

            headers = []
            if has_ese:
                headers.append("ESE")
            if has_cia:
                headers.append("CIA")

            headers += [
                "Total",
                "Credit Alloted",
                "Numeric Grade",
                "Credit Earned",
                "GP"
            ]
            for i, header in enumerate(headers):
                ws.cell(row=7, column=col_pointer + i).value = header

            # 🔥 DEBUG PRINT YAHAN LAGANA HAI
            print("------ HEADER COLUMN DEBUG ------")
            print("SUBJECT:", code)

            for i, header in enumerate(headers):
                print(f"{header} -> Column {col_pointer + i}")

            print("---------------------------------")
            for s in structures:
                print("LABEL:", s.label, "| MAX:", s.max_marks, "| MIN:", s.min_marks)


            self.subject_structure[code] = {
                "headers": headers,
                "credit": credit,
                "has_ese": has_ese,
                "has_cia": has_cia,
                "ese_pass": ese_pass,
                "cia_pass": cia_pass,
            }

            width = len(headers)

            ws.merge_cells(start_row=5, start_column=col_pointer,
                           end_row=5, end_column=col_pointer + width - 1)
            ws.cell(row=5, column=col_pointer).value = \
                self.subject_master.get(code, code)

            ws.merge_cells(start_row=6, start_column=col_pointer,
                           end_row=6, end_column=col_pointer + width - 1)
            ws.cell(row=6, column=col_pointer).value = code

            for i, header in enumerate(headers):
                ws.cell(row=7, column=col_pointer + i).value = header

            col_offset = col_pointer

            if has_ese:
                ws.cell(row=14, column=col_offset).value = ese_max
                ws.cell(row=15, column=col_offset).value = ese_pass
                col_offset += 1

            if has_cia:
                ws.cell(row=14, column=col_offset).value = cia_max
                ws.cell(row=15, column=col_offset).value = cia_pass
                col_offset += 1

            ws.cell(row=14, column=col_offset).value = ese_max + cia_max
            ws.cell(row=15, column=col_offset).value = ese_pass + cia_pass
            ws.cell(row=14, column=col_offset + 1).value = credit

            col_pointer += width
        self.gpa_start_col = col_pointer

    # ======================================================
    # HELPER FUNCTIONS
    # ======================================================

    def _get_student_marks(self, student_id, code):
        """Helper to fetch ESE and CIA marks for a specific student and paper."""
        res = {
            "ese": 0.0, "cia": 0.0, 
            "ese_found": False, "cia_found": False,
            "ese_absent": False, "cia_absent": False
        }
        for rec in self.student_map.get(student_id, []):
            if rec.paper_code == code:
                label = str(rec.label or "").upper()
                is_absent = bool(rec.ind_is_absent)
                marks = float(
                    rec.ind_final_marks_obtained
                    if rec.ind_final_marks_obtained is not None
                    else (rec.ind_marks_obtained or 0)
                )
                if label.startswith("ESE"):
                    res["ese"] = marks
                    res["ese_found"] = True
                    res["ese_absent"] = is_absent
                elif label.startswith("CIA"):
                    res["cia"] = marks
                    res["cia_found"] = True
                    res["cia_absent"] = is_absent
        return res

    def _determine_subject_result(self, marks_data, meta):
        """Helper to determine total, numeric, credit earned, and pass/fail for a subject."""
        ese_passed = True
        cia_passed = True
        
        if meta["has_ese"]:
            ese_passed = (marks_data["ese"] >= meta["ese_pass"]) if (marks_data["ese_found"] and not marks_data["ese_absent"]) else (meta["ese_pass"] <= 0)
        
        if meta["has_cia"]:
            cia_passed = (marks_data["cia"] >= meta["cia_pass"]) if (marks_data["cia_found"] and not marks_data["cia_absent"]) else (meta["cia_pass"] <= 0)

        total = marks_data["ese"] + marks_data["cia"]
        numeric = calculate_numeric_grade(total) if not (marks_data["ese_absent"] and marks_data["cia_absent"]) else 0
        
        appeared = (marks_data["ese_found"] and not marks_data["ese_absent"]) or (marks_data["cia_found"] and not marks_data["cia_absent"])
        
        # Credit logic: Must pass both ESE and CIA, and have numeric grade >= 5
        if ese_passed and cia_passed and numeric >= 5 and appeared:
            credit_earned = float(meta["credit"])
            failed = False
        else:
            credit_earned = 0.0
            failed = True
            
        return total, numeric, credit_earned, failed, appeared

    def _calculate_cgpa(self, student, current_gpa, result_status):
        """Helper to calculate CGPA based on previous semesters and current status."""
        if result_status in ["Fail", "Absent"]:
            return 0.00
            
        def to_float(val):
            try:
                clean_val = str(val or "0").strip()
                return float(clean_val) if clean_val else 0.0
            except:
                return 0.0

        s1 = to_float(student.sem_1_gpa)
        s2 = to_float(student.sem_2_gpa)
        s3 = to_float(student.sem_3_gpa)
        
        return round((s1 + s2 + s3 + current_gpa) / 4, 2)

    def _build_students(self, ws, student_chunk):

        row_pointer = self.DATA_START_ROW

        # SEM blocks start column (E = 5)
        sem1_col = 4
        sem2_col = 7
        sem3_col = 10

        for student in student_chunk:

            ws.cell(row=row_pointer, column=1).value = student.roll_no
            ws.cell(row=row_pointer, column=2).value = student.get_full_name()
            ws.cell(row=row_pointer, column=3).value = student.registration_no

            # Write SEM headers/GPAs
            ws.cell(row=row_pointer, column=sem1_col).value = student.roll_no
            ws.cell(row=row_pointer, column=sem1_col + 1).value = student.sem_1_gpa
            ws.cell(row=row_pointer, column=sem1_col + 2).value = student.sem_1_credit_earned

            ws.cell(row=row_pointer, column=sem2_col).value = student.roll_no
            ws.cell(row=row_pointer, column=sem2_col + 1).value = student.sem_2_gpa
            ws.cell(row=row_pointer, column=sem2_col + 2).value = student.sem_2_credit_earned

            ws.cell(row=row_pointer, column=sem3_col).value = student.roll_no
            ws.cell(row=row_pointer, column=sem3_col + 1).value = student.sem_3_gpa
            ws.cell(row=row_pointer, column=sem3_col + 2).value = student.sem_3_credit_earned

            col_pointer = self.COL_START
            total_credit_allotted = 0
            total_credit_earned = 0
            total_grade_points = 0
            total_marks_obtained = 0
            any_subject_failed = False
            student_appeared_at_all = False

            for code, meta in self.subject_structure.items():

                marks_data = self._get_student_marks(student.id, code)
                total, numeric, credit_earned, subj_failed, appeared = self._determine_subject_result(marks_data, meta)
                
                if appeared: student_appeared_at_all = True
                if subj_failed: any_subject_failed = True
                
                gp = calculate_grade_point(numeric, credit_earned)
                
                total_credit_allotted += float(meta["credit"])
                total_credit_earned += credit_earned
                total_grade_points += gp
                total_marks_obtained += total

                values = {
                    "ESE": "AB" if marks_data["ese_absent"] else marks_data["ese"],
                    "CIA": "AB" if marks_data["cia_absent"] else marks_data["cia"],
                    "Total": "AB" if (marks_data["ese_absent"] and marks_data["cia_absent"]) else total,
                    "Credit Alloted": float(meta["credit"]),
                    "Numeric Grade": numeric if appeared else 0,
                    "Credit Earned": credit_earned,
                    "GP": gp
                }

                total, numeric, credit_earned, subj_failed, appeared = self._determine_subject_result(marks_data, meta)
                print("\n==============================")
                print("STUDENT:", student.roll_no, "-", student.get_full_name())
                print("SUBJECT:", code)

                print(f"CIA  : {marks_data['cia']}  | Absent: {marks_data['cia_absent']}")
                print(f"ESE  : {marks_data['ese']}  | Absent: {marks_data['ese_absent']}")
                print(f"TOTAL: {total}")
                print(f"NUMERIC GRADE: {numeric}")
                print(f"CREDIT EARNED: {credit_earned}")
                print(f"FAILED?: {subj_failed}")
                print("==============================")

                headers = meta["headers"]
                for i, header in enumerate(headers):
                    ws.cell(row=row_pointer, column=col_pointer + i).value = values.get(header, "")

                col_pointer += len(headers)

            # Determine Final Status
            if not student_appeared_at_all:
                res_status = "Absent"
                gpa = 0.00
            elif any_subject_failed:
                res_status = "Fail"
                gpa = 0.00
            else:
                res_status = "Pass"
                gpa = round(total_grade_points / total_credit_allotted, 2) if total_credit_allotted > 0 else 0

            letter, desc = get_letter_and_description(gpa) if res_status == "Pass" else ("F", "Fail" if res_status == "Fail" else "Absent")
            cgpa = self._calculate_cgpa(student, gpa, res_status)

            # Write Summary
            ws.cell(row=row_pointer, column=54).value = total_marks_obtained   # BB
            ws.cell(row=row_pointer, column=55).value = total_credit_allotted # BC
            ws.cell(row=row_pointer, column=56).value = total_credit_earned   # BD
            ws.cell(row=row_pointer, column=57).value = gpa                  # BE
            ws.cell(row=row_pointer, column=58).value = cgpa                 # BF
            ws.cell(row=row_pointer, column=59).value = gpa                  # BG
            ws.cell(row=row_pointer, column=60).value = letter               # BH
            ws.cell(row=row_pointer, column=61).value = desc                 # BI
            ws.cell(row=row_pointer, column=62).value = res_status           # BJ

            row_pointer += 1

    def _build_footer(self, ws, student_chunk):
        MAX_STUDENTS_PER_PAGE = 5

        footer_row = self.DATA_START_ROW + MAX_STUDENTS_PER_PAGE - 1 + 4

        page_total = len(student_chunk)
        page_pass = 0
        page_fail = 0
        page_absent = 0
        page_pending = 0
        page_expelled = 0
        page_promoted = 0

        summary_col = 55
        result_status_col = 62 # BJ

        for i in range(len(student_chunk)):
            row = self.DATA_START_ROW + i
            status = ws.cell(row=row, column=result_status_col).value

            if status == "Pass":
                page_pass += 1
            elif status == "Fail":
                page_fail += 1
            elif status == "Absent":
                page_absent += 1

        # ===== LEFT BLOCK =====
        ws.cell(row=footer_row, column=2).value = f"No of Students : {page_total}"
        ws.cell(row=footer_row + 1, column=2).value = f"Pass : {page_pass}"
        ws.cell(row=footer_row + 2, column=2).value = f"Promoted : {page_promoted}"

        # ===== CENTER BLOCK =====
        ws.cell(row=footer_row, column=8).value = f"Expelled : {page_expelled}"
        ws.cell(row=footer_row + 1, column=8).value = f"Fail : {page_fail}"
        ws.cell(row=footer_row + 2, column=8).value = f"Qualified : 0"

        # ===== RIGHT BLOCK =====
        ws.cell(row=footer_row + 1 , column=14).value = f"Absent : {page_absent}"
        ws.cell(row=footer_row + 2, column=14).value = f"Result Pending : {page_pending}"

        # ===== SIGN AREA =====
        sign_row = footer_row + 2
        normal_font = Font(name="Calibri", size=11, color="000000")

        ws.cell(row=sign_row, column=22).value = "Compared By Address"
        ws.cell(row=sign_row, column=38).value = "Full Signature of Tabulator-cum-scrutinizer with date"
        ws.cell(row=sign_row, column=58).value = "Controller of Examination"
     