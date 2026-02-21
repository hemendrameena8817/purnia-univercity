from openpyxl.drawing.image import Image as XLImage
import math

GRADE_STRUCTURE = [
    {"min": 91, "max": 100, "numeric": 10, "letter": "O",   "desc": "Outstanding"},
    {"min": 81, "max": 90,  "numeric": 9,  "letter": "A++", "desc": "Excellent"},
    {"min": 71, "max": 80,  "numeric": 8,  "letter": "A+",  "desc": "Very Good"},
    {"min": 61, "max": 70,  "numeric": 7,  "letter": "A",   "desc": "Good"},
    {"min": 51, "max": 60,  "numeric": 6,  "letter": "B+",  "desc": "Average"},
    {"min": 45, "max": 50,  "numeric": 5,  "letter": "B",   "desc": "Pass"},
    {"min": 0,  "max": 44,  "numeric": 0,  "letter": "F",   "desc": "Fail"},
]

def generate_mba_admit_card_pdf(student, exam):
    from weasyprint import HTML, CSS
    import io, base64, os
    from django.conf import settings
    from django.template.loader import get_template
    from mba_sem.models import MBAExamCenterMapping, MBAExamSchedule
    from pup_umis_backend.utils.file_utils import image_to_base64
    import logging

    logger = logging.getLogger(__name__)

    # Get exam center mapping
    exam_center = None
    if student.college:
        mapping = MBAExamCenterMapping.objects.filter(
            exam=exam,
            attached_colleges=student.college
        ).first()
        print(f"{mapping = }")
        if mapping:
            exam_center = mapping.center
            print(f"{exam_center = }")
    
    # Get exam schedules
    # schedules = MBAExamSchedule.objects.filter(
    #     exam=exam
    # ).select_related('common_course_structure')

    # -----------------------------
    # Get exam schedules (discipline based + common)
    # -----------------------------
    discipline = ""

    if student.course and student.course.discipline_code:
        discipline = student.course.discipline_code.strip()

    print("FINAL DISCIPLINE =", discipline)

    # Get all schedules of exam
    schedules = MBAExamSchedule.objects.filter(exam=exam)

    filtered_schedules = []

    for schedule in schedules:
        code = schedule.common_course_structure.code or ""
        code = code.upper().strip()

        # COMMON SUBJECT → MB-101, MB-29, MB-90
        if code.startswith("MB-") and "-" not in code[3:]:
            filtered_schedules.append(schedule)

        # DISCIPLINE SUBJECT → MB-FC-120
        elif discipline and code.startswith(f"MB-{discipline}-"):
            filtered_schedules.append(schedule)

    print("TOTAL SUBJECTS =", len(filtered_schedules))

    for s in filtered_schedules:
        print("SUBJECT CODE =", s.common_course_structure.code)

    filtered_schedules.sort(
        key=lambda s: (s.exam_date is None, not s.exam_time, not s.sitting)
    )

    # use this in context
    schedules = filtered_schedules

    # Prepare context for template
    context = {
        'discipline_code': student.course,
        "exam": exam,
        "student": student,
        "center_mapping": mapping,
        "center_name": exam_center.name if exam_center else "-",
        "center_code": exam_center.center_code if exam_center else "-",
        "schedules": schedules,
        "university_logo": image_to_base64(os.path.join(settings.MEDIA_ROOT, "common/purnea-logo.png")),
        "watermark_logo": image_to_base64(os.path.join(settings.MEDIA_ROOT, "common/purnea-logo.png")),
        "student_photo": image_to_base64(student.profile_image.path if student.profile_image else None),
        "student_sig": image_to_base64(student.signature.path if student.signature else None),
        "controller_signature": image_to_base64(os.path.join(settings.MEDIA_ROOT, "common/controller-of-examination-signature.png")),
    }

    # Render HTML template
    html_string = get_template("mba_sem/admit_card.html").render(context)

    try:
        # Generate PDF using WeasyPrint
        pdf_file = HTML(string=html_string, base_url=settings.MEDIA_ROOT).write_pdf()
        
        logger.info(f"PDF generated successfully using WeasyPrint, size: {len(pdf_file)} bytes")
        return pdf_file
        
    except Exception as e:
        logger.error(f"PDF generation failed with WeasyPrint: {str(e)}")
        return None


def generate_mba_roll_sheet_pdf(exam, college, course):
    from weasyprint import HTML
    from django.conf import settings
    from django.template.loader import get_template
    from collections import OrderedDict
    from mba_sem.models import (
        MBAExamRegistration,
        MBAExamSchedule,
        MBAExamCenterMapping,
    )
    from pup_umis_backend.utils.file_utils import image_to_base64
    import os
    import logging

    logger = logging.getLogger(__name__)

    print("\n==============================")
    print("START MBA ROLL SHEET DEBUG")
    print("==============================")
    print("EXAM   :", exam.name)
    print("COLLEGE:", college.name)
    print("COURSE :", course.name)
    print("==============================\n")

    # =====================================================
    # 1️⃣ FETCH EXAM REGISTRATIONS (SOURCE OF TRUTH)
    # =====================================================
    registrations = (
        MBAExamRegistration.objects
        .filter(
            exam=exam,
            student__college=college,
            student__course=course,
        )
        .select_related(
            "student",
            "student__batch",
            "student__course",
        )
        .prefetch_related("exam_subjects")
        .order_by(
            "student__roll_no",
            "student__registration_no"
        )
    )

    print("🔎 REGISTRATIONS COUNT =", registrations.count())

    if not registrations.exists():
        print("❌ NO REGISTRATIONS FOUND")
        return None

    # =====================================================
    # 2️⃣ EXAM CENTER
    # =====================================================
    center_name = "-"

    center_mapping = (
        MBAExamCenterMapping.objects
        .filter(
            exam=exam,
            attached_colleges=college
        )
        .select_related("center")
        .first()
    )

    if center_mapping and center_mapping.center:
        center_name = center_mapping.center.name

    print("🏫 EXAM CENTER =", center_name)

    # =====================================================
    # 3️⃣ SUBJECTS (ONLY FROM EXAM REGISTRATION ✅)
    # =====================================================
    subjects_map = OrderedDict()

    for reg in registrations:
        print(
            f"DEBUG REG {reg.id} | SUBJECT COUNT =",
            reg.exam_subjects.count()
        )

        for subj in reg.exam_subjects.all():
            code = (subj.code or "").strip().upper()
            if not code:
                continue

            subjects_map[code] = {
                "id": subj.id,
                "code": subj.code,
                "course_name": subj.course_name,
            }

    subjects = list(subjects_map.values())

    print("📚 SUBJECTS FOUND =", len(subjects))

    if not subjects:
        print("❌ NO SUBJECTS FOUND FROM REGISTRATION")
        print("👉 FIX REQUIRED: exam_subjects M2M is empty")
        return None

    # =====================================================
    # 4️⃣ OPTIONAL: SCHEDULE MAP (DATE / TIME ONLY)
    # =====================================================
    schedule_map = {
        s.common_course_structure.code: s
        for s in MBAExamSchedule.objects.filter(
            exam=exam,
            common_course_structure__isnull=False
        )
    }

    # =====================================================
    # 5️⃣ STUDENT ROW DATA
    # =====================================================
    student_data = []

    for reg in registrations:
        student = reg.student

        subject_codes = list(
            reg.exam_subjects.values_list("code", flat=True)
        )

        print(
            f"👤 {student.roll_no} | SUBJECTS = {len(subject_codes)}"
        )

        student_data.append({
            "name": student.get_full_name(),
            "roll_no": student.roll_no or "-",
            "registration_no": student.registration_no or "-",
            "subject_codes": subject_codes,
        })

    print("👥 TOTAL STUDENTS =", len(student_data))

    # =====================================================
    # 6️⃣ TEMPLATE CONTEXT
    # =====================================================
    first_student = registrations.first().student

    context = {
        "exam": exam,
        "college": college,
        "course": course,
        "course_name": course.name,
        "discipline_code": course.discipline_code,
        "batch_name": first_student.batch.name if first_student.batch else "-",
        "semester": exam.semester or "-",
        "session": exam.session or "-",
        "exam_month_year": exam.exam_month_year or "-",
        "college_name": college.name,
        "center_name": center_name,
        "subjects": subjects,
        "year": f"{exam.exam_month_year}" if exam.exam_month_year else "-", 
        "student_data": student_data,
        "controller_signature": image_to_base64(
            os.path.join(
                settings.MEDIA_ROOT,
                "common/controller-of-examination-signature.png"
            )
        ),
    }

    print("🧾 CONTEXT READY — RENDERING PDF")

    # =====================================================
    # 7️⃣ RENDER PDF
    # =====================================================
    html_string = get_template(
        "mba_sem/roll_sheet.html"
    ).render(context)

    try:
        pdf_file = HTML(
            string=html_string,
            base_url=settings.MEDIA_ROOT
        ).write_pdf()

        print("✅ PDF GENERATED SUCCESSFULLY\n")
        return pdf_file

    except Exception as e:
        print("❌ PDF GENERATION FAILED:", str(e))
        logger.exception("MBA Roll Sheet PDF failed")
        return None


from mba_sem.models import *

def generate_mba_attendance_sheet_pdf(exam, college):
    """
    Generate student-wise attendance sheets for MBA students.
    """
    from weasyprint import HTML
    from django.conf import settings
    from django.template.loader import get_template
    from pup_umis_backend.utils.file_utils import image_to_base64, generate_barcode_base64
    import os, logging

    logger = logging.getLogger(__name__)

    # 1️⃣ Eligible students
    semester_regs = MBAExamRegistration.objects.filter(
        sem=exam.semester,
        student__college=college
    ).select_related(
        'student', 'student__course'
    ).order_by('student__roll_no')

    if not semester_regs.exists():
        logger.warning("No eligible MBA students found")
        return None

    # 2️⃣ Exam center
    mapping = MBAExamCenterMapping.objects.filter(
        exam=exam,
        attached_colleges=college
    ).select_related('center').first()

    exam_center = mapping.center if mapping else None

    # 3️⃣ University logo
    logo_path = os.path.join(settings.MEDIA_ROOT, "common/purnea-logo.png")
    university_logo = image_to_base64(logo_path) if os.path.exists(logo_path) else None

    # 4️⃣ All exam schedules (once)
    all_schedules = MBAExamSchedule.objects.filter(
        exam=exam
    ).select_related('common_course_structure').order_by(
        'exam_date', 'exam_time'
    )

    # 5️⃣ Build attendance data
    attendance_data = []

    for reg in semester_regs:
        student = reg.student

        # 🔹 Discipline (SHORT)
        discipline = (student.course.discipline_code or "").strip().upper() \
            if student.course else ""

        # 🔹 Filter schedules
        filtered_schedules = [
            s for s in all_schedules
            if s.common_course_structure
            and s.common_course_structure.code
            and (
                (
                    s.common_course_structure.code.upper().startswith("MB-")
                    and "-" not in s.common_course_structure.code[3:]
                ) or (
                    discipline
                    and s.common_course_structure.code.upper().startswith(f"MB-{discipline}-")
                )
            )
        ]

        # 🔹 Barcode
        barcode_text = (
            f"Roll:{student.roll_no or ''}, "
            f"Reg:{student.registration_no or ''}, "
            f"Name:{student.get_full_name()}, "
            f"Sem:{exam.semester}"
        )
        barcode_base64 = generate_barcode_base64(barcode_text)

        # 🔹 Photo
        photo_base64 = None
        if student.profile_image:
            try:
                photo_base64 = image_to_base64(student.profile_image.path)
            except Exception as e:
                logger.error(f"Photo error {student.registration_no}: {e}")

        # 🔹 Student schedules
        student_schedules = [
            {
                'date': s.exam_date.strftime('%d-%m-%Y') if s.exam_date else '-',
                'exam_time': s.exam_time or '',
                'sitting': s.sitting or '',
                'subject_name': s.common_course_structure.course_name,
                'subject_code': s.common_course_structure.code,
            }
            for s in filtered_schedules
        ]

        attendance_data.append({
            'name': student.get_full_name(),
            'roll_no': student.roll_no or 'N/A',
            'registration_no': student.registration_no or 'N/A',
            'photo': photo_base64,
            'college_name': college.name,
            'barcode': barcode_base64,
            'schedules': student_schedules,
        })

    # 6️⃣ Template context
    context = {
        'attendance_data': attendance_data,
        'university_logo': university_logo,
        'exam_header': f"{exam.name} (Semester {exam.semester})",
        'center_name': (
            f"{exam_center.center_code} - {exam_center.name}"
            if exam_center else "Not Assigned"
        ),
    }

    html_string = get_template('mba_sem/attendance_sheet.html').render(context)

    try:
        return HTML(string=html_string, base_url=settings.MEDIA_ROOT).write_pdf()
    except Exception as e:
        logger.error(f"MBA Attendance Sheet PDF error: {e}")
        return None


def generate_mba_student_course_assessment_pdf():
    from weasyprint import HTML, CSS
    import io, base64, os
    from django.conf import settings
    from django.template.loader import get_template
    from mba_sem.models import MBAExamCenterMapping, MBAExamSchedule
    from pup_umis_backend.utils.file_utils import image_to_base64
    import logging

    logger = logging.getLogger(__name__)

    # Get exam center mapping
    
    # Prepare context for template
    context = {
        "university_logo": image_to_base64(os.path.join(settings.MEDIA_ROOT, "common/purnea-logo.png")),
        "watermark_logo": image_to_base64(os.path.join(settings.MEDIA_ROOT, "common/purnea-logo.png")),
        # "student_photo": image_to_base64(student.profile_image.path if student.profile_image else None),
        # "student_sig": image_to_base64(student.signature.path if student.signature else None),
        "controller_signature": image_to_base64(os.path.join(settings.MEDIA_ROOT, "common/controller-of-examination-signature.png")),
    }

    # Render HTML template
    html_string = get_template("mba_sem/mba_result_sheet.html").render(context)

    try:
        # Generate PDF using WeasyPrint
        pdf_file = HTML(string=html_string, base_url=settings.MEDIA_ROOT).write_pdf()
        
        logger.info(f"PDF generated successfully using WeasyPrint, size: {len(pdf_file)} bytes")
        return pdf_file
        
    except Exception as e:
        logger.error(f"PDF generation failed with WeasyPrint: {str(e)}")
        return None

def generate_static_excel_pdf():
    import subprocess
    import os
    import shutil
    from django.conf import settings

    # ---- Excel Absolute Path ----
    excel_path = os.path.join(settings.BASE_DIR, "MBA 2nd Semester Result testing.xlsx")

    if not os.path.exists(excel_path):
        print("Excel file not found:", excel_path)
        return None

    # ---- Ensure LibreOffice Exists ----
    libreoffice_path = shutil.which("libreoffice")
    if not libreoffice_path:
        print("LibreOffice not installed inside environment.")
        return None

    output_dir = "/tmp"

    try:
        result = subprocess.run(
            [
                libreoffice_path,
                "--headless",
                "--convert-to", "pdf",
                excel_path,
                "--outdir", output_dir
            ],
            capture_output=True,
            text=True
        )

        if result.returncode != 0:
            print("LibreOffice Error:", result.stderr)
            return None

    except Exception as e:
        print("Subprocess Exception:", str(e))
        return None

    # ---- Detect Correct PDF Name ----
    base_name = os.path.splitext(os.path.basename(excel_path))[0]
    pdf_path = os.path.join(output_dir, f"{base_name}.pdf")

    if not os.path.exists(pdf_path):
        print("PDF not generated at:", pdf_path)
        return None

    # ---- Read PDF ----
    with open(pdf_path, "rb") as f:
        pdf_bytes = f.read()

    return pdf_bytes

def calculate_numeric_grade(total):
    try:
        total = float(total)
    except:
        return ""

    if total >= 91:
        return 10
    elif total >= 81:
        return 9
    elif total >= 71:
        return 8
    elif total >= 61:
        return 7
    elif total >= 51:
        return 6
    elif total >= 45:
        return 5
    else:
        return 0

def calculate_credit_obtained(ese, cia):
    """
    Returns:
    credit_obtained (int/float)
    """

    def safe(obj, field):
        if obj and hasattr(obj, field):
            val = getattr(obj, field)
            return val if val is not None else 0
        return 0

    # Fetch marks
    ese_marks = safe(ese, "ind_marks_obtained")
    cia_marks = safe(cia, "ind_marks_obtained")

    # Fetch pass marks
    ese_pass = safe(ese, "ind_pass_marks")
    cia_pass = safe(cia, "ind_pass_marks")

    # Fetch max credit
    max_credit = safe(ese or cia, "course_max_credits")

    try:
        ese_marks = float(ese_marks)
        cia_marks = float(cia_marks)
        ese_pass = float(ese_pass)
        cia_pass = float(cia_pass)
    except:
        return 0

    # BOTH PAPER PASS CONDITION
    if ese_marks >= ese_pass and cia_marks >= cia_pass:
        return max_credit
    else:
        return 0

def calculate_grade_point(numeric_grade, credit_obtained):
    """
    Returns:
    grade_point
    """

    try:
        numeric_grade = float(numeric_grade)
        credit_obtained = float(credit_obtained)
    except:
        return 0

    return numeric_grade * credit_obtained

def get_letter_and_description(gpa):

    percent_value = float(gpa) * 10

    if 91 <= percent_value <= 100:
        return "O", "Outstanding"
    elif 81 <= percent_value < 91:
        return "A++", "Excellent"
    elif 71 <= percent_value < 81:
        return "A+", "Very Good"
    elif 61 <= percent_value < 71:
        return "A", "Good"
    elif 51 <= percent_value < 61:
        return "B+", "Average"
    elif 45 <= percent_value < 51:
        return "B", "Pass"
    else:
        return "F", "Fail"

def write_grade_table(ws, start_row=5, start_col=53):

    headers = [
        "Percentage Range",
        "Numerical of Letter Grade",
        "Letter Grade",
        "Description of Grade"
    ]

    # Write headers
    for col, header in enumerate(headers):
        ws.cell(row=start_row, column=start_col + col).value = header

    # Write grade data dynamically
    for i, rule in enumerate(GRADE_STRUCTURE):

        if rule["min"] == 0:
            percentage_text = "<45"
            numeric_text = "<5"
        else:
            percentage_text = f"{rule['min']}-{rule['max']}"
            numeric_text = rule["numeric"]

        ws.cell(row=start_row + 1 + i, column=start_col).value = percentage_text
        ws.cell(row=start_row + 1 + i, column=start_col + 1).value = numeric_text
        ws.cell(row=start_row + 1 + i, column=start_col + 2).value = rule["letter"]
        ws.cell(row=start_row + 1 + i, column=start_col + 3).value = rule["desc"]

def generate_mba_result_pdf(students, college, semester, batch_uid=None):

    import os
    import uuid
    import shutil
    import subprocess
    from collections import defaultdict
    from django.conf import settings
    from openpyxl import load_workbook

    from mba_sem.models import MBAStudentCourseAssessment

    COL_START = 5
    SUBJECT_WIDTH = 7
    DATA_START_ROW = 17
    STUDENTS_PER_PAGE = 5

    def normalize(code):
        return (code or "").replace("-", "").replace(" ", "").upper().strip()

    # ===== TEMPLATE LOAD =====
    template_path = os.path.join(
                        settings.BASE_DIR,
                        "courses_data",
                        "mba",
                        "MBA_Result.xlsx"
                    )
    if not os.path.exists(template_path):
        return None

    temp_excel = os.path.join("/tmp", f"mba_result_{uuid.uuid4().hex}.xlsx")
    shutil.copy(template_path, temp_excel)

    wb = load_workbook(temp_excel)
    master_sheet = wb.active
    master_sheet.title = "MASTER_TEMPLATE"

    students = list(students)
    if not students:
        return None

    # ===== FETCH DATA =====
    qs = MBAStudentCourseAssessment.objects.filter(
        student__in=students,
        semester=semester,
        college_code=college.college_code
    ).select_related("student")

    if batch_uid:
        qs = qs.filter(batch__uid=batch_uid)

    all_assessments = list(qs)

    subject_master = {}
    student_map = defaultdict(list)

    for obj in all_assessments:
        code = normalize(obj.paper_code)
        if not code:
            continue
        subject_master[code] = obj.course_name or code
        student_map[obj.student.id].append(obj)

    SUBJECT_CODES = sorted(subject_master.keys())

    # ===== PAGINATION =====
    def chunked(lst, size):
        for i in range(0, len(lst), size):
            yield lst[i:i + size]

    pages = list(chunked(students, STUDENTS_PER_PAGE))

    # ===== PAGE LOOP =====
    for page_index, student_chunk in enumerate(pages):

        # 1️⃣ FIRST create worksheet
        ws = wb.copy_worksheet(master_sheet)
        ws.title = f"Page_{page_index + 1}"

        # ===== ADD UNIVERSITY LOGO (FIXED) =====
        logo_path = os.path.join(
            settings.MEDIA_ROOT,
            "common/purnea-logo.png"
        )

        if os.path.exists(logo_path):

            img = XLImage(logo_path)

            img.width = 120
            img.height = 120

            # Anchor position — adjust if needed
            img.anchor = "V1"

            ws.add_image(img)

            print("Logo added on sheet:", ws.title)

        else:
            print("Logo file not found:", logo_path)

        # ===== HEADER BUILD =====
        col_pointer = COL_START

        for code in SUBJECT_CODES:

            ws.merge_cells(start_row=5, start_column=col_pointer,
                           end_row=5, end_column=col_pointer + SUBJECT_WIDTH - 1)
            ws.cell(row=5, column=col_pointer).value = subject_master[code]

            ws.merge_cells(start_row=6, start_column=col_pointer,
                           end_row=6, end_column=col_pointer + SUBJECT_WIDTH - 1)
            ws.cell(row=6, column=col_pointer).value = code

            headers = ["End Semester Exam(ESE)", "Continious Internal Assessment(CIA)", "Total",
                       "Credit Alloted", "Numerical Of Letter Grade",
                       "Credit Earned", "GP = C.E. X N.G."]

            for i, title in enumerate(headers):
                ws.cell(row=7, column=col_pointer + i).value = title

            # Static full/pass marks
            ws.cell(row=14, column=col_pointer).value = 70
            ws.cell(row=14, column=col_pointer + 1).value = 30
            ws.cell(row=14, column=col_pointer + 2).value = 100
            ws.cell(row=14, column=col_pointer + 3).value = 4

            ws.cell(row=15, column=col_pointer).value = 31.5
            ws.cell(row=15, column=col_pointer + 1).value = 13.5

            col_pointer += SUBJECT_WIDTH

        # ===== GRADE TABLE =====
        # grade_table_start_col = col_pointer + 3
        # write_grade_table(ws, start_row=5, start_col=grade_table_start_col)

        # ===== FOOTER COUNTERS =====
        page_total = 0
        page_pass = 0
        page_fail = 0
        page_absent = 0
        page_expelled = 0
        page_pending = 0
        page_promoted = 0
        
        # ===== STUDENT LOOP =====
        row_pointer = DATA_START_ROW

        for student in student_chunk:

            ws.cell(row=row_pointer, column=1).value = student.roll_no
            ws.cell(row=row_pointer, column=2).value = student.get_full_name()
            ws.cell(row=row_pointer, column=3).value = student.registration_no

            total_credit_allotted = 0
            total_credit_earned = 0
            total_grade_points = 0
            failed = False
            
            col_pointer = COL_START

            for code in SUBJECT_CODES:

                ese = cia = None

                for rec in student_map.get(student.id, []):
                    if normalize(rec.paper_code) == code:
                        if "ESE" in (rec.label or "").upper():
                            ese = rec
                        if "CIA" in (rec.label or "").upper():
                            cia = rec

                def safe(obj, field):
                    if obj and hasattr(obj, field):
                        val = getattr(obj, field)
                        return val if val is not None else 0
                    return 0

                ese_marks = safe(ese, "ind_marks_obtained")
                cia_marks = safe(cia, "ind_marks_obtained")
                max_credit = safe(ese or cia, "course_max_credits")

                total = float(ese_marks) + float(cia_marks)
                numeric = calculate_numeric_grade(total)
                credit = calculate_credit_obtained(ese, cia)
                gp = calculate_grade_point(numeric, credit)

                total_credit_allotted += float(max_credit)
                total_credit_earned += float(credit)
                total_grade_points += float(gp)

                if credit == 0:
                    failed = True

                values = [
                    ese_marks,
                    cia_marks,
                    total,
                    max_credit,
                    numeric,
                    credit,
                    gp
                ]

                for i in range(SUBJECT_WIDTH):
                    ws.cell(row=row_pointer,
                            column=col_pointer + i).value = values[i]

                col_pointer += SUBJECT_WIDTH

            # ===== SUMMARY =====
            if total_credit_earned > 0:
                raw_gpa = total_grade_points / total_credit_earned
                gpa = round(raw_gpa,2)
            else:
                gpa = 0

            letter, desc = get_letter_and_description(gpa)
            result_status = "Fail" if failed else "Pass"
            # result_status = "Pass"

            page_total += 1

            if result_status == "Pass":
                page_pass += 1
            else:
                page_fail += 1


            ws.cell(row=row_pointer, column=47).value = total_credit_allotted
            ws.cell(row=row_pointer, column=48).value = total_credit_earned
            ws.cell(row=row_pointer, column=49).value = gpa
            ws.cell(row=row_pointer, column=50).value = letter
            ws.cell(row=row_pointer, column=51).value = desc
            ws.cell(row=row_pointer, column=52).value = result_status

            row_pointer += 1
        
        footer_row = row_pointer + 7

        # ===== LEFT BLOCK =====
        ws.cell(row=footer_row, column=2).value = f"No of Candidate : {page_total}"
        ws.cell(row=footer_row + 1, column=2).value = f"Pass : {page_pass}"
        ws.cell(row=footer_row + 2, column=2).value = f"Promoted : {page_promoted}"

        # ===== CENTER BLOCK =====
        ws.cell(row=footer_row, column=5).value = f"Expelled : {page_expelled}"
        ws.cell(row=footer_row + 1, column=5).value = f"Fail : {page_fail}"
        ws.cell(row=footer_row + 2, column=5).value = f"Qualified : {page_pass}"

        # ===== RIGHT BLOCK =====
        ws.cell(row=footer_row + 1, column=12).value = f"Absent : {page_absent}"
        ws.cell(row=footer_row + 2, column=12).value = f"Result Pending : {page_pending}"

        # ===== SIGN AREA =====
        sign_row = footer_row + 2

        ws.cell(row=sign_row, column=22).value = "Compared By Address"
        ws.cell(row=sign_row, column=35).value = "Full Signature of Tabulator-cum-scrutinizer with date"
        ws.cell(row=sign_row, column=50).value = "Controller of Examination"

        wb.remove(master_sheet)
        wb.save(temp_excel)

    subprocess.run([
        "soffice", "--headless",
        "--convert-to", "pdf:calc_pdf_Export",
        "--outdir", "/tmp",
        temp_excel
    ])

    pdf_path = temp_excel.replace(".xlsx", ".pdf")

    if not os.path.exists(pdf_path):
        return None

    with open(pdf_path, "rb") as f:
        pdf_bytes = f.read()

    os.remove(temp_excel)
    os.remove(pdf_path)

    return pdf_bytes

